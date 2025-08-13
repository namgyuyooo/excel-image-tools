#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import os
import locale

# 한글 자소 분리 문제 해결을 위한 인코딩 설정
if sys.platform.startswith('darwin'):  # macOS
    os.environ['LC_ALL'] = 'en_US.UTF-8'
    os.environ['LANG'] = 'en_US.UTF-8'
try:
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_ALL, 'C.UTF-8')
    except:
        pass

import glob
import csv
import unicodedata
from openpyxl import Workbook
from openpyxl.drawing import image as openpyxl_image
from openpyxl.styles import Alignment
from PIL import Image as PILImage
import uuid

def load_inference_results(base_path):
    """Load all inference_results.csv files and create a lookup dictionary"""
    print("=== 추론 결과 파일 검색 시작 ===")
    csv_files = glob.glob(os.path.join(base_path, "**", "inference_results.csv"), recursive=True)
    print(f"발견된 CSV 파일 수: {len(csv_files)}")
    
    results_dict = {}
    
    for i, csv_file in enumerate(csv_files):
        try:
            print(f"[{i+1}/{len(csv_files)}] CSV 파일 로딩 중: {csv_file}")
            row_count = 0
            with open(csv_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    filename = row.get('filename', '').replace('.bmp', '')
                    results_dict[filename] = {
                        'gt_status': row.get('gt_status', ''),
                        'pred_status': row.get('pred_status', ''),
                        'dominant_class': row.get('dominant_class', ''),
                        'csv_source': unicodedata.normalize('NFC', os.path.basename(os.path.dirname(csv_file)))
                    }
                    row_count += 1
            print(f"   └── {row_count}개 결과 로딩 완료")
        except Exception as e:
            print(f"   └── 오류 발생: {e}")
    
    print(f"=== 총 {len(results_dict)}개 파일의 추론 결과 로딩 완료 ===\n")
    return results_dict

def load_dmt_results(dmt_csv_path):
    """Load DMT CSV file and create a lookup dictionary"""
    print("=== DMT 상세결과 파일 로딩 시작 ===")
    dmt_dict = {}
    
    try:
        with open(dmt_csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            row_count = 0
            for row in reader:
                filename = row.get('filename', '').replace('.bmp', '')
                dmt_dict[filename] = {
                    'dmt_category': unicodedata.normalize('NFC', row.get('구분', '')),
                    'gt_status_real': unicodedata.normalize('NFC', row.get('gt_status_real', ''))
                }
                row_count += 1
        print(f"DMT 상세결과 {row_count}개 로딩 완료")
    except Exception as e:
        print(f"DMT 파일 로딩 오류: {e}")
    
    print(f"=== 총 {len(dmt_dict)}개 DMT 결과 로딩 완료 ===\n")
    return dmt_dict

def find_image_pairs(base_path):
    """Find image pairs (viz.png and .bmp files) in the given path"""
    print("=== 이미지 파일 검색 시작 ===")
    print("PNG 파일 (viz) 검색 중...")
    png_files = glob.glob(os.path.join(base_path, "**", "images", "**", "*_viz.png"), recursive=True)
    print(f"발견된 PNG 파일 수: {len(png_files)}")
    
    print("BMP 파일 검색 중...")
    bmp_files = glob.glob(os.path.join(base_path, "**", "images", "**", "*.bmp"), recursive=True)
    print(f"발견된 BMP 파일 수: {len(bmp_files)}")
    
    print("이미지 쌍 매칭 중...")
    pairs = []
    
    for i, png_file in enumerate(png_files):
        if i % 100 == 0:
            print(f"   진행률: {i}/{len(png_files)} PNG 파일 처리 중...")
        
        # Extract base name by removing "_viz.png"
        base_name = os.path.basename(png_file).replace("_viz.png", "")
        
        # Find corresponding .bmp file
        bmp_file = None
        for bmp in bmp_files:
            if os.path.basename(bmp).replace(".bmp", "") == base_name:
                bmp_file = bmp
                break
        
        if bmp_file:
            pairs.append({
                'filename': base_name,
                'viz_img_path': png_file,
                'img_path': bmp_file
            })
    
    print(f"=== 총 {len(pairs)}개 이미지 쌍 매칭 완료 ===\n")
    return pairs

def create_merged_excel(pairs, results_dict, dmt_dict, output_file):
    """Create Excel file with all data merged"""
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged Image Analysis Results"
        
        # Set headers
        ws['A1'] = 'filename'
        ws['B1'] = 'img'
        ws['C1'] = 'viz_img'
        ws['D1'] = 'gt_status'
        ws['E1'] = 'pred_status'
        ws['F1'] = 'dominant_class'
        ws['G1'] = 'csv_source'
        ws['H1'] = 'dmt_category'
        ws['I1'] = 'gt_status_real'
        
        # Set column widths to match image sizes (2 inches = ~19 Excel units)
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 19  # BMP (2 inches width for 1:1 ratio)
        ws.column_dimensions['C'].width = 38  # PNG (4 inches width for 2:1 ratio)  
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        current_row = 2
        temp_files = []
        
        # 모든 고유한 파일명 수집 (이미지 쌍 + DMT 데이터)
        all_filenames = set()
        
        # 이미지 쌍에서 파일명 수집
        for pair in pairs:
            all_filenames.add(pair['filename'])
        
        # DMT 데이터에서 파일명 수집
        for filename in dmt_dict.keys():
            all_filenames.add(filename)
        
        all_filenames = sorted(list(all_filenames))
        
        print("=== 병합된 엑셀 파일 생성 시작 ===")
        print(f"총 처리할 파일 수: {len(all_filenames)}")
        
        # 이미지 쌍을 딕셔너리로 변환 (빠른 검색을 위해)
        pairs_dict = {pair['filename']: pair for pair in pairs}
        
        for i, filename in enumerate(all_filenames):
            try:
                print(f"[{i+1}/{len(all_filenames)}] 처리 중: {filename}")
                
                # Add filename
                normalized_filename = unicodedata.normalize('NFC', filename)
                ws[f'A{current_row}'] = normalized_filename
                
                # 이미지가 있는 경우
                has_images = filename in pairs_dict
                
                if has_images:
                    pair = pairs_dict[filename]
                    
                    # Set row height for images (2 inches = 144 points)
                    ws.row_dimensions[current_row].height = 144
                    
                    # Add BMP image
                    print(f"   └── BMP 이미지 처리 중: {pair['img_path']}")
                    if os.path.exists(pair['img_path']):
                        try:
                            temp_bmp_path = f"temp_bmp_{uuid.uuid4().hex[:8]}.png"
                            temp_files.append(temp_bmp_path)
                            
                            with PILImage.open(pair['img_path']) as pil_img:
                                original_size = pil_img.size
                                if pil_img.mode != 'RGB':
                                    pil_img = pil_img.convert('RGB')
                                # BMP is 1:1 ratio, resize to fill cell height (144 points = 192 pixels)
                                target_size = 144  # 2 inches in pixels at 72 DPI
                                pil_img = pil_img.resize((target_size, target_size), PILImage.Resampling.LANCZOS)
                                pil_img.save(temp_bmp_path, "PNG")
                            
                            img = openpyxl_image.Image(temp_bmp_path)
                            img.width = target_size
                            img.height = target_size
                            img.anchor = f'B{current_row}'
                            ws.add_image(img)
                            
                            print(f"       └── BMP 이미지 삽입 완료 ({original_size[0]}x{original_size[1]} → {target_size}x{target_size})")
                            
                        except Exception as e:
                            ws[f'B{current_row}'] = f"Error: {str(e)}"
                            print(f"       └── BMP 이미지 처리 오류: {e}")
                    else:
                        ws[f'B{current_row}'] = "BMP not found"
                    
                    # Add PNG image
                    print(f"   └── PNG 이미지 처리 중: {pair['viz_img_path']}")
                    if os.path.exists(pair['viz_img_path']):
                        try:
                            temp_png_path = f"temp_png_{uuid.uuid4().hex[:8]}.png"
                            temp_files.append(temp_png_path)
                            
                            with PILImage.open(pair['viz_img_path']) as pil_img:
                                original_size = pil_img.size
                                if pil_img.mode != 'RGB':
                                    pil_img = pil_img.convert('RGB')
                                # PNG is 2:1 ratio, resize to fill cell height (144 points high, 288 wide for 2:1)
                                target_height = 144  # 2 inches
                                target_width = 288   # 4 inches for 2:1 ratio
                                pil_img = pil_img.resize((target_width, target_height), PILImage.Resampling.LANCZOS)
                                pil_img.save(temp_png_path, "PNG")
                            
                            img = openpyxl_image.Image(temp_png_path)
                            img.width = target_width
                            img.height = target_height
                            img.anchor = f'C{current_row}'
                            ws.add_image(img)
                            
                            print(f"       └── PNG 이미지 삽입 완료 ({original_size[0]}x{original_size[1]} → {target_width}x{target_height})")
                            
                        except Exception as e:
                            ws[f'C{current_row}'] = f"Error: {str(e)}"
                            print(f"       └── PNG 이미지 처리 오류: {e}")
                    else:
                        ws[f'C{current_row}'] = "PNG not found"
                else:
                    # 이미지가 없는 경우 빈 셀 처리
                    ws[f'B{current_row}'] = "No image"
                    ws[f'C{current_row}'] = "No image"
                    ws.row_dimensions[current_row].height = 25
                    print(f"   └── 이미지 없음")
                
                # Add inference results
                if filename in results_dict:
                    result = results_dict[filename]
                    ws[f'D{current_row}'] = unicodedata.normalize('NFC', result['gt_status'])
                    ws[f'E{current_row}'] = unicodedata.normalize('NFC', result['pred_status'])
                    ws[f'F{current_row}'] = unicodedata.normalize('NFC', result['dominant_class'])
                    ws[f'G{current_row}'] = unicodedata.normalize('NFC', result['csv_source'])
                    print(f"   └── 추론 결과 추가: GT={result['gt_status']}, Pred={result['pred_status']}")
                else:
                    ws[f'D{current_row}'] = 'Not found'
                    ws[f'E{current_row}'] = 'Not found'
                    ws[f'F{current_row}'] = 'Not found'
                    ws[f'G{current_row}'] = 'Not found'
                
                # Add DMT results
                if filename in dmt_dict:
                    dmt_result = dmt_dict[filename]
                    ws[f'H{current_row}'] = unicodedata.normalize('NFC', dmt_result['dmt_category'])
                    ws[f'I{current_row}'] = unicodedata.normalize('NFC', dmt_result['gt_status_real'])
                    print(f"   └── DMT 결과 추가: Category={dmt_result['dmt_category']}, GT_Real={dmt_result['gt_status_real']}")
                else:
                    ws[f'H{current_row}'] = 'Not found'
                    ws[f'I{current_row}'] = 'Not found'
                
                current_row += 1
                print(f"   └── 행 {current_row-1} 처리 완료\n")
                
            except Exception as e:
                print(f"   └── 파일 처리 오류: {e}\n")
                continue
        
        # Enable AutoFilter
        ws.auto_filter.ref = f"A1:I{current_row-1}"
        print("=== 자동 필터 설정 완료 ===")
        
        # Save workbook
        print("엑셀 파일 저장 중...")
        wb.save(output_file)
        
        # Clean up temporary files
        print("임시 파일 정리 중...")
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except Exception as e:
                print(f"임시 파일 정리 오류 {temp_file}: {e}")
        
        print(f"Excel file created successfully: {output_file}")
        print(f"Total files processed: {len(all_filenames)}")
        
        return True
        
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        # Clean up temporary files even on error
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except:
                pass
        return False

def main():
    base_path = "/Users/rtm/Downloads/v0.3"
    dmt_csv_path = "/Users/rtm/Downloads/v0.3/DMT_poc 상세결과 - 시트14 (1).csv"
    output_file = "/Users/rtm/Downloads/v0.3/merged_analysis_results.xlsx"
    
    print("추론 결과 로딩 중...")
    results_dict = load_inference_results(base_path)
    
    print("DMT 상세결과 로딩 중...")
    dmt_dict = load_dmt_results(dmt_csv_path)
    
    print("이미지 쌍 검색 중...")
    pairs = find_image_pairs(base_path)
    
    print(f"발견된 이미지 쌍: {len(pairs)}개")
    print(f"DMT 결과: {len(dmt_dict)}개")
    
    print("병합된 엑셀 파일 생성 중...")
    success = create_merged_excel(pairs, results_dict, dmt_dict, output_file)
    
    if success:
        print(f"Excel file created: {output_file}")
        print(f"파일이 성공적으로 생성되었습니다: {output_file}")
        print("이제 Excel에서 필터와 정렬 기능을 사용할 수 있습니다!")
    else:
        print("Failed to create Excel file")

if __name__ == "__main__":
    main()