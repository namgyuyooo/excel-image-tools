#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import glob
from openpyxl import Workbook
from openpyxl.drawing import image as openpyxl_image
from PIL import Image as PILImage
import tempfile

def find_image_pairs(base_path):
    """Find image pairs (viz.png and .bmp files) in the given path"""
    png_files = glob.glob(os.path.join(base_path, "**", "images", "**", "*_viz.png"), recursive=True)
    bmp_files = glob.glob(os.path.join(base_path, "**", "images", "**", "*.bmp"), recursive=True)
    
    pairs = []
    
    for png_file in png_files:
        # Extract base name by removing "_viz.png"
        base_name = os.path.basename(png_file).replace("_viz.png", "")
        
        # Find corresponding .bmp file
        bmp_file = None
        for bmp in bmp_files:
            if os.path.basename(bmp).replace(".bmp", "") == base_name:
                bmp_file = bmp
                break
        
        if bmp_file:
            pairs.append({
                'filename': base_name,
                'viz_img_path': png_file,
                'img_path': bmp_file
            })
    
    return pairs

def create_excel_with_images(pairs, output_file):
    """Create Excel file with images"""
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Image Pairs"
        
        # Set headers
        ws['A1'] = 'filename'
        ws['B1'] = 'img'
        ws['C1'] = 'viz_img'
        
        # Set column widths for images
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        current_row = 2
        
        for i, pair in enumerate(pairs):
            if i >= 50:  # Limit to first 50 pairs for performance
                break
                
            try:
                print(f"Processing pair {i+1}/{min(len(pairs), 50)}: {pair['filename']}")
                
                # Add filename
                ws[f'A{current_row}'] = pair['filename']
                
                # Set row height for images (300 pixels = about 225 points)
                ws.row_dimensions[current_row].height = 225
                
                # Add BMP image (img column)
                if os.path.exists(pair['img_path']):
                    try:
                        # Create temporary file for resized BMP image
                        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
                            temp_bmp_path = temp_file.name
                        
                        # Resize BMP image
                        with PILImage.open(pair['img_path']) as pil_img:
                            pil_img.thumbnail((300, 300), PILImage.Resampling.LANCZOS)
                            pil_img.save(temp_bmp_path, "PNG")
                        
                        # Add to Excel
                        img = openpyxl_image.Image(temp_bmp_path)
                        img.anchor = f'B{current_row}'
                        ws.add_image(img)
                        
                        # Clean up temp file
                        os.unlink(temp_bmp_path)
                        
                    except Exception as e:
                        ws[f'B{current_row}'] = f"Error: {str(e)}"
                        print(f"Error processing BMP {pair['img_path']}: {e}")
                else:
                    ws[f'B{current_row}'] = "BMP not found"
                
                # Add PNG image (viz_img column)
                if os.path.exists(pair['viz_img_path']):
                    try:
                        # Create temporary file for resized PNG image
                        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
                            temp_png_path = temp_file.name
                        
                        # Resize PNG image
                        with PILImage.open(pair['viz_img_path']) as pil_img:
                            pil_img.thumbnail((300, 300), PILImage.Resampling.LANCZOS)
                            pil_img.save(temp_png_path, "PNG")
                        
                        # Add to Excel
                        img = openpyxl_image.Image(temp_png_path)
                        img.anchor = f'C{current_row}'
                        ws.add_image(img)
                        
                        # Clean up temp file
                        os.unlink(temp_png_path)
                        
                    except Exception as e:
                        ws[f'C{current_row}'] = f"Error: {str(e)}"
                        print(f"Error processing PNG {pair['viz_img_path']}: {e}")
                else:
                    ws[f'C{current_row}'] = "PNG not found"
                
                current_row += 1
                
            except Exception as e:
                print(f"Error processing pair {pair['filename']}: {e}")
                continue
        
        # Save workbook
        wb.save(output_file)
        print(f"Excel file created successfully: {output_file}")
        print(f"Total pairs processed: {min(len(pairs), 50)}")
        
        return True
        
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return False

def main():
    base_path = "/Users/rtm/Downloads/v0.3"
    output_file = "/Users/rtm/Downloads/v0.3/image_pairs.xlsx"
    
    print("Searching for image pairs...")
    pairs = find_image_pairs(base_path)
    
    print(f"Found {len(pairs)} image pairs")
    
    if pairs:
        print("Creating Excel file with images...")
        success = create_excel_with_images(pairs, output_file)
        
        if success:
            print(f"Excel file created: {output_file}")
        else:
            print("Failed to create Excel file")
    else:
        print("No image pairs found")

if __name__ == "__main__":
    main()