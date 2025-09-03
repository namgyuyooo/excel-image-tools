#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Create an Excel file from segmentation inference CSV and images.

Inputs:
- IMAGES_BASE: Base directory that contains image files organized by class folders
- CSV_PATH: CSV file with headers: img_path, origin_class, error, pred_seg_results, seg_score

Output:
- Excel file containing rows per image with embedded thumbnails and parsed predictions
"""

import os
import sys
import csv
import glob
import uuid
from typing import Dict, List, Tuple, Optional

from openpyxl import Workbook
from openpyxl.drawing import image as openpyxl_image
from PIL import Image as PILImage
from openpyxl.utils import get_column_letter


def normalize_relative_path(path_from_csv: str) -> str:
    """Normalize a CSV-provided image path to be relative to images base.

    Examples:
    - "20250801_v0.2/1.bondfinger/0001.jpg" -> "1.bondfinger/0001.jpg"
    - "1.bondfinger/0001.jpg" -> unchanged
    - Handles backslashes and redundant separators
    """
    if not path_from_csv:
        return ""
    norm = path_from_csv.replace("\\", "/")
    # If the first segment looks like a dataset date/version, drop it
    if "/" in norm:
        first, rest = norm.split("/", 1)
        # Heuristic: drop a leading segment that contains digits and dots/underscores like a version string
        if any(ch.isdigit() for ch in first):
            return rest
    return norm


def resolve_image_path(images_base: str, csv_img_path: str) -> Optional[str]:
    """Resolve CSV image path to an absolute path under images_base.

    Enhanced Strategy for new /1/ based folder structure:
    - 새로운 기본 경로: /Users/yunamgyu/Downloads/test/1/0001/Unit/U0/BC
    - CSV File_path에서 /1/ 이하 상대 경로 추출 및 매칭
    - /1/ 구조를 고려한 다중 경로 검색

    Search patterns:
    1) /1/ 이하 상대 경로 추출 및 결합
    2) /img 폴더에서도 검색 (실제 파일 위치)
    3) Unit 폴더 간 매칭 (U0, U9, U12 등)
    4) 확장자 변환 및 재귀 검색
    """
    if not csv_img_path:
        return None

    # 절대 경로로 존재하는지 확인
    if os.path.isabs(csv_img_path) and os.path.exists(csv_img_path):
        return csv_img_path

    print(f"🔍 패턴 기반 초유연 매칭 시작: {csv_img_path}")

    # 0. Windows 백슬래시를 Unix 슬래시로 변환 (새로운 패턴 지원)
    normalized_csv_path = csv_img_path.replace('\\', '/')
    if normalized_csv_path != csv_img_path:
        print(f"🔄 Windows 경로 변환: {csv_img_path} -> {normalized_csv_path}")
        csv_img_path = normalized_csv_path

    # 1. CSV 경로 구조 분석 및 패턴 추출
    import re

    # /숫자/ 패턴 추출 (개선된 정규식)
    number_pattern = re.search(r'/(\d+)/', csv_img_path)
    csv_structure = {}

    if number_pattern:
        csv_number = number_pattern.group(1)
        csv_structure['number'] = csv_number

        # /숫자/ 이후 경로 분석
        after_number = csv_img_path.split(f'/{csv_number}/', 1)[1]
        path_parts = after_number.split('/')

        if len(path_parts) >= 4:
            csv_structure.update({
                'part1': path_parts[0],  # 0001 (4자리 숫자)
                'part2': path_parts[1],  # Unit (고정)
                'part3': path_parts[2],  # U12, U70 등 (U+숫자)
                'part4': path_parts[3],  # BC, FC 등 (타입)
                'filename': path_parts[-1] if len(path_parts) > 4 else path_parts[-1]
            })

        print(f"📊 CSV 구조 분석: {csv_structure}")
        print(f"🎯 매칭 패턴: {csv_number}/{csv_structure.get('part1', '*')}/{csv_structure.get('part2', '*')}/{csv_structure.get('part3', 'U*')}/{csv_structure.get('filename', '*.jpg')}")

    # 2. 기본 경로 구조 분석
    base_structure = {}
    base_number_pattern = re.search(r'/test/(\d+)/', images_base)
    if base_number_pattern:
        base_number = base_number_pattern.group(1)
        base_structure['number'] = base_number

        # 기본 경로의 나머지 부분 분석
        after_base_number = images_base.split(f'/test/{base_number}/', 1)[1]
        base_path_parts = after_base_number.split('/')

        if len(base_path_parts) >= 4:
            base_structure.update({
                'part1': base_path_parts[0],
                'part2': base_path_parts[1],
                'part3': base_path_parts[2],
                'part4': base_path_parts[3],
            })

        print(f"📊 기본 경로 구조: {base_structure}")

    # 3. 패턴 기반 최적화 매칭
    search_patterns = []
    basename = os.path.basename(csv_img_path)
    base_no_ext, original_ext = os.path.splitext(basename)

    # 우선순위 1: 정확한 구조 매칭 (CSV 구조를 기본 경로에 적용)
    if csv_structure and base_structure:
        target_number = base_structure.get('number', csv_structure.get('number', '1'))

        # U* 패턴 우선: U로 시작하는 폴더만 검색 (최적화)
        u_folders = []
        # 기존 unit_folders + 동적 생성
        base_u_folders = ['U0', 'U1', 'U2', 'U6', 'U7', 'U8', 'U9', 'U10', 'U11', 'U12', 'U13', 'U14', 'U15', 'U16', 'U19']

        # CSV의 U+숫자 패턴이 있다면 우선 추가
        if csv_structure.get('part3', '').startswith('U'):
            csv_u_folder = csv_structure['part3']
            if csv_u_folder not in base_u_folders:
                base_u_folders.insert(0, csv_u_folder)  # 우선순위 높임

        u_folders = base_u_folders
        type_folders = ['BC', 'FC', 'DC']

        # 4자리 숫자 우선 매칭 (0001, 0002 등)
        part1_candidates = []
        if len(csv_structure.get('part1', '')) == 4 and csv_structure['part1'].isdigit():
            # CSV의 4자리 숫자를 우선 사용
            part1_candidates.append(csv_structure['part1'])
        part1_candidates.extend([base_structure.get('part1', '0001'), '0001', '0002', '0003'])

        for part1 in part1_candidates[:3]:  # 상위 3개만
            for unit in u_folders[:10]:  # 상위 10개 U 폴더만 (성능 최적화)
                for type_folder in type_folders:
                    # 구조 기반 경로 생성
                    struct_path = f"{target_number}/{part1}/{base_structure.get('part2', 'Unit')}/{unit}/{type_folder}/{basename}"
                    search_patterns.append(os.path.join(images_base.split(f'/test/{base_structure.get("number", "1")}/')[0], "test", struct_path))

                    # /img 폴더에서도 검색
                    img_base = images_base.replace('/test/', '/test/img/')
                    if img_base != images_base:
                        search_patterns.append(os.path.join(img_base.split(f'/test/img/{base_structure.get("number", "1")}/')[0], "test", "img", struct_path))

    # 우선순위 2: p* 패턴 우선 검색 (p로 시작하는 파일만)
    if basename.startswith('p') or base_no_ext.startswith('p'):
        # p로 시작하는 파일 패턴들
        p_patterns = [
            os.path.join(images_base, "**", f"p*.{original_ext[1:]}" if original_ext else "p*"),
            os.path.join(images_base, "**", f"p*.jpg"),
            os.path.join(images_base, "**", f"p*.png"),
            os.path.join(images_base, "**", f"p*.bmp"),
        ]
        search_patterns.extend(p_patterns)

    # 우선순위 3: 기본 파일명 패턴들
    search_patterns.extend([
        os.path.join(images_base, "**", basename),
        os.path.join(images_base, "**", f"{base_no_ext}.*"),
    ])

    # 4. 확장자 변환 패턴 추가 (우선순위 4)
    ext_mapping = {
        '.bmp': ['.jpg', '.jpeg', '.png'],
        '.png': ['.jpg', '.jpeg', '.bmp'],
        '.jpg': ['.jpeg', '.png', '.bmp'],
        '.jpeg': ['.jpg', '.png', '.bmp']
    }

    if original_ext.lower() in ext_mapping:
        for new_ext in ext_mapping[original_ext.lower()]:
            search_patterns.extend([
                os.path.join(images_base, "**", f"{base_no_ext}{new_ext}"),
                os.path.join(images_base.replace('/test/', '/test/img/'), "**", f"{base_no_ext}{new_ext}")
            ])

    # 5. 중복 제거 및 우선순위 정렬
    seen_patterns = set()
    unique_patterns = []
    for pattern in search_patterns:
        if pattern not in seen_patterns:
            seen_patterns.add(pattern)
            unique_patterns.append(pattern)

    print(f"🔍 생성된 검색 패턴 수: {len(unique_patterns)} (우선순위 최적화)")

    # 6. 실제 검색 수행 (우선순위 기반)
    all_matches = []

    # 우선순위 그룹별 검색
    priority_groups = [
        unique_patterns[:10],    # 구조 매칭 우선 (상위 10개)
        unique_patterns[10:30],  # p* 패턴 (다음 20개)
        unique_patterns[30:50],  # 기본 패턴 (다음 20개)
        unique_patterns[50:70],  # 확장자 변환 (다음 20개)
    ]

    for group_idx, group in enumerate(priority_groups):
        if not group:
            continue

        print(f"🎯 우선순위 그룹 {group_idx + 1} 검색 중...")
        for pattern in group:
            try:
                if os.path.exists(pattern):
                    print(f"✅ 우선순위 {group_idx + 1} - 직접 경로 매치: {pattern}")
                    return pattern

                matches = glob.glob(pattern, recursive=True)
                if matches:
                    for match in matches:
                        match_basename = os.path.basename(match)
                        if match_basename == basename:
                            print(f"✅ 우선순위 {group_idx + 1} - 정확한 파일명 매치: {match}")
                            return match
                        all_matches.append(match)
            except:
                continue

    # 7. 매치 결과 반환
    if all_matches:
        best_match = all_matches[0]
        print(f"✅ 구조 기반 매칭 성공: {best_match}")
        return best_match

    print(f"❌ 모든 패턴으로 검색했으나 파일을 찾을 수 없음: {basename}")
    return None


def find_viz_image(original_image_path: str) -> Optional[str]:
    """Try to find a corresponding visualization image near the original image.

    Looks for variants like:
    - <name>_viz.png
    - <name>_viz.jpg/jpeg
    - any *<name>*viz*.png in the same directory
    """
    if not original_image_path:
        return None
    directory = os.path.dirname(original_image_path)
    base, _ = os.path.splitext(os.path.basename(original_image_path))

    candidates = [
        os.path.join(directory, f"{base}_viz.png"),
        os.path.join(directory, f"{base}_viz.jpg"),
        os.path.join(directory, f"{base}_viz.jpeg"),
    ]
    for cand in candidates:
        if os.path.exists(cand):
            return cand

    wildcard = glob.glob(os.path.join(directory, f"*{base}*viz*.png"))
    if wildcard:
        return wildcard[0]

    return None


def parse_prediction_fields(pred_seg_results: str, seg_score: str) -> Tuple[str, str, str, str]:
    """Parse semicolon-separated predictions and scores.

    Returns: (top_pred, top_score, all_preds, all_scores)
    """
    preds = [p.strip() for p in (pred_seg_results or "").split(";") if p.strip()]
    scores = [s.strip() for s in (seg_score or "").split(";") if s.strip()]

    top_pred = preds[0] if preds else ""
    top_score = scores[0] if scores else ""
    all_preds = "; ".join(preds)
    all_scores = "; ".join(scores)
    return top_pred, top_score, all_preds, all_scores


def create_excel_from_csv(images_base: str, csv_path: str, output_file: str, limit: Optional[int] = None) -> bool:
    """Create an Excel workbook that mirrors the CSV columns and appends one last image column.

    limit: cap the number of rows processed (None for all)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "inference_results"

    temp_files: List[str] = []

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        fieldnames: List[str] = reader.fieldnames or []

        # Headers = CSV headers + last image column
        headers = fieldnames + ["img"]
        for col_index, name in enumerate(headers, start=1):
            ws.cell(row=1, column=col_index, value=name)

        # Make the last column wide for images
        last_col_letter = get_column_letter(len(headers))
        ws.column_dimensions[last_col_letter].width = 30

        current_row = 2
        for i, row in enumerate(reader):
            if limit is not None and i >= limit:
                break

            # Write CSV fields as-is in order
            for col_index, key in enumerate(fieldnames, start=1):
                ws.cell(row=current_row, column=col_index, value=row.get(key, ""))

            # Resolve and insert image at the last column
            csv_img_path = row.get('img_path', '')
            abs_img_path = resolve_image_path(images_base, csv_img_path)

            # Set row height for image visibility
            ws.row_dimensions[current_row].height = 180

            image_col_letter = last_col_letter
            if abs_img_path and os.path.exists(abs_img_path):
                try:
                    temp_img_path = f"temp_img_{uuid.uuid4().hex[:8]}.png"
                    with PILImage.open(abs_img_path) as pil_img:
                        if pil_img.mode != 'RGB':
                            pil_img = pil_img.convert('RGB')
                        pil_img.thumbnail((300, 300), PILImage.Resampling.LANCZOS)
                        pil_img.save(temp_img_path, "PNG")
                    temp_files.append(temp_img_path)

                    img = openpyxl_image.Image(temp_img_path)
                    img.width = 180
                    img.height = 180
                    img.anchor = f'{image_col_letter}{current_row}'
                    ws.add_image(img)
                except Exception as e:
                    ws[f'{image_col_letter}{current_row}'] = f"Error: {e}"
            else:
                ws[f'{image_col_letter}{current_row}'] = "Image not found"

            current_row += 1

    wb.save(output_file)

    for path in temp_files:
        try:
            if os.path.exists(path):
                os.unlink(path)
        except Exception:
            pass

    return True


def main():
    # Configure your paths here (can be turned into CLI args if needed)
    IMAGES_BASE = \
        "/Users/rtm/Downloads/seg/v0.3_inference_20250801_v0.2/images"
    CSV_PATH = \
        "/Users/rtm/Downloads/seg/v0.3_inference_20250801_v0.2/inference_results.csv"

    output_file = \
        "/Users/rtm/Downloads/seg/v0.3_inference_20250801_v0.2/inference_results_with_image.xlsx"

    # Safety checks
    if not os.path.isdir(IMAGES_BASE):
        print(f"❌ Images base directory not found: {IMAGES_BASE}")
        sys.exit(1)
    if not os.path.isfile(CSV_PATH):
        print(f"❌ CSV file not found: {CSV_PATH}")
        sys.exit(1)

    print("📁 Images base:", IMAGES_BASE)
    print("📄 CSV path:", CSV_PATH)
    print("📊 Output:", output_file)

    success = create_excel_from_csv(IMAGES_BASE, CSV_PATH, output_file, limit=None)
    if success:
        print(f"✅ Excel created: {output_file}")
    else:
        print("❌ Failed to create Excel")
        sys.exit(1)


if __name__ == "__main__":
    main()


