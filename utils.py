#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import json
import hashlib
import re
from datetime import datetime
try:
    # Python 3.11+
    from datetime import UTC as _UTC
except Exception:
    from datetime import timezone as _tz
    _UTC = _tz.utc
from typing import Dict, List, Optional, Tuple
import gc
import psutil

import pandas as pd
from PySide6 import QtCore, QtGui, QtWidgets

from openpyxl import load_workbook

# Reuse path resolution from the existing module
from create_excel_from_seg_csv import resolve_image_path

    # CSV 타입별 경로 설정
CSV_CONFIGS = {
    "inference": {
        "csv_path": "/Users/yunamgyu/Downloads/v0.5/v0.5_inference_20250818_v0.2/inference_results.csv",
        "images_base": "/Users/yunamgyu/Downloads/v0.5/v0.5_inference_20250818_v0.2/images",
        "json_base": "/Users/yunamgyu/Downloads/v0.5/v0.5_inference_20250818_v0.2/result"  # JSON 파일이 있는 기본 경로
    },
    "report": {
        "csv_path": "/Users/yunamgyu/Downloads/report/2025-07-31.csv",
        "images_base": "/Users/yunamgyu/Downloads/report/images",  # report 폴더에 images 하위폴더가 있다고 가정
        "json_base": "/Users/yunamgyu/Downloads/report 2/result"  # JSON 파일이 있는 기본 경로
    }
}

def detect_csv_type(csv_path: str) -> str:
    """CSV 파일 경로를 기반으로 타입을 감지합니다."""
    if "inference" in csv_path.lower():
        return "inference"
    elif "report" in csv_path.lower():
        return "report"
    else:
        # 기본값으로 inference 사용
        return "inference"

def get_csv_config(csv_path: str) -> dict:
    """CSV 파일 경로에 맞는 설정을 반환합니다."""
    csv_type = detect_csv_type(csv_path)
    return CSV_CONFIGS.get(csv_type, CSV_CONFIGS["inference"])

# Memory management utilities
def get_memory_usage():
    """Get current memory usage in MB"""
    try:
        process = psutil.Process(os.getpid())
        return process.memory_info().rss / 1024 / 1024
    except:
        return 0

def check_memory_limit(limit_mb=1024):
    """Check if memory usage exceeds limit"""
    return get_memory_usage() > limit_mb

def force_garbage_collection():
    """Force garbage collection to free memory"""
    gc.collect()

def get_system_memory():
    """Get total system memory in MB"""
    try:
        return psutil.virtual_memory().total / 1024 / 1024
    except:
        return 8192  # Default to 8GB if can't detect


def parse_pred_list(value) -> List[str]:
    """Parse Unique_seg_result value into a list of strings.
    Handles JSON arrays, python-like list strings, or comma-separated strings.
    """
    try:
        if isinstance(value, (list, tuple, set)):
            return [str(x).strip() for x in value]
        s = str(value).strip()
        if not s:
            return []
        # Try JSON first
        if s.startswith("[") or s.startswith("{"):
            try:
                data = json.loads(s)
                if isinstance(data, (list, tuple, set)):
                    return [str(x).strip() for x in data]
            except Exception:
                pass
        # Fallback: strip brackets and split by semicolon/comma variants
        s2 = s.strip().strip('[](){}')
        parts = [p.strip().strip("'\"") for p in re.split(r"[;,\uFF1B\uFF0C]+", s2) if p.strip()]
        return parts
    except Exception:
        return []


def extract_detail_from_json(json_path: str) -> List[str]:
    """JSON 파일에서 detail 정보를 추출합니다."""
    if not json_path or not os.path.exists(json_path):
        return []

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        details = []

        # JSON 구조에 따라 detail 정보 추출
        if isinstance(data, dict):
            # 어노테이션 정보에서 detail 추출
            if 'annotations' in data and isinstance(data['annotations'], list):
                for ann in data['annotations']:
                    if isinstance(ann, dict):
                        label = ann.get('label', '')
                        score = ann.get('score', 0.0)
                        bbox = ann.get('bbox', [])
                        if label:
                            detail = f"{label} (신뢰도: {score:.3f})"
                            if bbox and len(bbox) == 4:
                                detail += f" 위치: [{bbox[0]}, {bbox[1]}, {bbox[2]}, {bbox[3]}]"
                            details.append(detail)

            # 기존 detail 키가 있는 경우
            if 'detail' in data:
                detail_data = data['detail']
                if isinstance(detail_data, list):
                    details.extend([str(item) for item in detail_data])
                elif isinstance(detail_data, str):
                    details.append(detail_data)
                elif isinstance(detail_data, dict):
                    # detail이 dict인 경우 모든 값 추출
                    for key, value in detail_data.items():
                        details.append(f"{key}: {value}")

            # 다른 가능한 구조들
            elif 'details' in data:
                detail_data = data['details']
                if isinstance(detail_data, list):
                    details.extend([str(item) for item in detail_data])
                elif isinstance(detail_data, str):
                    details.append(detail_data)

            # 전체 데이터에서 특정 패턴 찾기
            else:
                # defects, issues 등의 키 탐색
                for key in ['defects', 'issues', 'problems', 'anomalies']:
                    if key in data:
                        items = data[key]
                        if isinstance(items, list):
                            details.extend([str(item) for item in items])
                        break

        elif isinstance(data, list):
            # JSON이 리스트인 경우
            details.extend([str(item) for item in data])

        return details

    except Exception as e:
        print(f"JSON 파일 파싱 오류 ({json_path}): {e}")
        return []


def ensure_object_dtype(df: pd.DataFrame, column: str) -> None:
    try:
        df[column] = df[column].astype("object")
    except Exception:
        pass


def default_json_path(xlsx_path: str) -> str:
    base = os.path.basename(xlsx_path)
    root, _ = os.path.splitext(base)
    parent = os.path.dirname(xlsx_path) or os.getcwd()
    # Preferred naming: if Excel is *_labeled.xlsx → JSON is *_labels.json
    if root.endswith("_labeled"):
        new_root = root[: -len("_labeled")] + "_labels"
    else:
        new_root = root + "_labels"
    new_path = os.path.join(parent, f"{new_root}.json")
    # Legacy path compatibility: previously always appended _labels
    legacy_path = os.path.join(parent, f"{root}_labels.json")
    # If legacy exists and new doesn't, keep using legacy for seamless migration
    if os.path.exists(legacy_path) and not os.path.exists(new_path):
        return legacy_path
    return new_path


def load_label_store(json_path: str) -> dict:
    if not json_path or not os.path.exists(json_path):
        return {"version": 1, "updated_at": None, "labels": {}}
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, dict) and "labels" in data:
                return data
    except Exception:
        pass
    return {"version": 1, "updated_at": None, "labels": {}}


def save_label_store(json_path: str, store: dict) -> bool:
    """Atomically persist the label store. Returns True on success."""
    try:
        store["updated_at"] = datetime.now(_UTC).isoformat()
        tmp = json_path + ".tmp"
        os.makedirs(os.path.dirname(json_path) or ".", exist_ok=True)
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(store, f, ensure_ascii=False, indent=2)
        os.replace(tmp, json_path)
        return True
    except Exception:
        return False


def apply_json_to_excel(json_path: str, xlsx_path: str, sheet_name: str, col_indices: Dict[str, int], df: pd.DataFrame) -> int:
    store = load_label_store(json_path)
    labels = store.get("labels", {})
    applied = 0
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    for key, entry in labels.items():
        try:
            row_idx = int(key)
        except Exception:
            continue
        excel_row = row_idx + 2
        values = entry.get("values", {})
        for col_name, val in values.items():
            idx = col_indices.get(col_name)
            if idx is None:
                continue
            try:
                ws.cell(row=excel_row, column=idx, value=val)
                applied += 1
                if col_name in df.columns:
                    df.at[row_idx, col_name] = val
            except Exception:
                pass
    wb.save(xlsx_path)
    wb.close()
    return applied


def get_json_entry(json_path: str, row_idx: int) -> dict:
    store = load_label_store(json_path)
    key = str(row_idx)
    return store.get("labels", {}).get(key) or {}


def upsert_json_entry(json_path: str, row_idx: int, updater: Dict[str, object]) -> None:
    store = load_label_store(json_path)
    key = str(row_idx)
    entry = store["labels"].get(key) or {}
    for k, v in updater.items():
        entry[k] = v
    store["labels"][key] = entry
    save_label_store(json_path, store)


def merge_json_into_df(json_path: str, df: pd.DataFrame, label_columns: List[str]) -> None:
    """Load existing JSON labels and reflect into DataFrame so work can resume after restart."""
    store = load_label_store(json_path)
    labels = store.get("labels", {})
    for key, entry in labels.items():
        try:
            ridx = int(key)
        except Exception:
            continue
        if ridx not in df.index:
            continue
        values = entry.get("values", {})
        for col, val in values.items():
            if col in label_columns:
                try:
                    curr = None
                    try:
                        curr = df.at[ridx, col]
                    except Exception:
                        curr = None
                    is_empty = (curr is None) or (str(curr) == "")
                    try:
                        if not is_empty:
                            is_empty = bool(pd.isna(curr))
                    except Exception:
                        pass
                    if is_empty:
                        df.at[ridx, col] = val
                except Exception:
                    pass

def is_xlsx(path: str) -> bool:
    try:
        return os.path.isfile(path) and path.lower().endswith(".xlsx")
    except Exception:
        return False


def thumb_cache_path(images_base: str, resolved_path: str, target_edge: int) -> str:
    rel = os.path.relpath(resolved_path, images_base)
    key = hashlib.md5(f"{rel}|{target_edge}".encode("utf-8")).hexdigest()
    cache_dir = os.path.join(images_base, ".thumb_cache")
    os.makedirs(cache_dir, exist_ok=True)
    return os.path.join(cache_dir, f"{key}.png")


def build_thumb_if_needed(images_base: str, resolved_path: str, target_edge: int) -> str:
    thumb = thumb_cache_path(images_base, resolved_path, target_edge)
    try:
        src_mtime = os.path.getmtime(resolved_path)
        if os.path.exists(thumb) and os.path.getmtime(thumb) >= src_mtime:
            return thumb
        # Use Qt to scale and save
        img = QtGui.QImage(resolved_path)
        if img.isNull():
            return resolved_path
        w, h = img.width(), img.height()
        scale = target_edge / float(max(w, h))
        if scale < 1.0:
            img = img.scaled(int(w * scale), int(h * scale), QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        img.save(thumb, "PNG")
        return thumb
    except Exception:
        return resolved_path
