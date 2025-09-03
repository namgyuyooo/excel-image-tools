#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import json
import hashlib
import re
from datetime import datetime
try:
    # Python 3.11+
    from datetime import UTC as _UTC
except Exception:
    from datetime import timezone as _tz
    _UTC = _tz.utc
from typing import Dict, List, Optional, Tuple
import gc
import psutil

import pandas as pd
from PySide6 import QtCore, QtGui, QtWidgets

from openpyxl import load_workbook

# Reuse path resolution from the existing module
from create_excel_from_seg_csv import resolve_image_path

    # CSV 타입별 기본 템플릿 (동적 경로 생성용)
CSV_CONFIGS = {
    "report": {
        "csv_path": "",  # 동적으로 설정됨
        "images_base": "",  # 동적으로 설정됨
        "json_base": ""  # 동적으로 설정됨
    }
}

def auto_detect_paths(csv_path: str) -> dict:
    """CSV 파일 경로를 기반으로 이미지와 JSON 경로를 자동 탐색합니다."""
    import os
    import glob

    csv_dir = os.path.dirname(csv_path)
    csv_name = os.path.basename(csv_path)
    csv_base = os.path.splitext(csv_name)[0]

    print(f"🔍 자동 경로 탐색 시작: {csv_path}")
    print(f"📁 CSV 디렉토리: {csv_dir}")

    # 가능한 이미지 경로 후보들
    image_candidates = [
        os.path.join(csv_dir, "images"),
        os.path.join(csv_dir, "img"),
        os.path.join(csv_dir, "1"),  # /1/ 구조
        os.path.join(csv_dir, "1", "0001"),  # /1/0001/ 구조
        os.path.join(csv_dir, "1", "0001", "Unit"),  # /1/0001/Unit/ 구조
        os.path.join(csv_dir, "Unit"),  # Unit 폴더
        csv_dir,  # CSV 파일과 같은 디렉토리
    ]

    # 가능한 JSON 경로 후보들
    json_candidates = [
        os.path.join(csv_dir, "json"),
        os.path.join(csv_dir, "result"),
        os.path.join(csv_dir, "Unit"),  # Unit 폴더
        os.path.join(csv_dir, "1"),  # /1/ 구조
        os.path.join(csv_dir, "1", "0001"),  # /1/0001/ 구조
        csv_dir,  # CSV 파일과 같은 디렉토리
    ]

    # 이미지 경로 탐색
    images_base = None
    for candidate in image_candidates:
        if os.path.exists(candidate):
            # 이미지 파일이 있는지 확인
            image_files = glob.glob(os.path.join(candidate, "**", "*.jpg"), recursive=True)
            image_files.extend(glob.glob(os.path.join(candidate, "**", "*.png"), recursive=True))
            image_files.extend(glob.glob(os.path.join(candidate, "**", "*.bmp"), recursive=True))

            if image_files:
                images_base = candidate
                print(f"✅ 이미지 경로 발견: {candidate} ({len(image_files)}개 파일)")
                break

    # JSON 경로 탐색
    json_base = None
    for candidate in json_candidates:
        if os.path.exists(candidate):
            # JSON 파일이 있는지 확인
            json_files = glob.glob(os.path.join(candidate, "**", "*.json"), recursive=True)

            if json_files:
                json_base = candidate
                print(f"✅ JSON 경로 발견: {candidate} ({len(json_files)}개 파일)")
                break

    # 기본값 설정 (발견되지 않은 경우)
    if not images_base:
        images_base = csv_dir
        print(f"⚠️ 이미지 경로를 찾을 수 없어 기본값 사용: {images_base}")

    if not json_base:
        json_base = csv_dir
        print(f"⚠️ JSON 경로를 찾을 수 없어 기본값 사용: {json_base}")

    return {
        "csv_path": csv_path,
        "images_base": images_base,
        "json_base": json_base
    }

def detect_csv_type(csv_path: str) -> str:
    """CSV 파일 경로를 기반으로 타입을 감지합니다."""
    # 모든 CSV 파일을 report 타입으로 처리
    return "report"

def get_csv_config(csv_path: str) -> dict:
    """CSV 파일 경로에 맞는 설정을 반환합니다."""
    csv_type = detect_csv_type(csv_path)

    # 동적 경로 탐색을 통해 설정 생성
    if csv_path and os.path.exists(csv_path):
        detected_config = auto_detect_paths(csv_path)
        print(f"🎯 최종 설정: {detected_config}")
        return detected_config
    else:
        # 기본 설정 반환
        return CSV_CONFIGS.get(csv_type, CSV_CONFIGS["report"])

# Memory management utilities
def get_memory_usage():
    """Get current memory usage in MB"""
    try:
        process = psutil.Process(os.getpid())
        return process.memory_info().rss / 1024 / 1024
    except:
        return 0

def check_memory_limit(limit_mb=1024):
    """Check if memory usage exceeds limit"""
    return get_memory_usage() > limit_mb

def force_garbage_collection():
    """Force garbage collection to free memory"""
    gc.collect()

def get_system_memory():
    """Get total system memory in MB"""
    try:
        return psutil.virtual_memory().total / 1024 / 1024
    except:
        return 8192  # Default to 8GB if can't detect


def parse_pred_list(value) -> List[str]:
    """Parse Unique_seg_result value into a list of strings.
    Handles JSON arrays, python-like list strings, or comma-separated strings.
    """
    try:
        if isinstance(value, (list, tuple, set)):
            return [str(x).strip() for x in value]
        s = str(value).strip()
        if not s:
            return []
        # Try JSON first
        if s.startswith("[") or s.startswith("{"):
            try:
                data = json.loads(s)
                if isinstance(data, (list, tuple, set)):
                    return [str(x).strip() for x in data]
            except Exception:
                pass
        # Fallback: strip brackets and split by semicolon/comma variants
        s2 = s.strip().strip('[](){}')
        parts = [p.strip().strip("'\"") for p in re.split(r"[;,\uFF1B\uFF0C]+", s2) if p.strip()]
        return parts
    except Exception:
        return []





def extract_bbox_from_json(json_path: str) -> List[dict]:
    """JSON 파일에서 bbox 정보를 추출합니다."""
    if not json_path or not os.path.exists(json_path):
        return []

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        bboxes = []

        # JSON이 리스트 형태인 경우 (주어진 예시처럼)
        if isinstance(data, list):
            for item in data:
                if isinstance(item, dict) and 'bbox' in item:
                    bbox = item['bbox']
                    # 리스트 형태에서는 imageWidth/Height 정보가 없으므로 기본값 사용
                    processed_bbox = process_bbox_coordinates(bbox)

                    bbox_info = {
                        'bbox': processed_bbox,  # 변환된 [x1, y1, x2, y2]
                        'label': item.get('label', 'Unknown'),
                        'score': item.get('score', 0.0),
                        'type': item.get('type', 'unknown'),
                        'original_bbox': bbox,  # 원본 bbox도 저장
                        'json_img_width': None,
                        'json_img_height': None
                    }
                    bboxes.append(bbox_info)

        # JSON이 딕셔너리 형태인 경우
        elif isinstance(data, dict):
            # JSON에서 원본 이미지 크기 정보 추출
            json_img_width = data.get('imageWidth', None)
            json_img_height = data.get('imageHeight', None)
            
            # 어노테이션 정보에서 bbox 추출
            if 'annotations' in data and isinstance(data['annotations'], list):
                for ann in data['annotations']:
                    if isinstance(ann, dict) and 'bbox' in ann:
                        bbox = ann['bbox']
                        # bbox 좌표 변환 (이미지 크기 정보 포함)
                        processed_bbox = process_bbox_coordinates(bbox, json_img_width, json_img_height)

                        bbox_info = {
                            'bbox': processed_bbox,
                            'label': ann.get('label', 'Unknown'),
                            'score': ann.get('score', 0.0),
                            'type': ann.get('type', 'unknown'),
                            'original_bbox': bbox,  # 원본 bbox도 저장
                            'json_img_width': json_img_width,
                            'json_img_height': json_img_height
                        }
                        bboxes.append(bbox_info)

        return bboxes

    except Exception as e:
        print(f"bbox 추출 중 오류: {e}")
        return []


def process_bbox_coordinates(bbox: list, json_img_width: int = None, json_img_height: int = None) -> list:
    """bbox 좌표를 처리하여 올바른 포맷으로 변환합니다."""
    if not isinstance(bbox, list) or len(bbox) != 4:
        print(f"⚠️ 잘못된 bbox 형식: {bbox}")
        return bbox

    x1, y1, x2, y2 = bbox

    # 좌표 유효성 검사
    if x1 >= x2 or y1 >= y2:
        print(f"⚠️ bbox 좌표 순서 오류: {bbox}")
        return bbox

    # 정규화 좌표인지 확인 (0.0-1.0 범위)
    is_normalized = all(0.0 <= coord <= 1.0 for coord in [x1, y1, x2, y2])

    if is_normalized:
        # 정규화 좌표를 절대 좌표로 변환
        # JSON에 이미지 크기 정보가 있으면 사용, 없으면 기본값
        img_width = json_img_width if json_img_width else 250
        img_height = json_img_height if json_img_height else 250
        
        abs_x1 = int(x1 * img_width)
        abs_y1 = int(y1 * img_height)
        abs_x2 = int(x2 * img_width)
        abs_y2 = int(y2 * img_height)

        print(f"🔄 정규화→절대 변환 ({img_width}x{img_height}): [{x1:.3f}, {y1:.3f}, {x2:.3f}, {y2:.3f}] → [{abs_x1}, {abs_y1}, {abs_x2}, {abs_y2}]")
        return [abs_x1, abs_y1, abs_x2, abs_y2]
    else:
        # 절대 좌표인 경우 - JSON 이미지 크기 정보가 있으면 로그에 표시
        if json_img_width and json_img_height:
            print(f"✅ 절대 좌표 확인 (JSON: {json_img_width}x{json_img_height}): [{x1}, {y1}, {x2}, {y2}]")
        else:
            print(f"✅ 절대 좌표 확인: [{x1}, {y1}, {x2}, {y2}]")
        return bbox


def generate_label_color(label: str) -> QtGui.QColor:
    """label을 기반으로 유니크한 색상을 생성합니다."""
    # label을 해시해서 일관된 색상 생성
    hash_value = hash(label) % 360  # 0-359 범위

    # HSL에서 색상 생성 (채도는 높게, 밝기는 중간으로)
    hue = hash_value
    saturation = 200  # 높은 채도
    lightness = 150   # 중간 밝기

    color = QtGui.QColor()
    color.setHsl(hue, saturation, lightness)
    return color


def extract_detail_from_json(json_path: str) -> List[str]:
    """JSON 파일에서 detail 정보를 추출합니다."""
    if not json_path or not os.path.exists(json_path):
        return []

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        details = []

        # JSON 구조에 따라 detail 정보 추출
        if isinstance(data, dict):
            # 어노테이션 정보에서 detail 추출
            if 'annotations' in data and isinstance(data['annotations'], list):
                for ann in data['annotations']:
                    if isinstance(ann, dict):
                        label = ann.get('label', '')
                        score = ann.get('score', 0.0)
                        bbox = ann.get('bbox', [])
                        if label:
                            detail = f"{label} (신뢰도: {score:.3f})"
                            if bbox and len(bbox) == 4:
                                detail += f" 위치: [{bbox[0]}, {bbox[1]}, {bbox[2]}, {bbox[3]}]"
                            details.append(detail)

            # 기존 detail 키가 있는 경우
            if 'detail' in data:
                detail_data = data['detail']
                if isinstance(detail_data, list):
                    details.extend([str(item) for item in detail_data])
                elif isinstance(detail_data, str):
                    details.append(detail_data)
                elif isinstance(detail_data, dict):
                    # detail이 dict인 경우 모든 값 추출
                    for key, value in detail_data.items():
                        details.append(f"{key}: {value}")

            # 다른 가능한 구조들
            elif 'details' in data:
                detail_data = data['details']
                if isinstance(detail_data, list):
                    details.extend([str(item) for item in detail_data])
                elif isinstance(detail_data, str):
                    details.append(detail_data)

            # 전체 데이터에서 특정 패턴 찾기
            else:
                # defects, issues 등의 키 탐색
                for key in ['defects', 'issues', 'problems', 'anomalies']:
                    if key in data:
                        items = data[key]
                        if isinstance(items, list):
                            details.extend([str(item) for item in items])
                        break

        elif isinstance(data, list):
            # JSON이 리스트인 경우
            details.extend([str(item) for item in data])

        return details

    except Exception as e:
        print(f"JSON 파일 파싱 오류 ({json_path}): {e}")
        return []


def ensure_object_dtype(df: pd.DataFrame, column: str) -> None:
    try:
        df[column] = df[column].astype("object")
    except Exception:
        pass


def default_json_path(xlsx_path: str) -> str:
    base = os.path.basename(xlsx_path)
    root, _ = os.path.splitext(base)
    parent = os.path.dirname(xlsx_path) or os.getcwd()
    # Preferred naming: if Excel is *_labeled.xlsx → JSON is *_labels.json
    if root.endswith("_labeled"):
        new_root = root[: -len("_labeled")] + "_labels"
    else:
        new_root = root + "_labels"
    new_path = os.path.join(parent, f"{new_root}.json")
    # Legacy path compatibility: previously always appended _labels
    legacy_path = os.path.join(parent, f"{root}_labels.json")
    # If legacy exists and new doesn't, keep using legacy for seamless migration
    if os.path.exists(legacy_path) and not os.path.exists(new_path):
        return legacy_path
    return new_path


def load_label_store(json_path: str) -> dict:
    if not json_path or not os.path.exists(json_path):
        return {"version": 1, "updated_at": None, "labels": {}}
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, dict) and "labels" in data:
                return data
    except Exception:
        pass
    return {"version": 1, "updated_at": None, "labels": {}}


def save_label_store(json_path: str, store: dict) -> bool:
    """Atomically persist the label store. Returns True on success."""
    try:
        store["updated_at"] = datetime.now(_UTC).isoformat()
        tmp = json_path + ".tmp"
        os.makedirs(os.path.dirname(json_path) or ".", exist_ok=True)
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(store, f, ensure_ascii=False, indent=2)
        os.replace(tmp, json_path)
        return True
    except Exception:
        return False


def apply_json_to_excel(json_path: str, xlsx_path: str, sheet_name: str, col_indices: Dict[str, int], df: pd.DataFrame) -> int:
    store = load_label_store(json_path)
    labels = store.get("labels", {})
    applied = 0
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    for key, entry in labels.items():
        try:
            row_idx = int(key)
        except Exception:
            continue
        excel_row = row_idx + 2
        values = entry.get("values", {})
        for col_name, val in values.items():
            idx = col_indices.get(col_name)
            if idx is None:
                continue
            try:
                ws.cell(row=excel_row, column=idx, value=val)
                applied += 1
                if col_name in df.columns:
                    df.at[row_idx, col_name] = val
            except Exception:
                pass
    wb.save(xlsx_path)
    wb.close()
    return applied


def get_json_entry(json_path: str, row_idx: int) -> dict:
    store = load_label_store(json_path)
    key = str(row_idx)
    return store.get("labels", {}).get(key) or {}


def upsert_json_entry(json_path: str, row_idx: int, updater: Dict[str, object]) -> None:
    store = load_label_store(json_path)
    key = str(row_idx)
    entry = store["labels"].get(key) or {}
    for k, v in updater.items():
        entry[k] = v
    store["labels"][key] = entry
    save_label_store(json_path, store)


def merge_json_into_df(json_path: str, df: pd.DataFrame, label_columns: List[str]) -> None:
    """Load existing JSON labels and reflect into DataFrame so work can resume after restart."""
    store = load_label_store(json_path)
    labels = store.get("labels", {})
    for key, entry in labels.items():
        try:
            ridx = int(key)
        except Exception:
            continue
        if ridx not in df.index:
            continue
        values = entry.get("values", {})
        for col, val in values.items():
            if col in label_columns:
                try:
                    curr = None
                    try:
                        curr = df.at[ridx, col]
                    except Exception:
                        curr = None
                    is_empty = (curr is None) or (str(curr) == "")
                    try:
                        if not is_empty:
                            is_empty = bool(pd.isna(curr))
                    except Exception:
                        pass
                    if is_empty:
                        df.at[ridx, col] = val
                except Exception:
                    pass

def is_xlsx(path: str) -> bool:
    try:
        return os.path.isfile(path) and path.lower().endswith(".xlsx")
    except Exception:
        return False


def thumb_cache_path(images_base: str, resolved_path: str, target_edge: int) -> str:
    rel = os.path.relpath(resolved_path, images_base)
    key = hashlib.md5(f"{rel}|{target_edge}".encode("utf-8")).hexdigest()
    cache_dir = os.path.join(images_base, ".thumb_cache")
    os.makedirs(cache_dir, exist_ok=True)
    return os.path.join(cache_dir, f"{key}.png")


def build_thumb_if_needed(images_base: str, resolved_path: str, target_edge: int) -> str:
    thumb = thumb_cache_path(images_base, resolved_path, target_edge)
    try:
        src_mtime = os.path.getmtime(resolved_path)
        if os.path.exists(thumb) and os.path.getmtime(thumb) >= src_mtime:
            return thumb
        # Use Qt to scale and save
        img = QtGui.QImage(resolved_path)
        if img.isNull():
            return resolved_path
        w, h = img.width(), img.height()
        scale = target_edge / float(max(w, h))
        if scale < 1.0:
            img = img.scaled(int(w * scale), int(h * scale), QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        img.save(thumb, "PNG")
        return thumb
    except Exception:
        return resolved_path


class SetupWindow(QtWidgets.QDialog):
    """설정 페이지 - CSV 파일과 이미지 폴더 경로를 설정하고 매칭 테스트를 수행합니다."""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("CSV 라벨링 도구 설정")
        self.resize(900, 700)
        self.setModal(True)
        
        # 설정값
        self.csv_path = ""
        self.images_base = ""
        self.json_base = ""
        self.csv_type = "report"
        
        self._build_ui()
        self._load_default_paths()
    
    def _build_ui(self):
        """UI 구성"""
        main_layout = QtWidgets.QVBoxLayout(self)

        # 제목
        title_label = QtWidgets.QLabel("CSV 라벨링 도구 설정")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #1976d2; margin: 10px;")
        title_label.setAlignment(QtCore.Qt.AlignCenter)
        main_layout.addWidget(title_label)

        # 스크롤 영역 생성
        scroll_area = QtWidgets.QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)

        # 스크롤될 컨텐츠 위젯
        content_widget = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(content_widget)

        scroll_area.setWidget(content_widget)
        main_layout.addWidget(scroll_area)
        
        # CSV 파일 선택
        csv_group = QtWidgets.QGroupBox("CSV 파일 선택")
        csv_layout = QtWidgets.QVBoxLayout(csv_group)
        
        csv_info = QtWidgets.QLabel("라벨링할 CSV 파일을 선택하세요.")
        csv_layout.addWidget(csv_info)
        
        csv_path_layout = QtWidgets.QHBoxLayout()
        self.csv_path_edit = QtWidgets.QLineEdit()
        self.csv_path_edit.setPlaceholderText("CSV 파일 경로를 입력하거나 선택하세요")
        self.csv_path_edit.setReadOnly(True)
        csv_path_layout.addWidget(self.csv_path_edit)
        
        self.csv_browse_btn = QtWidgets.QPushButton("파일 찾기")
        self.csv_browse_btn.clicked.connect(self._browse_csv)
        csv_path_layout.addWidget(self.csv_browse_btn)
        
        csv_layout.addLayout(csv_path_layout)
        layout.addWidget(csv_group)

        # 라벨링 옵션
        labeling_group = QtWidgets.QGroupBox("라벨링 옵션")
        labeling_layout = QtWidgets.QVBoxLayout(labeling_group)

        labeling_info = QtWidgets.QLabel("새로운 CSV 파일을 시작할 때의 라벨링 설정을 선택하세요.")
        labeling_layout.addWidget(labeling_info)

        # 새로운 라벨링 시작 옵션
        self.skip_existing_labels_chk = QtWidgets.QCheckBox("새로운 라벨링 시작 (기존 라벨링 데이터 무시)")
        self.skip_existing_labels_chk.setToolTip("체크하면 기존 JSON 파일의 라벨링 데이터를 로드하지 않고 완전히 새로운 상태로 시작합니다.")
        labeling_layout.addWidget(self.skip_existing_labels_chk)

        layout.addWidget(labeling_group)
        
        # 이미지 폴더 선택
        images_group = QtWidgets.QGroupBox("이미지 폴더 선택")
        images_layout = QtWidgets.QVBoxLayout(images_group)
        
        images_info = QtWidgets.QLabel("CSV 파일의 이미지들이 저장된 폴더를 선택하세요.")
        images_layout.addWidget(images_info)
        
        images_path_layout = QtWidgets.QHBoxLayout()
        self.images_path_edit = QtWidgets.QLineEdit()
        self.images_path_edit.setPlaceholderText("이미지 폴더 경로를 입력하거나 선택하세요")
        self.images_path_edit.setReadOnly(True)
        images_path_layout.addWidget(self.images_path_edit)
        
        self.images_browse_btn = QtWidgets.QPushButton("폴더 찾기")
        self.images_browse_btn.clicked.connect(self._browse_images)
        images_path_layout.addWidget(self.images_browse_btn)
        
        images_layout.addLayout(images_path_layout)
        layout.addWidget(images_group)

        # JSON 폴더 선택
        json_group = QtWidgets.QGroupBox("JSON 폴더 선택")
        json_layout = QtWidgets.QVBoxLayout(json_group)

        json_info = QtWidgets.QLabel("JSON 파일들이 저장된 폴더를 선택하세요.")
        json_layout.addWidget(json_info)

        json_path_layout = QtWidgets.QHBoxLayout()
        self.json_path_edit = QtWidgets.QLineEdit()
        self.json_path_edit.setPlaceholderText("JSON 폴더 경로를 입력하거나 선택하세요")
        self.json_path_edit.setReadOnly(True)
        json_path_layout.addWidget(self.json_path_edit)

        self.json_browse_btn = QtWidgets.QPushButton("폴더 찾기")
        self.json_browse_btn.clicked.connect(self._browse_json)
        json_path_layout.addWidget(self.json_browse_btn)

        json_layout.addLayout(json_path_layout)
        layout.addWidget(json_group)
        
        # CSV 타입 선택 (리포트 단일로 고정)
        type_group = QtWidgets.QGroupBox("CSV 타입 설정")
        type_layout = QtWidgets.QVBoxLayout(type_group)
        
        type_info = QtWidgets.QLabel("CSV 타입: 리포트 단일 (Report Single)")
        type_info.setStyleSheet("font-weight: bold; color: #1976d2;")
        type_layout.addWidget(type_info)
        
        # 고정된 타입 표시
        type_display = QtWidgets.QLabel("📊 리포트 CSV 파일을 사용합니다.")
        type_display.setStyleSheet("color: #666; font-style: italic;")
        type_layout.addWidget(type_display)

        layout.addWidget(type_group)

        # 마지막 경로 설정 복원 버튼
        restore_group = QtWidgets.QGroupBox("저장된 경로 복원")
        restore_layout = QtWidgets.QVBoxLayout(restore_group)
        
        restore_info = QtWidgets.QLabel("이전에 사용한 경로 설정을 복원할 수 있습니다.")
        restore_layout.addWidget(restore_info)
        
        restore_buttons_layout = QtWidgets.QHBoxLayout()
        
        self.btn_restore_paths = QtWidgets.QPushButton("저장된 경로 복원")
        self.btn_restore_paths.clicked.connect(self._restore_saved_paths)
        restore_buttons_layout.addWidget(self.btn_restore_paths)
        
        restore_layout.addLayout(restore_buttons_layout)
        layout.addWidget(restore_group)
        
        # 매칭 테스트 결과
        test_group = QtWidgets.QGroupBox("매칭 테스트 결과")
        test_layout = QtWidgets.QVBoxLayout(test_group)
        
        self.test_result_label = QtWidgets.QLabel("CSV 파일과 이미지 폴더를 선택한 후 테스트를 실행하세요.")
        self.test_result_label.setWordWrap(True)
        test_layout.addWidget(self.test_result_label)
        
        test_buttons_layout = QtWidgets.QHBoxLayout()
        self.test_btn = QtWidgets.QPushButton("매칭 테스트 실행")
        self.test_btn.clicked.connect(self._run_matching_test)
        self.test_btn.setEnabled(False)
        test_buttons_layout.addWidget(self.test_btn)
        test_layout.addLayout(test_buttons_layout)
        
        layout.addWidget(test_group)

        # 스트레치 추가로 스크롤 영역을 채움
        layout.addStretch()

        # 진행 버튼 (스크롤 영역 바깥에 위치)
        button_layout = QtWidgets.QHBoxLayout()

        self.cancel_btn = QtWidgets.QPushButton("취소")
        self.cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_btn)

        self.start_btn = QtWidgets.QPushButton("라벨링 시작")
        self.start_btn.clicked.connect(self.accept)
        self.start_btn.setEnabled(False)
        button_layout.addWidget(self.start_btn)

        main_layout.addLayout(button_layout)
    
    def _load_default_paths(self):
        """기본 경로 로드"""
        # 리포트 단일 타입이 기본값
        self.csv_type = "report"
        self.csv_path = CSV_CONFIGS["report"]["csv_path"]
        self.images_base = CSV_CONFIGS["report"]["images_base"]
        self.json_base = CSV_CONFIGS["report"]["json_base"]

        self.csv_path_edit.setText(self.csv_path)
        self.images_path_edit.setText(self.images_base)
        self.json_path_edit.setText(self.json_base)

        self._update_test_button_state()
        
        # 저장된 경로가 있으면 복원 시도
        self._try_restore_saved_paths()
    
    def _restore_saved_paths(self):
        """저장된 경로 설정을 복원"""
        self.load_paths_from_settings()
        self.csv_path_edit.setText(self.csv_path)
        self.images_path_edit.setText(self.images_base)
        self.json_path_edit.setText(self.json_base)
            
        self._update_test_button_state()
        print("저장된 경로 설정이 복원되었습니다.")
    
    def _try_restore_saved_paths(self):
        """초기화 시 저장된 경로가 있으면 자동으로 복원 시도"""
        settings = QtCore.QSettings("rtm", "inference_labeler")
        last_csv_path = settings.value("last_csv_path", "")
        last_images_base = settings.value("last_images_base", "")
        last_json_base = settings.value("last_json_base", "")
        
        if last_csv_path and os.path.exists(last_csv_path):
            self.csv_path = last_csv_path
            self.csv_path_edit.setText(self.csv_path)
            
        if last_images_base and os.path.exists(last_images_base):
            self.images_base = last_images_base
            self.images_path_edit.setText(self.images_base)
            
        if last_json_base and os.path.exists(last_json_base):
            self.json_base = last_json_base
            self.json_path_edit.setText(self.json_base)
            
        if self.csv_path != CSV_CONFIGS[self.csv_type]["csv_path"] or \
           self.images_base != CSV_CONFIGS[self.csv_type]["images_base"] or \
           self.json_base != CSV_CONFIGS[self.csv_type]["json_base"]:
            print("저장된 경로 설정이 복원되었습니다.")
    

    
    def _browse_csv(self):
        """CSV 파일 찾기"""
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "CSV 파일 선택",
            os.path.expanduser("~/Downloads"),
            "CSV 파일 (*.csv);;모든 파일 (*)"
        )
        
        if file_path:
            self.csv_path = file_path
            self.csv_path_edit.setText(file_path)
            
            # 리포트 단일 타입으로 고정되어 있으므로 별도 처리 불필요
            self._update_test_button_state()
    
    def _browse_images(self):
        """이미지 폴더 찾기"""
        folder_path = QtWidgets.QFileDialog.getExistingDirectory(
            self,
            "이미지 폴더 선택",
            os.path.expanduser("~/Downloads")
        )

        if folder_path:
            self.images_base = folder_path
            self.images_path_edit.setText(folder_path)
            self._update_test_button_state()

    def _browse_json(self):
        """JSON 폴더 찾기"""
        folder_path = QtWidgets.QFileDialog.getExistingDirectory(
            self,
            "JSON 폴더 선택",
            os.path.expanduser("~/Downloads")
        )

        if folder_path:
            self.json_base = folder_path
            self.json_path_edit.setText(folder_path)
            self._update_test_button_state()
    
    def _update_test_button_state(self):
        """테스트 버튼 활성화 상태 업데이트"""
        can_test = bool(self.csv_path and self.images_base and self.json_base and
                       os.path.exists(self.csv_path) and os.path.exists(self.images_base) and os.path.exists(self.json_base))
        self.test_btn.setEnabled(can_test)
    
    def _run_matching_test(self):
        """매칭 테스트 실행"""
        print("🧪 매칭 테스트 시작...")
        # 상태 바에 진행 상황 표시
        if hasattr(self, 'parent') and hasattr(self.parent(), 'status'):
            self.parent().status.showMessage("🧪 매칭 테스트 실행 중...")

        if not self.csv_path or not self.images_base:
            print("❌ CSV 경로 또는 이미지 경로가 설정되지 않음")
            return

        try:
            print(f"📄 CSV 파일 로드 중: {self.csv_path}")
            # 상태 바 업데이트
            if hasattr(self, 'parent') and hasattr(self.parent(), 'status'):
                self.parent().status.showMessage("📄 CSV 파일 로드 중...")

            # CSV 파일 로드
            df = pd.read_csv(self.csv_path, nrows=100)  # 처음 100행만 테스트
            print(f"✅ CSV 로드 완료: {len(df)} 행, 컬럼: {list(df.columns)}")

            if "File_path" not in df.columns:
                print("❌ CSV 파일에 'File_path' 컬럼이 없음")
                self.test_result_label.setText("❌ CSV 파일에 'File_path' 컬럼이 없습니다.")
                return
            
            # 이미지 및 JSON 매칭 테스트 (최대 10개만 테스트)
            total_rows = len(df)
            test_count = min(10, total_rows)  # 최대 10개만 테스트
            image_matched_count = 0
            json_matched_count = 0
            sample_image_matches = []
            sample_json_matches = []
            print(f"🔍 매칭 테스트 시작: {test_count}개 행 검사 (총 {total_rows}개 중)")

            for idx in range(test_count):
                row = df.iloc[idx]
                file_path = row["File_path"]
                if pd.isna(file_path) or not str(file_path).strip():
                    continue

                print(f"🔎 행 {idx} 검사: {os.path.basename(file_path)}")

                # 1. 이미지 파일 매칭 테스트
                resolved_path = resolve_image_path(self.images_base, str(file_path))
                if resolved_path and os.path.exists(resolved_path):
                    image_matched_count += 1
                    print(f"  ✅ 이미지: {os.path.basename(resolved_path)}")
                    if len(sample_image_matches) < 2:
                        sample_image_matches.append(os.path.basename(resolved_path))
                else:
                    print(f"  ❌ 이미지: 찾을 수 없음")

                # 2. JSON 파일 매칭 테스트 (개선된 알고리즘)
                json_found = False
                if self.json_base:
                    # File_path에서 파일명 추출
                    filename = os.path.basename(file_path)
                    name_without_ext = os.path.splitext(filename)[0]

                    print(f"  🔍 JSON 검색: {filename} -> {name_without_ext}.json")

                    # 개선된 JSON 검색 패턴들
                    json_candidates = []

                    # 기본 패턴들
                    json_candidates.extend([
                        os.path.join(self.json_base, name_without_ext + '.json'),
                        os.path.join(self.json_base, filename + '.json'),
                        os.path.join(self.json_base, filename),
                        os.path.join(self.json_base, name_without_ext, name_without_ext + '.json'),
                    ])

                    # 초유연 구조 기반 JSON 검색
                    import re

                    # CSV File_path 구조 분석
                    csv_number_pattern = re.search(r'/(\d+)/', file_path)
                    csv_structure = {}

                    if csv_number_pattern:
                        csv_number = csv_number_pattern.group(1)
                        csv_structure['number'] = csv_number

                        # /숫자/ 이후 경로 분석
                        after_number = file_path.split(f'/{csv_number}/', 1)[1]
                        path_parts = after_number.split('/')

                        if len(path_parts) >= 4:
                            csv_structure.update({
                                'part1': path_parts[0],  # 0001
                                'part2': path_parts[1],  # Unit
                                'part3': path_parts[2],  # U12, U70 등
                                'part4': path_parts[3],  # BC, FC 등
                            })

                    # JSON 기본 경로 구조 분석
                    base_number_pattern = re.search(r'/test/(\d+)/', self.json_base)
                    base_structure = {}

                    if base_number_pattern:
                        base_number = base_number_pattern.group(1)
                        base_structure['number'] = base_number

                        # 기본 경로의 나머지 부분 분석
                        after_base_number = self.json_base.split(f'/test/{base_number}/', 1)[1]
                        base_path_parts = after_base_number.split('/')

                        if len(base_path_parts) >= 4:
                            base_structure.update({
                                'part1': base_path_parts[0],
                                'part2': base_path_parts[1],
                                'part3': base_path_parts[2],
                                'part4': base_path_parts[3],
                            })

                    # 구조 기반 JSON 패턴 생성
                    if csv_structure and base_structure:
                        # 다양한 Unit 폴더 조합 생성
                        unit_folders = ['U0', 'U1', 'U2', 'U6', 'U7', 'U8', 'U9', 'U10', 'U11', 'U12', 'U13', 'U14', 'U15', 'U16', 'U19']
                        type_folders = ['BC', 'FC', 'DC']

                        for unit in unit_folders:
                            for type_folder in type_folders:
                                # JSON 구조 기반 경로 생성
                                json_struct_path = f"{base_number}/{base_structure.get('part2', 'Unit')}/{unit}/{type_folder}/{name_without_ext}.json"
                                json_candidates.append(os.path.join(self.json_base.split(f'/test/{base_number}/')[0], "test", json_struct_path))

                                # /Unit 폴더에서도
                                unit_base = os.path.join(self.json_base.split(f'/test/{base_number}/')[0], "test", "Unit", f"{name_without_ext}.json")
                                json_candidates.append(unit_base)

                                # /img 폴더에서도
                                img_base = os.path.join(self.json_base.split(f'/test/{base_number}/')[0], "test", "img", json_struct_path)
                                json_candidates.append(img_base)

                    # 기존 방식도 유지 (폴백)
                    number_pattern_in_base = re.search(r'/test/(\d+)/', self.json_base)
                    if number_pattern_in_base:
                        number_part = number_pattern_in_base.group(1)
                        unit_base = self.json_base.replace(f'/test/{number_part}/', '/test/Unit/')
                        json_candidates.extend([
                            os.path.join(unit_base, name_without_ext + '.json'),
                            os.path.join(unit_base, filename),
                        ])

                    # Unit 폴더 간 검색 (U0 -> U9, U12 등)
                    if '/Unit/' in self.json_base:
                        # Unit 폴더 목록
                        unit_folders = ['U0', 'U1', 'U2', 'U6', 'U7', 'U8', 'U9', 'U10', 'U11', 'U12', 'U13', 'U14', 'U15', 'U16', 'U19']

                        for unit_folder in unit_folders:
                            # 기본 Unit 경로에서
                            unit_path = self.json_base.replace('/Unit/U0/', f'/Unit/{unit_folder}/')
                            json_candidates.extend([
                                os.path.join(unit_path, name_without_ext + '.json'),
                                os.path.join(unit_path, filename),
                            ])

                            # /img Unit 경로에서도
                            if number_pattern_in_base:
                                number_part = number_pattern_in_base.group(1)
                                img_unit_path = self.json_base.replace(f'/test/{number_part}/', '/test/img/1/').replace('/Unit/U0/', f'/Unit/{unit_folder}/')
                                json_candidates.extend([
                                    os.path.join(img_unit_path, name_without_ext + '.json'),
                                    os.path.join(img_unit_path, filename),
                                ])

                        # 기존 Unit 기반 폴백
                        unit_parent = os.path.dirname(self.json_base)
                        json_candidates.extend([
                            os.path.join(unit_parent, "**", name_without_ext + '.json'),
                            os.path.join(unit_parent, "**", filename),
                        ])

                    # 모든 후보 경로에서 검색
                    for candidate in json_candidates:
                        if os.path.exists(candidate):
                            json_matched_count += 1
                            print(f"  ✅ JSON: {os.path.basename(candidate)} (경로: {os.path.dirname(candidate)})")
                            if len(sample_json_matches) < 2:
                                sample_json_matches.append(os.path.basename(candidate))
                            json_found = True
                            break

                        # glob 패턴으로 검색 (디렉토리인 경우)
                        if os.path.isdir(os.path.dirname(candidate)):
                            import glob
                            pattern = os.path.join(os.path.dirname(candidate), "**", os.path.basename(candidate))
                            matches = glob.glob(pattern, recursive=True)
                            if matches:
                                json_matched_count += 1
                                found_path = matches[0]
                                print(f"  ✅ JSON: {os.path.basename(found_path)} (재귀 검색)")
                                if len(sample_json_matches) < 2:
                                    sample_json_matches.append(os.path.basename(found_path))
                                json_found = True
                                break

                    if not json_found:
                        print(f"  ❌ JSON: 찾을 수 없음 (기본경로: {self.json_base})")
                else:
                    print(f"  ⚠️ JSON 경로 설정 안됨")

            print(f"📈 매칭 결과:")
            print(f"   이미지: {image_matched_count}/{test_count}개 찾음")
            print(f"   JSON: {json_matched_count}/{test_count}개 찾음")
            
            # 결과 표시 (이미지와 JSON 모두 고려)
            image_match_rate = (image_matched_count / test_count * 100) if test_count > 0 else 0
            json_match_rate = (json_matched_count / test_count * 100) if test_count > 0 else 0
            overall_match_rate = ((image_matched_count + json_matched_count) / (test_count * 2) * 100) if test_count > 0 else 0

            if overall_match_rate > 75:
                status = "✅"
                color = "green"
                self.start_btn.setEnabled(True)
            elif overall_match_rate > 40:
                status = "⚠️"
                color = "orange"
                self.start_btn.setEnabled(True)
            else:
                status = "❌"
                color = "red"
                self.start_btn.setEnabled(False)
            
            result_text = f"{status} 매칭 테스트 결과:\n"
            result_text += f"전체 행: {total_rows:,}개\n"
            result_text += f"테스트 행: {test_count}개\n"
            result_text += f"이미지 매칭: {image_matched_count}/{test_count} ({image_match_rate:.1f}%)\n"
            result_text += f"JSON 매칭: {json_matched_count}/{test_count} ({json_match_rate:.1f}%)\n"
            result_text += f"종합 매칭률: {overall_match_rate:.1f}%\n\n"

            if sample_image_matches:
                result_text += f"샘플 이미지 파일:\n"
                for match in sample_image_matches:
                    result_text += f"  📷 {match}\n"

            if sample_json_matches:
                result_text += f"샘플 JSON 파일:\n"
                for match in sample_json_matches:
                    result_text += f"  📄 {match}\n"

            # bbox 정보도 표시 (JSON이 있는 경우)
            if sample_json_matches and json_matched_count > 0:
                result_text += f"\n💡 JSON 파일이 있으면 bbox 오버레이 표시 가능"
            
            self.test_result_label.setText(result_text)
            self.test_result_label.setStyleSheet(f"color: {color}; font-weight: bold;")
            
        except Exception as e:
            self.test_result_label.setText(f"❌ 테스트 실행 중 오류 발생:\n{str(e)}")
            self.test_result_label.setStyleSheet("color: red; font-weight: bold;")
    
    def get_settings(self):
        """설정값 반환"""
        return {
            "csv_path": self.csv_path,
            "images_base": self.images_base,
            "json_base": self.json_base,
            "csv_type": self.csv_type,
            "skip_existing_labels": self.skip_existing_labels_chk.isChecked()
        }

    def save_paths_to_settings(self):
        """경로 설정을 QSettings에 저장"""
        settings = QtCore.QSettings("rtm", "inference_labeler")
        settings.setValue("last_csv_path", self.csv_path)
        settings.setValue("last_images_base", self.images_base)
        settings.setValue("last_json_base", self.json_base)
        settings.setValue("last_csv_type", self.csv_type)
        print(f"경로 설정 저장됨: CSV={self.csv_path}, 이미지={self.images_base}, JSON={self.json_base}")

    def load_paths_from_settings(self):
        """QSettings에서 마지막 경로 설정을 로드"""
        settings = QtCore.QSettings("rtm", "inference_labeler")
        last_csv_path = settings.value("last_csv_path", "")
        last_images_base = settings.value("last_images_base", "")
        last_json_base = settings.value("last_json_base", "")
        last_csv_type = settings.value("last_csv_type", "report")
        
        if last_csv_path and os.path.exists(last_csv_path):
            self.csv_path = last_csv_path
        if last_images_base and os.path.exists(last_images_base):
            self.images_base = last_images_base
        if last_json_base and os.path.exists(last_json_base):
            self.json_base = last_json_base
        if last_csv_type == "report":
            self.csv_type = last_csv_type
            
        print(f"저장된 경로 설정 로드됨: CSV={self.csv_path}, 이미지={self.images_base}, JSON={self.json_base}")

    def accept(self):
        """라벨링 시작 버튼 클릭 시 설정값 검증"""
        # 설정값 검증
        if not self.csv_path or not os.path.exists(self.csv_path):
            QtWidgets.QMessageBox.critical(self, "오류", "CSV 파일을 찾을 수 없거나 선택되지 않았습니다.")
            return

        if not self.images_base or not os.path.exists(self.images_base):
            QtWidgets.QMessageBox.critical(self, "오류", "이미지 폴더를 찾을 수 없거나 선택되지 않았습니다.")
            return

        if not self.json_base or not os.path.exists(self.json_base):
            QtWidgets.QMessageBox.critical(self, "오류", "JSON 폴더를 찾을 수 없거나 선택되지 않았습니다.")
            return

        # 모든 검증 통과 시 부모의 accept() 호출
        super().accept()


class InferenceLabelerWindow(QtWidgets.QMainWindow):
    def __init__(self, settings: dict = None) -> None:
        super().__init__()
        print("🚀 InferenceLabelerWindow 초기화 시작")
        self.setWindowTitle("추론 결과 라벨링 도구")
        self.resize(1400, 900)
        # UI readiness flag
        self._ui_ready = False

        # 설정에서 경로 가져오기
        if settings:
            self.csv_path = settings.get("csv_path", CSV_CONFIGS["report"]["csv_path"])
            self.images_base = settings.get("images_base", CSV_CONFIGS["report"]["images_base"])
            self.json_base = settings.get("json_base", CSV_CONFIGS["report"]["json_base"])
            csv_type = settings.get("csv_type", "report")
            self.skip_existing_labels = settings.get("skip_existing_labels", False)
            self.setWindowTitle(f"추론 결과 라벨링 도구 - {csv_type.upper()} ({os.path.basename(self.csv_path)})")
        else:
            # 기본값 사용
            self.csv_path = CSV_CONFIGS["report"]["csv_path"]
            self.images_base = CSV_CONFIGS["report"]["images_base"]
            self.json_base = CSV_CONFIGS["report"]["json_base"]
            self.skip_existing_labels = False
        
        # State
        self.df: Optional[pd.DataFrame] = None
        self.json_path: str = ""
        self.col_indices: Dict[str, int] = {}
        
        # Memory management and performance settings - optimized for 50k+ records
        system_memory = get_system_memory()
        self.max_memory_mb = min(2048, system_memory * 0.3)  # Increased for large datasets
        self.chunk_size = 1000  # Increased chunk size
        self.max_table_rows = 200  # Reduced for better performance
        self.image_cache_size = 12  # Increased for better performance and faster navigation
        self._image_cache: Dict[str, QtGui.QPixmap] = {}
        self._lazy_loading = True
        
        # Performance optimization flags
        self._filter_cache: Optional[pd.Series] = None
        self._last_filter_hash: Optional[str] = None
        
        # Unified labeling approach - single active column with as-is/to-be integration
        self.active_label_col: str = "Manual_Label"
        self.label_choices: List[str] = [
            "OK",
            "애매한 OK", 
            "NG", 
            "애매한 NG",
            "보류",
            "SRLogicOK",
        ]
        
        self.current_idx: int = 0
        self.filtered_indices: List[int] = []
        self.fit_to_window: bool = True
        self.tobe_choices: List[str] = [
            "돌기",
            "흑점", 
            "색상얼룩",
            "찍힘",
            "SR이물",
            "SR금속",
        ]
        
        # Unique_seg_result filter choices
        self.pred_filter_choices: List[str] = []
        self.selected_pred_filters: set = set()
        self.pred_filter_checkboxes: Dict[str, QtWidgets.QCheckBox] = {}
        
        # Auto-advance settings
        self.auto_advance_enabled: bool = True
        
        # AS-IS/TO-BE mode settings
        self.as_is_tobe_mode: bool = False

        # 오버레이 표시 설정
        self.show_overlay: bool = True
        

        
        # Performance monitoring
        self._label_count = 0
        self._session_start_time = datetime.now()
        
        # Settings and navigation
        self.settings = QtCore.QSettings("rtm", "inference_labeler")
        self._navigating: bool = False
        
        # Batched JSON save - optimized for large datasets
        self._pending_ops: List[Tuple[str, int, Dict[str, object], Dict[str, str]]] = []
        self._pending_json_path: str = ""
        self._save_timer = QtCore.QTimer(self)
        self._save_timer.setSingleShot(True)
        self._save_timer.setInterval(500)  # Reduced to 500ms for faster response
        self._save_timer.timeout.connect(self._flush_pending_ops)
        
        # Performance optimization settings
        self._ui_update_throttle = QtCore.QTimer(self)
        self._ui_update_throttle.setSingleShot(True)
        self._ui_update_throttle.setInterval(50)  # Reduced to 50ms for faster UI updates
        self._ui_update_throttle.timeout.connect(self._deferred_ui_update)
        self._pending_ui_update = False
        
        # Periodic session saving
        self._session_save_timer = QtCore.QTimer(self)
        self._session_save_timer.timeout.connect(self.save_session_state)
        self._session_save_timer.start(60000)  # Save every 60 seconds for better performance
        
        # Performance optimization flags
        self._last_table_update = 0
        self._table_update_throttle = 100  # Reduced to 100ms between table updates
        self._last_image_path = ""
        self._image_update_throttle = 200  # Increased to 200ms between image updates

        # UI
        self._build_ui()
        self._connect_shortcuts()
        
        # Defer data loading until UI is ready - wait for UI to settle
        QtCore.QTimer.singleShot(100, self._auto_load_data)
        # Defer session restore even more to ensure data is loaded first
        QtCore.QTimer.singleShot(1000, self.restore_session_state)

    def _auto_load_data(self):
        """Automatically load the CSV data on startup"""
        if not getattr(self, "_ui_ready", False):
            print("⏸️ _auto_load_data: UI not ready yet, retrying in 200ms...")
            QtCore.QTimer.singleShot(200, self._auto_load_data)
            return
        
        print("📊 UI 준비 완료, 데이터 로드 시작...")
        if os.path.exists(self.csv_path):
            self.load_csv_data()
        else:
            self.status.showMessage(f"CSV 파일을 찾을 수 없음: {self.csv_path}")

    def _build_ui(self) -> None:
        """UI 빌드 - 고정된 순서로 안정적 초기화"""
        print("🔧 UI 빌드 시작...")
        
        # Step 1: Central widget + QSplitter 생성
        print("1️⃣ Central Widget + QSplitter 생성...")
        self.splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        self.setCentralWidget(self.splitter)
        
        # Step 2: 좌/중/우 패널과 핵심 위젯들 모두 생성
        print("2️⃣ 모든 핵심 위젯 생성...")
        self._create_all_core_widgets()
        
        # Step 3: 시그널 연결
        print("3️⃣ 시그널 연결...")
        self._connect_all_signals()
        
        # Step 4: 기타 UI 요소들
        print("4️⃣ 기타 UI 요소 생성...")
        try:
            self.status = self.statusBar()
            self._create_status_widgets()
            self._apply_theme()
            self._create_toolbar()
            self._create_menus()
        except Exception as e:
            print(f"❌ 기타 UI 요소 생성 오류: {e}")
        
        # UI 완전히 구축 완료
        self._ui_ready = True
        print("✅ UI 빌드 완료 (ui_ready=True)")
        print("📊 모든 핵심 위젯이 안정적으로 초기화됨")

    def _create_all_core_widgets(self) -> None:
        """모든 핵심 위젯들을 고정된 순서로 생성"""
        print("🏗️ 핵심 위젯 생성 시작...")
        
        # 좌측 패널: 이미지 뷰어
        print("📸 이미지 패널 생성...")
        self.scroll_area = QtWidgets.QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.image_label = QtWidgets.QLabel(alignment=QtCore.Qt.AlignCenter)
        self.image_label.setScaledContents(False)
        self.image_label.setBackgroundRole(QtGui.QPalette.Base)
        self.scroll_area.setWidget(self.image_label)
        
        self.image_status_bar = QtWidgets.QLabel("")
        self.image_status_bar.setAlignment(QtCore.Qt.AlignCenter)
        self.image_status_bar.setStyleSheet("""
            QLabel {
                background-color: #f0f0f0;
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 8px;
                margin: 2px;
                font-weight: bold;
                color: #333;
            }
        """)
        
        self.path_label = QtWidgets.QLabel("")
        self.path_label.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse)
        self.path_label.setWordWrap(True)
        
        image_panel = QtWidgets.QWidget()
        image_layout = QtWidgets.QVBoxLayout(image_panel)
        image_layout.setContentsMargins(0, 0, 0, 0)
        image_layout.setSpacing(2)
        image_layout.addWidget(self.image_status_bar)
        image_layout.addWidget(self.scroll_area)
        image_layout.addWidget(self.path_label)
        
        # 중간 패널: 컨트롤들
        print("🎛️ 컨트롤 패널 생성...")
        controls_panel = QtWidgets.QWidget()
        controls_layout = QtWidgets.QVBoxLayout(controls_panel)
        controls_layout.setSpacing(3)
        controls_layout.setContentsMargins(3, 3, 3, 3)
        
        self.controls_scroll_area = QtWidgets.QScrollArea()
        self.controls_scroll_area.setWidgetResizable(True)
        self.controls_scroll_area.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.controls_scroll_area.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOn)
        
        self.controls_widget = QtWidgets.QWidget()
        self.controls_layout = QtWidgets.QVBoxLayout(self.controls_widget)
        self.controls_layout.setSpacing(3)
        self.controls_layout.setContentsMargins(3, 3, 3, 3)
        
        self.controls_scroll_area.setWidget(self.controls_widget)
        controls_layout.addWidget(self.controls_scroll_area)
        
        # 우측 패널: 테이블
        print("📊 테이블 패널 생성...")  
        table_panel = QtWidgets.QWidget()
        table_layout = QtWidgets.QVBoxLayout(table_panel)
        table_layout.setContentsMargins(3, 3, 3, 3)
        table_layout.setSpacing(3)
        
        table_label = QtWidgets.QLabel("데이터 미리보기")
        table_label.setSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Fixed)
        table_layout.addWidget(table_label)
        
        self.table = QtWidgets.QTableWidget()
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setWordWrap(False)
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        self.table.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        self.table.setMinimumHeight(400)
        table_layout.addWidget(self.table, 1)
        
        # 패널들을 스플리터에 추가
        self.splitter.addWidget(image_panel)
        self.splitter.addWidget(controls_panel)
        self.splitter.addWidget(table_panel)
        self.splitter.setSizes([600, 400, 400])
        
        # 컨트롤 패널의 내부 위젯들 생성
        self._create_control_contents()
        
        print("✅ 모든 핵심 위젯 생성 완료")
    
    def _create_control_contents(self) -> None:
        """컨트롤 패널의 내부 컨텐츠 생성"""
        print("🔧 컨트롤 내용 생성...")
        
        # Progress dashboard
        progress_dashboard = self._create_progress_dashboard()
        self.controls_layout.addWidget(progress_dashboard)

        # Current row info
        self.lbl_current_info = QtWidgets.QLabel("데이터가 로드되지 않음")
        self.controls_layout.addWidget(self.lbl_current_info)

        # Bookmark section
        grp_bookmark_memo = QtWidgets.QGroupBox("북마크")
        bookmark_memo_layout = QtWidgets.QVBoxLayout(grp_bookmark_memo)
        bookmark_memo_layout.setContentsMargins(5, 5, 5, 5)
        bookmark_memo_layout.setSpacing(5)
        
        bookmark_controls = QtWidgets.QHBoxLayout()
        self.btn_toggle_bookmark = QtWidgets.QPushButton("북마크 토글 (B)")
        self.lbl_bookmark_status = QtWidgets.QLabel("북마크: ❌")
        bookmark_controls.addWidget(self.btn_toggle_bookmark)
        bookmark_controls.addWidget(self.lbl_bookmark_status)
        bookmark_controls.addStretch()
        bookmark_memo_layout.addLayout(bookmark_controls)
        self.controls_layout.addWidget(grp_bookmark_memo)

        # Quick labeling section
        grp_labeling = QtWidgets.QGroupBox()
        labeling_main_layout = QtWidgets.QVBoxLayout(grp_labeling)
        labeling_title = QtWidgets.QLabel("빠른 라벨링")
        labeling_main_layout.addWidget(labeling_title)
        
        self.quick_labeling_container = QtWidgets.QWidget()
        quick_labeling_layout = QtWidgets.QVBoxLayout(self.quick_labeling_container)
        quick_labeling_layout.setSpacing(2)
        quick_labeling_layout.setContentsMargins(5, 2, 5, 2)
        
        self.choice_buttons_scroll = QtWidgets.QScrollArea()
        self.choice_buttons_scroll.setWidgetResizable(True)
        self.choice_buttons_scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.choice_buttons_scroll.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        
        self.choice_buttons_container = QtWidgets.QWidget()
        self.choice_buttons_layout = QtWidgets.QVBoxLayout(self.choice_buttons_container)
        self.choice_buttons_layout.setSpacing(4)
        self.choice_buttons_layout.setContentsMargins(4, 4, 4, 4)
        
        self.choice_buttons_scroll.setWidget(self.choice_buttons_container)
        quick_labeling_layout.addWidget(self.choice_buttons_scroll)
        labeling_main_layout.addWidget(self.quick_labeling_container)
        self.choice_buttons_scroll.setMinimumHeight(100)
        grp_labeling.setSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        self.controls_layout.addWidget(grp_labeling)

        # AS-IS/TO-BE section
        grp_as_is_tobe = QtWidgets.QGroupBox()
        as_is_tobe_main_layout = QtWidgets.QVBoxLayout(grp_as_is_tobe)
        as_is_tobe_title = QtWidgets.QLabel("AS-IS → TO-BE 라벨링")
        as_is_tobe_main_layout.addWidget(as_is_tobe_title)
        
        self.as_is_tobe_container = QtWidgets.QWidget()
        self.as_is_tobe_container.setFixedHeight(200)
        self.as_is_tobe_layout = QtWidgets.QVBoxLayout(self.as_is_tobe_container)
        self.as_is_tobe_layout.setSpacing(5)
        self.as_is_tobe_layout.setContentsMargins(5, 5, 5, 5)
        self.as_is_tobe_container.setVisible(False)
        self.as_is_tobe_container.setMaximumHeight(0)
        
        as_is_tobe_main_layout.addWidget(self.as_is_tobe_container)
        grp_as_is_tobe.setSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        self.controls_layout.addWidget(grp_as_is_tobe)

        # Filter controls
        grp_filter = QtWidgets.QGroupBox("필터 / 탐색")
        grp_filter_layout = QtWidgets.QVBoxLayout(grp_filter)
        
        quick_filter_widget = self._create_quick_filters()
        grp_filter_layout.addWidget(quick_filter_widget)
        
        basic_filters_title = QtWidgets.QLabel("기본 필터")
        grp_filter_layout.addWidget(basic_filters_title)
        
        self.basic_filters_widget = QtWidgets.QWidget()
        fl = QtWidgets.QGridLayout(self.basic_filters_widget)
        fl.setSpacing(3)
        fl.setContentsMargins(5, 2, 5, 2)
        
        self.chk_unlabeled = QtWidgets.QCheckBox("라벨 없는 항목만")
        self.cmb_label_state = QtWidgets.QComboBox()
        self.cmb_label_state.addItems(["전체", "라벨됨", "라벨안됨"])
        self.cmb_label_value = QtWidgets.QComboBox()
        self.cmb_result_filter = QtWidgets.QComboBox()
        self.cmb_background_filter = QtWidgets.QComboBox()
        self.chk_bookmarks = QtWidgets.QCheckBox("북마크만")

        fl.addWidget(self.chk_unlabeled, 0, 0)
        fl.addWidget(self.cmb_label_state, 0, 1)
        fl.addWidget(QtWidgets.QLabel("라벨 값:"), 1, 0)
        fl.addWidget(self.cmb_label_value, 1, 1)
        fl.addWidget(QtWidgets.QLabel("기본결과:"), 2, 0)
        fl.addWidget(self.cmb_result_filter, 2, 1)
        # 배경결과 필터 (강조 표시)
        bg_label = QtWidgets.QLabel("🖼️ 배경:")
        bg_label.setStyleSheet("font-weight: bold; color: #2196f3;")
        fl.addWidget(bg_label, 3, 0)
        fl.addWidget(self.cmb_background_filter, 3, 1)
        fl.addWidget(self.chk_bookmarks, 4, 0)

        # 필터 적용 버튼 (강조 표시)
        self.btn_apply_filters = QtWidgets.QPushButton("🔍 필터 적용")
        self.btn_apply_filters.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-weight: bold;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
        """)
        self.btn_apply_filters.clicked.connect(self.apply_filters)
        fl.addWidget(self.btn_apply_filters, 5, 0, 1, 2)  # 2열에 걸쳐 표시
        
        # 필터 새로고침 버튼 추가
        self.btn_refresh_filters = QtWidgets.QPushButton("🔄 필터 새로고침")
        self.btn_refresh_filters.setToolTip("CSV 데이터에서 필터 옵션을 다시 로드합니다")
        self.btn_refresh_filters.clicked.connect(self._refresh_basic_filters)
        fl.addWidget(self.btn_refresh_filters, 6, 0, 1, 2)

        self.chk_show_overlay = QtWidgets.QCheckBox("JSON 오버레이 표시")
        self.chk_show_overlay.setChecked(self.show_overlay)
        fl.addWidget(self.chk_show_overlay, 7, 0)

        grp_filter_layout.addWidget(self.basic_filters_widget)
        grp_filter.setSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        self.controls_layout.addWidget(grp_filter)

        # Pred filter section
        grp_pred_filter = QtWidgets.QGroupBox()
        pred_filter_main_layout = QtWidgets.QVBoxLayout(grp_pred_filter)
        pred_filters_title = QtWidgets.QLabel("예측 결과 필터")
        pred_filter_main_layout.addWidget(pred_filters_title)
        
        self.pred_filters_container = QtWidgets.QWidget()
        pred_filter_layout = QtWidgets.QVBoxLayout(self.pred_filters_container)
        
        self.btn_clear_pred_filters = QtWidgets.QPushButton("모든 필터 해제")
        pred_filter_layout.addWidget(self.btn_clear_pred_filters)
        
        self.pred_filter_scroll = QtWidgets.QScrollArea()
        self.pred_filter_scroll.setMaximumHeight(200)
        self.pred_filter_widget = QtWidgets.QWidget()
        self.pred_filter_checkboxes_layout = QtWidgets.QVBoxLayout(self.pred_filter_widget)
        self.pred_filter_scroll.setWidget(self.pred_filter_widget)
        self.pred_filter_scroll.setWidgetResizable(True)
        pred_filter_layout.addWidget(self.pred_filter_scroll)
        
        pred_filter_main_layout.addWidget(self.pred_filters_container)
        grp_pred_filter.setSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        self.controls_layout.addWidget(grp_pred_filter)

        # Navigation
        nav_widget = QtWidgets.QWidget()
        nav_layout = QtWidgets.QVBoxLayout(nav_widget)
        
        nav_buttons = QtWidgets.QHBoxLayout()
        self.btn_prev = QtWidgets.QPushButton("이전")
        self.btn_next = QtWidgets.QPushButton("다음")
        nav_buttons.addWidget(self.btn_prev)
        nav_buttons.addWidget(self.btn_next)
        nav_layout.addLayout(nav_buttons)
        
        self.chk_auto_advance = QtWidgets.QCheckBox("리뷰 완료 후 자동 다음 이동")
        self.chk_auto_advance.setChecked(self.auto_advance_enabled)
        nav_layout.addWidget(self.chk_auto_advance)
        
        nav_widget.setSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        self.controls_layout.addWidget(nav_widget)
        
        print("✅ 컨트롤 내용 생성 완료")
    
    def _connect_all_signals(self) -> None:
        """모든 시그널 한번에 연결"""
        print("🔗 시그널 연결 시작...")
        
        # Table signals
        self.table.cellDoubleClicked.connect(self.on_table_double_click)
        self.table.cellClicked.connect(self.on_table_click)
        self.table.itemSelectionChanged.connect(self._on_table_selection_changed)
        self.table.verticalScrollBar().valueChanged.connect(self._on_table_scroll)
        
        # Navigation buttons
        self.btn_prev.clicked.connect(self.on_prev)
        self.btn_next.clicked.connect(self.on_next)
        
        # Bookmark
        self.btn_toggle_bookmark.clicked.connect(self.toggle_bookmark)
        
        # Auto-advance
        self.chk_auto_advance.toggled.connect(self.on_auto_advance_toggled)
        
        # Overlay toggle
        self.chk_show_overlay.toggled.connect(self._on_overlay_toggled)
        
        # Filter controls
        self.chk_unlabeled.toggled.connect(self.apply_filters)
        self.cmb_label_state.currentTextChanged.connect(self.apply_filters)
        self.cmb_label_value.currentTextChanged.connect(self.apply_filters)
        self.cmb_result_filter.currentTextChanged.connect(self.apply_filters)
        self.cmb_background_filter.currentTextChanged.connect(self.apply_filters)
        self.chk_bookmarks.toggled.connect(self.apply_filters)
        
        # Pred filter
        self.btn_clear_pred_filters.clicked.connect(self.clear_pred_filters)
        
        print("✅ 시그널 연결 완료")

    def _create_minimal_ui(self):
        """최소한의 UI 생성 - 디버깅용"""
        print("🚨 최소 UI 생성 시작...")
        try:
            # 기본 레이아웃 생성
            central_widget = QtWidgets.QWidget()
            self.setCentralWidget(central_widget)
            layout = QtWidgets.QVBoxLayout(central_widget)

            # 간단한 라벨 추가
            label = QtWidgets.QLabel("UI 테스트 - 최소 모드")
            label.setAlignment(QtCore.Qt.AlignCenter)
            label.setStyleSheet("font-size: 20px; color: red; font-weight: bold;")
            layout.addWidget(label)

            # 상태 정보 표시
            info_label = QtWidgets.QLabel(f"CSV: {self.csv_path}\n이미지: {self.images_base}\nJSON: {self.json_base}")
            info_label.setStyleSheet("font-size: 12px; color: blue;")
            layout.addWidget(info_label)

            # 버튼 추가
            test_btn = QtWidgets.QPushButton("테스트 버튼")
            test_btn.clicked.connect(lambda: print("테스트 버튼 클릭됨"))
            layout.addWidget(test_btn)

            print("✅ 최소 UI 생성 완료")
        except Exception as e:
            print(f"❌ 최소 UI 생성 오류: {e}")
            import traceback
            traceback.print_exc()


    def _create_menus(self):
        """메뉴 생성"""
        # File menu
        file_menu = self.menuBar().addMenu("파일")
        act_reload = file_menu.addAction("데이터 새로고침")
        act_export = file_menu.addAction("라벨을 엑셀로 내보내기")
        file_menu.addSeparator()
        act_save_session = file_menu.addAction("세션 상태 저장")
        act_load_session = file_menu.addAction("세션 상태 복원")
        file_menu.addSeparator()
        act_quit = file_menu.addAction("종료")

        act_quit.triggered.connect(self.close)
        act_reload.triggered.connect(self.load_csv_data)
        act_export.triggered.connect(self.on_export_labels)
        act_save_session.triggered.connect(self.save_session_state)
        act_load_session.triggered.connect(self.restore_session_state)

        # Memory management menu
        memory_menu = self.menuBar().addMenu("메모리")
        act_clear_cache = memory_menu.addAction("이미지 캐시 삭제")
        act_clear_cache.triggered.connect(self._clear_image_cache)
        act_memory_info = memory_menu.addAction("메모리 정보")
        act_memory_info.triggered.connect(self._show_memory_info)
        act_force_cleanup = memory_menu.addAction("메모리 정리")
        act_force_cleanup.triggered.connect(self._force_memory_cleanup)
        act_performance_stats = memory_menu.addAction("성능 통계")
        act_performance_stats.triggered.connect(self._show_performance_stats)

        # Image matching debugging
        act_image_debug = memory_menu.addAction("이미지 매칭 디버그")
        act_image_debug.triggered.connect(self._debug_image_matching)

        # Manual image path correction
        act_fix_image_path = memory_menu.addAction("이미지 경로 수동 수정")
        act_fix_image_path.triggered.connect(self._fix_image_path_manually)

        # Quick fix for wrong matches
        act_quick_fix = memory_menu.addAction("잘못된 매칭 빠른 수정")
        act_quick_fix.triggered.connect(self._quick_fix_wrong_match)

    def _apply_theme(self):
        """모던 테마 적용"""
        pass  # 현재는 빈 함수로 두고 나중에 구현


    def _on_overlay_toggled(self, checked: bool):
        """오버레이 표시 토글"""
        print(f"🔄 오버레이 토글: {'켜짐' if checked else '꺼짐'}")
        self.show_overlay = checked

        # 현재 표시된 이미지가 있다면 다시 로드하여 오버레이 적용/해제
        if hasattr(self, 'current_idx') and self.df is not None:
            if self.current_idx < len(self.filtered_indices):
                row_idx = self.filtered_indices[self.current_idx]
                print(f"📸 이미지 다시 로드 중... (행: {row_idx})")
                self._load_image_for_row(row_idx)
                print("✅ 이미지 다시 로드 완료")
            else:
                print("⚠️ 유효한 이미지 인덱스가 없음")
        else:
            print("⚠️ 현재 인덱스나 데이터프레임이 없음")

    def _create_status_widgets(self) -> None:
        """Create status bar widgets for real-time information display"""
        # Create status widgets
        self.lbl_save_status = QtWidgets.QLabel("저장 상태: 대기")
        self.lbl_save_status.setStyleSheet("color: #666; font-size: 11px; padding: 2px 8px;")
        
        self.lbl_memory_status = QtWidgets.QLabel("메모리: --")
        self.lbl_memory_status.setStyleSheet("color: #666; font-size: 11px; padding: 2px 8px;")
        
        self.lbl_progress_status = QtWidgets.QLabel("진행률: --")
        self.lbl_progress_status.setStyleSheet("color: #666; font-size: 11px; padding: 2px 8px;")
        
        self.lbl_current_position = QtWidgets.QLabel("위치: --")
        self.lbl_current_position.setStyleSheet("color: #666; font-size: 11px; padding: 2px 8px;")
        
        # Add separators
        separator1 = QtWidgets.QLabel("|")
        separator1.setStyleSheet("color: #ccc; font-size: 11px; padding: 2px 4px;")
        separator2 = QtWidgets.QLabel("|")
        separator2.setStyleSheet("color: #ccc; font-size: 11px; padding: 2px 4px;")
        separator3 = QtWidgets.QLabel("|")
        separator3.setStyleSheet("color: #ccc; font-size: 11px; padding: 2px 4px;")
        
        # Add widgets to status bar
        self.status.addPermanentWidget(self.lbl_save_status)
        self.status.addPermanentWidget(separator1)
        self.status.addPermanentWidget(self.lbl_memory_status)
        self.status.addPermanentWidget(separator2)
        self.status.addPermanentWidget(self.lbl_progress_status)
        self.status.addPermanentWidget(separator3)
        self.status.addPermanentWidget(self.lbl_current_position)
        
        # Start timer for periodic updates
        self._status_update_timer = QtCore.QTimer()
        self._status_update_timer.timeout.connect(self._update_status_widgets)
        self._status_update_timer.start(2000)  # Update every 2 seconds

    def _update_status_widgets(self) -> None:
        """Update status bar widgets with real-time information"""
        try:
            # Update memory status
            memory_mb = get_memory_usage()
            self.lbl_memory_status.setText(f"메모리: {memory_mb:.1f}MB")
            
            # Update progress status
            if self.df is not None:
                total_rows = len(self.df)
                labeled_rows = len(self.df[~(self.df[self.active_label_col].isna() | (self.df[self.active_label_col] == ""))])
                progress = (labeled_rows / total_rows * 100) if total_rows > 0 else 0
                self.lbl_progress_status.setText(f"진행률: {labeled_rows:,}/{total_rows:,} ({progress:.1f}%)")
            else:
                self.lbl_progress_status.setText("진행률: --")
            
            # Update current position
            if self.df is not None and self.filtered_indices:
                current_pos = self.current_idx + 1 if self.current_idx < len(self.filtered_indices) else 0
                total_filtered = len(self.filtered_indices)
                self.lbl_current_position.setText(f"위치: {current_pos}/{total_filtered}")
            else:
                self.lbl_current_position.setText("위치: --")
                
        except Exception as e:
            print(f"상태 업데이트 오류: {e}")

    def _update_save_status(self, status: str, color: str = "#666") -> None:
        """Update save status with custom color"""
        self.lbl_save_status.setText(f"저장 상태: {status}")
        self.lbl_save_status.setStyleSheet(f"color: {color}; font-size: 11px; padding: 2px 8px;")

    def _create_toolbar(self) -> None:
        """Create modern toolbar with frequently used actions"""
        toolbar = self.addToolBar("도구")
        toolbar.setMovable(False)

        
        # Navigation actions
        prev_action = toolbar.addAction("⬅️ 이전 (←/A)")
        prev_action.triggered.connect(self.on_prev)
        prev_action.setShortcut(QtGui.QKeySequence(QtCore.Qt.Key_Left))
        
        next_action = toolbar.addAction("➡️ 다음 (→/D/Space)")
        next_action.triggered.connect(self.on_next)
        next_action.setShortcut(QtGui.QKeySequence(QtCore.Qt.Key_Right))
        
        toolbar.addSeparator()
        
        # Labeling actions
        bookmark_action = toolbar.addAction("🔖 북마크 (B)")
        bookmark_action.triggered.connect(self.toggle_bookmark)
        bookmark_action.setShortcut("B")
        
        toolbar.addSeparator()
        
        # View actions
        stats_action = toolbar.addAction("📊 통계")
        stats_action.triggered.connect(self._show_performance_stats)
        
        memory_action = toolbar.addAction("💾 메모리")
        memory_action.triggered.connect(self._show_memory_info)
        
        toolbar.addSeparator()
        
        # Settings
        reload_action = toolbar.addAction("🔄 새로고침")
        reload_action.triggered.connect(self.load_csv_data)

    def _create_progress_dashboard(self) -> QtWidgets.QWidget:
        """Create progress dashboard with statistics"""
        dashboard = QtWidgets.QWidget()

        dashboard.setMaximumHeight(80)
        
        layout = QtWidgets.QVBoxLayout(dashboard)
        layout.setSpacing(4)
        layout.setContentsMargins(12, 8, 12, 8)
        
        # Title
        title_label = QtWidgets.QLabel("📊 진행 현황")
        title_label.setStyleSheet("font-size: 14px; font-weight: 600; color: #1976d2;")
        layout.addWidget(title_label)
        
        # Progress bar and stats in horizontal layout
        progress_layout = QtWidgets.QHBoxLayout()
        
        # Progress bar
        self.progress_bar = QtWidgets.QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        progress_layout.addWidget(self.progress_bar, 2)
        
        # Stats label
        self.stats_label = QtWidgets.QLabel("로딩 중...")
        self.stats_label.setStyleSheet("font-size: 12px; color: #424242; min-width: 200px;")
        progress_layout.addWidget(self.stats_label, 1)
        
        layout.addLayout(progress_layout)
        
        return dashboard

    def _create_modern_label_button(self, text: str, shortcut: int) -> QtWidgets.QPushButton:
        """Create a modern styled label button with shortcut hint"""
        # Choose color based on label type
        color_map = {
            'OK': '#4caf50',
            'NG': '#f44336',
            'NG_BUT': '#ff9800',
            '보류': '#9c27b0',
            'SR-이물->OK': '#2196f3',
            'SR-이물->도금-찍힘': '#795548'
        }
        
        color = color_map.get(text, '#757575')  # Default gray
        
        btn = QtWidgets.QPushButton(f"{text} ({shortcut})")
        
        # Add tooltip with shortcut info
        btn.setToolTip(f"단축키: {shortcut}")
        
        return btn



    def _create_quick_filters(self) -> QtWidgets.QWidget:
        """Create quick filter buttons for common operations"""
        widget = QtWidgets.QWidget()
        
        layout = QtWidgets.QHBoxLayout(widget)
        layout.setSpacing(6)
        layout.setContentsMargins(4, 4, 4, 4)
        
        title_label = QtWidgets.QLabel("빠른 필터:")
        layout.addWidget(title_label)
        
        # Quick filter buttons
        quick_filters = [
            ("라벨없음", self._filter_unlabeled, "#ff9800"),
            ("OK만", self._filter_ok_only, "#4caf50"),
            ("NG만", self._filter_ng_only, "#f44336"), 
            ("배경SR", self._filter_sr_background, "#2196f3"),
            ("북마크", self._filter_bookmarks, "#9c27b0"),
            ("전체보기", self._show_all, "#757575")
        ]
        
        for text, func, color in quick_filters:
            btn = QtWidgets.QPushButton(text)
            btn.clicked.connect(func)
            layout.addWidget(btn)
        
        layout.addStretch()
        return widget



    def _filter_unlabeled(self) -> None:
        """Quick filter: show only unlabeled items"""
        self.cmb_label_state.setCurrentText("라벨안됨")
        self.chk_bookmarks.setChecked(False)

    def _filter_ok_only(self) -> None:
        """Quick filter: show only OK items"""
        self.cmb_label_state.setCurrentText("라벨됨")
        self.cmb_label_value.setCurrentText("OK")
        self.chk_bookmarks.setChecked(False)

    def _filter_ng_only(self) -> None:
        """Quick filter: show only NG items"""
        self.cmb_label_state.setCurrentText("라벨됨")
        self.cmb_label_value.setCurrentText("NG")
        self.chk_bookmarks.setChecked(False)





    def _filter_sr_background(self) -> None:
        """Quick filter: show only items with SR background"""
        # Clear other filters first
        self.cmb_label_state.setCurrentText("전체")
        self.cmb_label_value.setCurrentText("전체")
        self.chk_bookmarks.setChecked(False)

        # Clear existing pred filters
        for checkbox in self.pred_filter_checkboxes.values():
            checkbox.setChecked(False)
        self.selected_pred_filters.clear()

        # Set background filter to SR
        if "Background_result" in self.df.columns:
            self.cmb_background_filter.setCurrentText("SR")

        self.apply_filters()

    def _filter_bookmarks(self) -> None:
        """Quick filter: show only bookmarked items"""
        self.chk_bookmarks.setChecked(True)

    def _show_all(self) -> None:
        """Quick filter: show all items"""
        self.cmb_label_state.setCurrentText("전체")
        self.cmb_label_value.setCurrentText("전체")
        self.chk_bookmarks.setChecked(False)
        # Clear pred filters
        for checkbox in self.pred_filter_checkboxes.values():
            checkbox.setChecked(False)
        self.selected_pred_filters.clear()

    def _update_progress_dashboard(self) -> None:
        """Update progress dashboard with current statistics"""
        if self.df is None or not hasattr(self, 'progress_bar'):
            return
            
        try:
            total_items = len(self.df)
            if total_items == 0:
                return
                
            # Count labeled items
            labeled_count = 0
            if self.active_label_col in self.df.columns:
                labeled_mask = ~pd.isna(self.df[self.active_label_col]) & (self.df[self.active_label_col] != "")
                labeled_count = labeled_mask.sum()
            
            # Calculate progress percentage
            progress_percent = (labeled_count / total_items) * 100 if total_items > 0 else 0
            
            # Update progress bar with null checks
            if hasattr(self, 'progress_bar') and self.progress_bar is not None:
                try:
                    self.progress_bar.setValue(int(progress_percent))
                    self.progress_bar.setFormat(f"{progress_percent:.1f}% ({labeled_count:,}/{total_items:,})")
                except RuntimeError:
                    pass  # Progress bar widget has been deleted
            
            # Count different label types
            label_stats = {}
            if self.active_label_col in self.df.columns:
                label_counts = self.df[self.active_label_col].value_counts()
                for label, count in label_counts.items():
                    if pd.notna(label) and label != "":
                        label_stats[str(label)] = int(count)
            
            # Create stats text
            remaining = total_items - labeled_count
            filtered_total = len(self.filtered_indices) if self.filtered_indices else 0
            
            stats_text = f"✅ {labeled_count:,} 완료 | ⏳ {remaining:,} 남음 | 🎯 {progress_percent:.1f}%"
            if filtered_total != total_items:
                stats_text += f" | 🔍 필터됨: {filtered_total:,}/{total_items:,}"
            
            if hasattr(self, 'stats_label') and self.stats_label is not None:
                try:
                    self.stats_label.setText(stats_text)
                except RuntimeError:
                    pass  # Stats label widget has been deleted
            
        except Exception as e:
            print(f"Progress dashboard update error: {e}")

    def _create_collapsible_section_button(self, title: str, is_expanded: bool = True) -> QtWidgets.QPushButton:
        """Create a styled collapsible section toggle button"""
        arrow = "▼" if is_expanded else "▶"
        btn = QtWidgets.QPushButton(f"{arrow} {title}")
        btn.setStyleSheet("""
            QPushButton {
                text-align: left;
                font-weight: 700;
                font-size: 14px;
                background: qlineargradient(y1:0, y2:1, stop:0 #4caf50, stop:1 #388e3c);
                border: 2px solid #2e7d32;
                border-radius: 8px;
                padding: 12px 16px;
                margin: 3px 0px;
                color: white;
                min-height: 25px;
            }
            QPushButton:hover {
                background: qlineargradient(y1:0, y2:1, stop:0 #66bb6a, stop:1 #4caf50);
                border-color: #1b5e20;
                font-size: 15px;
            }
            QPushButton:pressed {
                background: qlineargradient(y1:0, y2:1, stop:0 #2e7d32, stop:1 #1b5e20);
                border: 3px solid #1b5e20;
                font-weight: 800;
            }
        """)
        return btn

    def _connect_shortcuts(self) -> None:
        """Connect keyboard shortcuts for labeling"""
        shortcuts = [
            # Labeling shortcuts
            ("1", lambda: self._assign_by_index(0)),
            ("2", lambda: self._assign_by_index(1)),
            ("3", lambda: self._assign_by_index(2)),
            ("4", lambda: self._assign_by_index(3)),
            ("5", lambda: self._assign_by_index(4)),
            ("6", lambda: self._assign_by_index(5)),
            ("7", self._toggle_as_is_tobe_mode),
            ("Return", self._apply_all_tobe_selections),  # Enter key for apply all
            
            # Navigation shortcuts - only secondary keys (primary keys handled by QAction)
            ("Up", self.on_prev),
            ("Down", self.on_next),
            ("a", self.on_prev),
            ("d", self.on_next),
            ("A", self.on_prev),
            ("D", self.on_next),
            ("Space", self.on_next),
        ]
        
        for key, func in shortcuts:
            shortcut = QtGui.QShortcut(QtGui.QKeySequence(key), self)
            shortcut.activated.connect(func)
            shortcut.setContext(QtCore.Qt.ApplicationShortcut)  # Make shortcuts work globally

    def keyPressEvent(self, event: QtGui.QKeyEvent) -> None:
        """Handle key press events for navigation and labeling"""
        key = event.key()
        modifiers = event.modifiers()
        
        # Navigation keys
        # Handle navigation keys directly since QAction shortcuts might have focus issues
        if key == QtCore.Qt.Key_Left:
            print("Left key pressed in keyPressEvent")
            self.on_prev()
            event.accept()
            return
        elif key == QtCore.Qt.Key_Right:
            print("Right key pressed in keyPressEvent")
            self.on_next()
            event.accept()
            return
        
        # Number keys for labeling
        if key == QtCore.Qt.Key_1:
            self._assign_by_index(0)
            event.accept()
            return
        elif key == QtCore.Qt.Key_2:
            self._assign_by_index(1)
            event.accept()
            return
        elif key == QtCore.Qt.Key_3:
            self._assign_by_index(2)
            event.accept()
            return
        elif key == QtCore.Qt.Key_4:
            self._assign_by_index(3)
            event.accept()
            return
        elif key == QtCore.Qt.Key_5:
            self._assign_by_index(4)
            event.accept()
            return
        elif key == QtCore.Qt.Key_6:
            self._assign_by_index(5)
            event.accept()
            return
        elif key == QtCore.Qt.Key_7:
            self._toggle_as_is_tobe_mode()
            event.accept()
            return
        elif key == QtCore.Qt.Key_Return or key == QtCore.Qt.Key_Enter:
            self._apply_all_tobe_selections()
            event.accept()
            return
        
        # Call parent class for other keys
        super().keyPressEvent(event)

    def _get_smart_visible_indices(self) -> List[int]:
        """Get smart visible indices ensuring current row is always visible"""
        if not self.filtered_indices:
            return []
        
        # Configuration
        window_size = 100  # Number of rows to show around current position
        buffer_size = 50   # Buffer size for smooth scrolling
        
        # Calculate the range to show
        current_pos = self.current_idx
        total_items = len(self.filtered_indices)
        
        # Calculate start and end positions to center current row
        start_pos = max(0, current_pos - buffer_size)
        end_pos = min(total_items, current_pos + buffer_size + 1)
        
        # Adjust if we're near the beginning or end
        if end_pos - start_pos < window_size:
            if start_pos == 0:
                # Near beginning, extend to the right
                end_pos = min(total_items, start_pos + window_size)
            elif end_pos == total_items:
                # Near end, extend to the left
                start_pos = max(0, end_pos - window_size)
        
        # Get the visible indices
        visible_indices = self.filtered_indices[start_pos:end_pos]
        
        # Store the mapping for later use
        self._table_start_pos = start_pos
        self._table_end_pos = end_pos
        
        # Store current row position for scroll restoration
        self._current_row_in_window = current_pos - start_pos
        print(f"_get_smart_visible_indices: current_pos={current_pos}, start_pos={start_pos}, _current_row_in_window={self._current_row_in_window}")
        
        return visible_indices

    def load_csv_data(self) -> None:
        """Load the CSV data and set up the interface - optimized for large files"""
        print(f"📄 CSV 데이터 로드 시작: {self.csv_path}")
        if not os.path.exists(self.csv_path):
            print(f"❌ CSV 파일 없음: {self.csv_path}")
            QtWidgets.QMessageBox.warning(self, "오류", f"CSV 파일을 찾을 수 없음: {self.csv_path}")
            return

        try:
            print("⏳ CSV 파일 로드 중...")
            # Show loading progress for large files
            self.status.showMessage("📄 대용량 CSV 파일 로드 중...")
            QtWidgets.QApplication.processEvents()  # Allow UI to update

            # Load CSV with optimized settings for large files
            print("📊 pandas로 CSV 읽는 중...")
            self.df = pd.read_csv(
                self.csv_path,
                low_memory=False,  # Read entire file at once for consistency
                dtype_backend='numpy_nullable',  # Use nullable dtypes for better memory usage
                engine='c'  # Use C engine for better performance
            )
            
            # Force garbage collection after loading
            force_garbage_collection()
            
            self.status.showMessage(f"{self.csv_path}에서 {len(self.df):,}개 행 로드됨")
            
            # Set up JSON path
            self.json_path = default_json_path(self.csv_path.replace('.csv', '.xlsx'))
            
            # Manual_Label should NOT exist in CSV - it's for user labeling only
            if self.active_label_col in self.df.columns:
                print(f"⚠️ 경고: CSV 파일에 '{self.active_label_col}' 컬럼이 이미 존재합니다.")
                print("   이 컬럼은 사용자가 직접 채워야 하는 컬럼입니다.")
                print("   기존 데이터를 백업하고 컬럼을 제거합니다.")

                # Remove the existing Manual_Label column since it shouldn't be in CSV
                self.df = self.df.drop(columns=[self.active_label_col])
                print(f"✅ 기존 '{self.active_label_col}' 컬럼을 제거했습니다.")

            # Create fresh Manual_Label column for user labeling
                self.df[self.active_label_col] = ""
                ensure_object_dtype(self.df, self.active_label_col)
            print(f"✅ 새로운 빈 '{self.active_label_col}' 컬럼을 생성했습니다.")
            print("   이 컬럼은 사용자가 직접 라벨링한 결과를 저장합니다.")

            # Extract details from JSON files
            if "detail" not in self.df.columns:
                self.df["detail"] = ""
                ensure_object_dtype(self.df, "detail")

            if "Result_path" in self.df.columns:
                for idx, row in self.df.iterrows():
                    result_path = row["Result_path"]
                    if pd.notna(result_path) and str(result_path).strip():
                        # JSON 파일 경로 추출 (파일명에서 .json 확장자 추가)
                        json_file_path = str(result_path).strip()
                        if not json_file_path.endswith('.json'):
                            json_file_path += '.json'

                        # JSON 파일에서 detail 정보 추출
                        details = extract_detail_from_json(json_file_path)
                        if details:
                            self.df.at[idx, "detail"] = "; ".join(details)
            
            # Manual_Label은 사용자가 직접 채워야 하는 컬럼이므로 기존 데이터를 로드하지 않음
            # 다른 컬럼들의 기존 데이터는 로드할 수 있음 (필요시)
            print(f"ℹ️ '{self.active_label_col}' 컬럼은 사용자가 직접 라벨링해야 하는 빈 컬럼입니다.")
            print("   기존 JSON 데이터를 로드하지 않습니다.")

            # 다른 컬럼들의 기존 데이터 로드 (skip_existing_labels 설정에 따라)
            other_columns = [col for col in self.df.columns if col != self.active_label_col]
            if not getattr(self, 'skip_existing_labels', False) and other_columns:
                merge_json_into_df(self.json_path, self.df, other_columns)
                print("✅ 다른 컬럼들의 기존 데이터를 로드했습니다.")
            elif other_columns:
                print("⏭️ 새로운 시작 - 다른 컬럼들의 기존 데이터도 무시합니다.")
            
            # Extract TO-BE choices from Unique_seg_result
            self.compute_tobe_choices()
            self.compute_pred_filter_choices()
            self.setup_result_filter()
            self.setup_background_filter()

            # Force refresh of basic filters to ensure they display properly
            QtCore.QTimer.singleShot(100, self._refresh_basic_filters)

            # Debug: Check loaded data
            print(f"Loaded DataFrame shape: {self.df.shape}")
            print(f"Columns: {list(self.df.columns)}")
            print(f"Sample data (first 3 rows):")
            print(self.df.head(3))
            print(f"Active label column '{self.active_label_col}' values:")
            if self.active_label_col in self.df.columns:
                print(self.df[self.active_label_col].value_counts().head())
            
            # Set up UI with progress updates
            self.status.showMessage("UI 초기화 중...")
            QtWidgets.QApplication.processEvents()
            
            # UI updates with better error handling
            try:
                self.refresh_label_controls()
                print("Label controls refreshed successfully")
            except Exception as e:
                print(f"Error in refresh_label_controls: {e}")

            try:
                self.refresh_pred_filter_controls()
                print("Pred filter controls refreshed successfully")
            except Exception as e:
                print(f"Error in refresh_pred_filter_controls: {e}")

            try:
                self.refresh_as_is_tobe_panel()
                print("AS-IS/TO-BE panel refreshed successfully")
            except Exception as e:
                print(f"Error in refresh_as_is_tobe_panel: {e}")

            try:
                self._initialize_after_load()
                print("Post-load initialization completed successfully")
                
                # Force refresh view and table after initialization
                if hasattr(self, 'filtered_indices') and self.filtered_indices:
                    print(f"Forcing refresh_view and refresh_table with {len(self.filtered_indices)} filtered items")
                    self.refresh_view()
                    self.refresh_table()
                else:
                    print("No filtered indices available for refresh")
                    
            except Exception as e:
                print(f"Error in _initialize_after_load: {e}")
            
            self.status.showMessage(f"로드 완료: {len(self.df):,}개 행 준비됨", 2000)
            
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "오류", f"CSV 로드 실패: {str(e)}")

    def _initialize_after_load(self) -> None:
        """Initialize UI elements after CSV data is loaded"""
        print("📊 _initialize_after_load 시작")
        try:
            # Apply filters first
            print("🔍 필터 적용 시작")
            self.apply_filters()
            print("✅ 필터 적용 완료")
            
            # Initialize image status bar if we have data
            if self.df is not None and self.filtered_indices and self.current_idx < len(self.filtered_indices):
                row_idx = self.filtered_indices[self.current_idx]
                self._update_image_status_bar(row_idx)
                print(f"Initialized image status bar for row {row_idx}")
            else:
                self._safe_set_text(self.image_status_bar, "데이터가 로드되지 않음")
                print("No data available for image status bar initialization")
                
        except Exception as e:
            print(f"Error in _initialize_after_load: {e}")
            import traceback
            traceback.print_exc()

    def compute_tobe_choices(self) -> None:
        """Extract unique values from Unique_seg_result for TO-BE choices"""
        if self.df is None or "Unique_seg_result" not in self.df.columns:
            return

        choices = set()
        for val in self.df["Unique_seg_result"].dropna():
            pred_list = parse_pred_list(val)
            choices.update(pred_list)
        
        # Combine with standard choices
        all_choices = set(self.label_choices + list(choices))
        self.tobe_choices = sorted(all_choices)

    def compute_pred_filter_choices(self) -> None:
        """Extract unique Unique_seg_result values for filtering"""
        if self.df is None or "Unique_seg_result" not in self.df.columns:
            return

        choices = set()
        for val in self.df["Unique_seg_result"].dropna():
            pred_list = parse_pred_list(val)
            choices.update(pred_list)
        
        self.pred_filter_choices = sorted(choices)



    def setup_result_filter(self) -> None:
        """Set up Result filter dropdown"""
        if not hasattr(self, 'cmb_result_filter') or self.cmb_result_filter is None:
            print("⚠️ cmb_result_filter가 생성되지 않음")
            return

        print(f"🔄 Result 필터 설정 시작...")
        self.cmb_result_filter.clear()
        self.cmb_result_filter.addItem("전체")

        if self.df is None or "Result" not in self.df.columns:
            print("ℹ️ Result 컬럼이 없음")
            return

        # Get unique Result values, filtering out NaN and empty strings
        result_series = self.df["Result"].dropna()
        result_series = result_series[result_series.astype(str).str.strip() != ""]
        unique_results = sorted(result_series.unique())
        
        if unique_results:
            self.cmb_result_filter.addItems([str(x) for x in unique_results])
            print(f"✅ Result 필터 설정됨: {len(unique_results)}개 값 - {unique_results}")
        else:
            print("⚠️ Result 컬럼에 유효한 데이터 없음")

    def setup_background_filter(self) -> None:
        """Set up Background_result filter dropdown"""
        if not hasattr(self, 'cmb_background_filter') or self.cmb_background_filter is None:
            print("⚠️ cmb_background_filter가 생성되지 않음")
            return

        print(f"🔄 Background_result 필터 설정 시작...")
        self.cmb_background_filter.clear()
        self.cmb_background_filter.addItem("전체")

        if self.df is None or "Background_result" not in self.df.columns:
            print("ℹ️ Background_result 컬럼이 없음")
            return

        # Get unique Background_result values, filtering out NaN and empty strings
        bg_series = self.df["Background_result"].dropna()
        bg_series = bg_series[bg_series.astype(str).str.strip() != ""]
        unique_bg_results = sorted(bg_series.unique())
        
        if unique_bg_results:
            self.cmb_background_filter.addItems([str(x) for x in unique_bg_results])
            print(f"✅ Background_result 필터 설정됨: {len(unique_bg_results)}개 값 - {unique_bg_results}")
        else:
            print("⚠️ Background_result 컬럼에 유효한 데이터 없음")

    def _refresh_basic_filters(self) -> None:
        """Refresh all basic filter dropdowns to ensure they display properly"""
        try:
            print("🔄 기본 필터 드롭다운 새로고침...")

            # Refresh label value filter
            if hasattr(self, 'cmb_label_value') and self.cmb_label_value:
                self.cmb_label_value.clear()
                self.cmb_label_value.addItem("전체")
                if self.df is not None and self.active_label_col in self.df.columns:
                    # Get unique labels, excluding empty/NaN values
                    label_series = self.df[self.active_label_col]
                    unique_labels = sorted(label_series.dropna().unique())
                    # Only add non-empty labels
                    for label in unique_labels:
                        if str(label).strip():
                            self.cmb_label_value.addItem(str(label))
                    print(f"✅ 라벨 값 필터 설정됨: {len(unique_labels)}개 값")

            # Refresh result filter
            print("🔄 Result 필터 새로고침 중...")
            self.setup_result_filter()

            # Refresh background filter  
            print("🔄 Background_result 필터 새로고침 중...")
            self.setup_background_filter()

            # Ensure UI updates
            QtWidgets.QApplication.processEvents()

            print("✅ 기본 필터 드롭다운 새로고침 완료")

        except Exception as e:
            print(f"❌ 기본 필터 드롭다운 새로고침 오류: {e}")

    def refresh_pred_filter_controls(self) -> None:
        """Update Unique_seg_result filter checkboxes"""
        # Check if layout is valid
        if not hasattr(self, 'pred_filter_checkboxes_layout') or self.pred_filter_checkboxes_layout is None:
            print("Warning: pred_filter_checkboxes_layout not initialized")
            return

        try:
            # Clear existing checkboxes safely
            while self.pred_filter_checkboxes_layout.count() > 0:
                item = self.pred_filter_checkboxes_layout.takeAt(0)
                if item and item.widget():
                    item.widget().setParent(None)
                    item.widget().deleteLater()
        except Exception as e:
            print(f"Error clearing pred_filter_checkboxes_layout: {e}")
            return
        
        self.pred_filter_checkboxes.clear()
        
        # Create checkboxes for each unique prediction value (no categorization)
        for choice in sorted(self.pred_filter_choices):
            checkbox = QtWidgets.QCheckBox(choice)
            checkbox.toggled.connect(self.on_pred_filter_changed)
            self.pred_filter_checkboxes[choice] = checkbox
            self.pred_filter_checkboxes_layout.addWidget(checkbox)

        # Add stretch at the end
        self.pred_filter_checkboxes_layout.addStretch()

    def on_pred_filter_changed(self):
        """Handle pred filter checkbox changes"""
        self.selected_pred_filters.clear()
        for choice, checkbox in self.pred_filter_checkboxes.items():
            if checkbox.isChecked():
                self.selected_pred_filters.add(choice)
        self.apply_filters()

    def clear_pred_filters(self):
        """Clear all pred filter selections"""
        for checkbox in self.pred_filter_checkboxes.values():
            checkbox.setChecked(False)
        self.selected_pred_filters.clear()
        self.apply_filters()

    def refresh_label_controls(self) -> None:
        """Update the labeling button controls"""
        # Check if layout is valid
        if not hasattr(self, 'choice_buttons_layout') or self.choice_buttons_layout is None:
            print("Warning: choice_buttons_layout not initialized")
            return

        try:
            # Clear existing buttons safely
            while self.choice_buttons_layout.count() > 0:
                item = self.choice_buttons_layout.takeAt(0)
                if item and item.widget():
                    item.widget().setParent(None)
                    item.widget().deleteLater()
        except Exception as e:
            print(f"Error clearing choice_buttons_layout: {e}")
            return
        
        # Create buttons for label choices - vertical layout
        for i, choice in enumerate(self.label_choices):
            btn = self._create_modern_label_button(choice, i+1)
            btn.clicked.connect(lambda _, idx=i: self._assign_by_index(idx))
            self.choice_buttons_layout.addWidget(btn)
        
        # Add AS-IS/TO-BE mode toggle button (7번)
        btn_tobe_mode = self._create_modern_label_button("AS-IS/TO-BE", 7)
        btn_tobe_mode.clicked.connect(self._toggle_as_is_tobe_mode)
        
        # Set button style based on mode
        if self.as_is_tobe_mode:
            btn_tobe_mode.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: 2px solid #45a049;
                    border-radius: 4px;
                    padding: 8px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
                QPushButton:pressed {
                    background-color: #3d8b40;
                }
            """)
        else:
            btn_tobe_mode.setStyleSheet("""
                QPushButton {
                    background-color: #f0f0f0;
                    color: #333;
                    border: 2px solid #ddd;
                    border-radius: 4px;
                    padding: 8px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #e0e0e0;
                }
                QPushButton:pressed {
                    background-color: #d0d0d0;
                }
            """)
        
        # Add the button to the layout (next row after existing buttons)
        next_row = (len(self.label_choices) + 2) // 3
        self.choice_buttons_layout.addWidget(btn_tobe_mode, next_row, 0, 1, 3)  # Span all 3 columns
        
        # Adjust container size based on number of buttons (including AS-IS/TO-BE button)
        num_rows = ((len(self.label_choices) + 2) // 3) + 1  # +1 for AS-IS/TO-BE button row
        button_height = 35  # Normal height per button including margins
        container_height = num_rows * button_height + 10
        self.choice_buttons_container.setMinimumHeight(container_height)
        
        # Update label value filter
        self.cmb_label_value.clear()
        self.cmb_label_value.addItem("전체")
        self.cmb_label_value.addItems(self.label_choices)



    def refresh_as_is_tobe_panel(self) -> None:
        """Update the AS-IS → TO-BE mapping panel"""
        # Check if layout is valid
        if not hasattr(self, 'as_is_tobe_layout') or self.as_is_tobe_layout is None:
            print("Warning: as_is_tobe_layout not initialized")
            return

        try:
            # Clear existing widgets safely
            while self.as_is_tobe_layout.count() > 0:
                item = self.as_is_tobe_layout.takeAt(0)
                if item and item.widget():
                    try:
                        item.widget().setParent(None)
                        item.widget().deleteLater()
                    except RuntimeError:
                        pass  # Widget already deleted
        except (RuntimeError, AttributeError) as e:
            print(f"Error clearing as_is_tobe_layout: {e}")
            return
        
        if self.df is None or self.current_idx >= len(self.filtered_indices):
            return
            
        row_idx = self.filtered_indices[self.current_idx]
        if "Unique_seg_result" not in self.df.columns:
            return

        pred_val = self.df.at[row_idx, "Unique_seg_result"]
        pred_list = parse_pred_list(pred_val)
        
        if not pred_list:
            lbl = QtWidgets.QLabel("AS-IS 매핑용 예측 데이터 없음")
            self.as_is_tobe_layout.addWidget(lbl)
            return
        
        # Create AS-IS → TO-BE mappings
        self.tobe_combos = []  # Store combos for batch apply
        for i, as_is_val in enumerate(pred_list[:5]):  # Limit to 5 items
            lbl_as_is = QtWidgets.QLabel(f"AS-IS: {as_is_val}")
            cmb_tobe = QtWidgets.QComboBox()
            cmb_tobe.addItems(["→"] + self.tobe_choices)
            self.tobe_combos.append(cmb_tobe)
            
            # Try to pre-select best match
            best_match = self._find_best_tobe_match(as_is_val)
            if best_match and best_match in self.tobe_choices:
                cmb_tobe.setCurrentText(best_match)
            
            # Set tab order for keyboard navigation
            if i == 0:
                # First combo gets focus when AS-IS/TO-BE mode is activated
                cmb_tobe.setFocusPolicy(QtCore.Qt.StrongFocus)
            else:
                cmb_tobe.setFocusPolicy(QtCore.Qt.StrongFocus)
            
            # Create horizontal layout for each AS-IS/TO-BE pair
            pair_layout = QtWidgets.QHBoxLayout()
            pair_layout.addWidget(lbl_as_is)
            pair_layout.addWidget(cmb_tobe)
            self.as_is_tobe_layout.addLayout(pair_layout)
        
        # Add "Apply All" button (only this button, no individual apply buttons)
        if len(pred_list) > 0:
            btn_apply_all = QtWidgets.QPushButton("모두 적용 (Enter)")
            btn_apply_all.clicked.connect(self._apply_all_tobe_selections)
            btn_apply_all.setFocusPolicy(QtCore.Qt.StrongFocus)
            self.as_is_tobe_layout.addWidget(btn_apply_all, len(pred_list), 0, 1, 2)
            
            # Set tab order: combos -> apply all button
            for i, combo in enumerate(self.tobe_combos):
                if i < len(self.tobe_combos) - 1:
                    QtWidgets.QWidget.setTabOrder(combo, self.tobe_combos[i + 1])
                else:
                    QtWidgets.QWidget.setTabOrder(combo, btn_apply_all)

    def _find_best_tobe_match(self, as_is_val: str) -> Optional[str]:
        """Find the best TO-BE match for an AS-IS value using simple heuristics"""
        as_is_lower = as_is_val.lower()
        
        # Direct match
        if as_is_val in self.tobe_choices:
            return as_is_val
            
        # Partial matching heuristics
        for choice in self.tobe_choices:
            choice_lower = choice.lower()
            if as_is_lower in choice_lower or choice_lower in as_is_lower:
                return choice
        
        return None

    def _apply_tobe_selection(self, combo: QtWidgets.QComboBox) -> None:
        """Apply the selected TO-BE value to the current row"""
        tobe_val = combo.currentText()
        if tobe_val == "→" or not tobe_val:
            return
            
        if self.df is None or self.current_idx >= len(self.filtered_indices):
            return
            
        row_idx = self.filtered_indices[self.current_idx]
        
        # 기존 라벨을 완전히 덮어쓰기 (추가가 아닌 교체)
        new_label = tobe_val
        
        # 즉시 DataFrame 업데이트
        self.df.at[row_idx, self.active_label_col] = new_label
        
        # 즉시 UI 업데이트
        self._update_current_label_display(row_idx, new_label)
        
        # 배치 저장 (지연 없이)
        self._batch_save_json_entry(row_idx, {"values": {self.active_label_col: new_label}})
        
        # 즉시 테이블 새로고침
        self.refresh_table()
        
        self.status.showMessage(f"TO-BE 라벨 적용됨: {tobe_val}")
        
        # 다음 이미지로 자동 이동
        if self.current_idx < len(self.filtered_indices) - 1:
            self.current_idx += 1
            self.refresh_view()
            self.status.showMessage("다음 이미지로 이동됨", 1000)

    def _apply_all_tobe_selections(self) -> None:
        """Apply all TO-BE selections at once"""
        if self.df is None or self.current_idx >= len(self.filtered_indices):
            return
            
        row_idx = self.filtered_indices[self.current_idx]
        
        # 모든 선택된 TO-BE 라벨 수집 (기존 라벨 무시하고 새로 생성)
        selected_labels = []
        for combo in self.tobe_combos:
            tobe_val = combo.currentText()
            if tobe_val != "→" and tobe_val:
                selected_labels.append(tobe_val)
        
        # 선택된 라벨들을 세미콜론으로 구분
        new_label = ';'.join(selected_labels) if selected_labels else ""
        
        # 즉시 DataFrame 업데이트
        self.df.at[row_idx, self.active_label_col] = new_label
        
        # 즉시 UI 업데이트
        self._update_current_label_display(row_idx, new_label)
        
        # 배치 저장 (지연 없이)
        self._batch_save_json_entry(row_idx, {"values": {self.active_label_col: new_label}})
        
        # 즉시 테이블 새로고침
        self.refresh_table()
        
        self.status.showMessage(f"모든 TO-BE 라벨 적용됨: {new_label}")
        

        
        # 모두 적용 후 AS-IS/TO-BE 모드 비활성화 및 다음 이미지로 이동
        if self.as_is_tobe_mode:
            self.as_is_tobe_mode = False
            if hasattr(self, 'as_is_tobe_container'):
                self.as_is_tobe_container.setVisible(False)
                self.as_is_tobe_container.setStyleSheet("")
            self.status.showMessage("AS-IS/TO-BE 모드 비활성화됨", 2000)
            # Refresh label controls to update button style
            self.refresh_label_controls()
            
            # 다음 이미지로 자동 이동
            if self.current_idx < len(self.filtered_indices) - 1:
                self.current_idx += 1
                self.refresh_view()
                self.status.showMessage("다음 이미지로 이동됨", 1000)

    def _assign_label_without_advance(self, row_idx: int, label_value: str) -> None:
        """Assign a label to a specific row without auto-advance"""
        if self.df is None:
            return
            
        # Update DataFrame
        self.df.at[row_idx, self.active_label_col] = label_value
        
        # Update performance metrics
        self._label_count += 1
        
        # Save to JSON (batched)
        self._batch_save_json_entry(row_idx, {"values": {self.active_label_col: label_value}})
        

        
        # Immediately update current label display only
        self._update_current_label_display(row_idx, label_value)
        
        # No auto-advance - just update the display
        self._pending_ui_update = True
        self._ui_update_throttle.start()
        
        # 강제로 테이블 새로고침
        QtCore.QTimer.singleShot(100, self.refresh_table)
        


    def _toggle_as_is_tobe_mode(self) -> None:
        """Toggle AS-IS/TO-BE mode for multi-labeling"""
        # Remember current scroll positions and current row
        controls_scrollbar = self.controls_scroll_area.verticalScrollBar()
        current_scroll_pos = controls_scrollbar.value() if controls_scrollbar else 0
        current_row_idx = self.current_idx if hasattr(self, 'current_idx') else 0
        
        self.as_is_tobe_mode = not self.as_is_tobe_mode
        
        if self.as_is_tobe_mode:
            self.status.showMessage("AS-IS/TO-BE 모드 활성화 - 다중 라벨링 가능", 2000)
            # Show AS-IS/TO-BE container and highlight it
            if hasattr(self, 'as_is_tobe_container'):
                self.as_is_tobe_container.setVisible(True)
                self.as_is_tobe_container.setMaximumHeight(200)  # Expand to full height
                self.as_is_tobe_container.setStyleSheet("QGroupBox { border: 2px solid #4CAF50; background-color: #E8F5E8; }")
        else:
            self.status.showMessage("AS-IS/TO-BE 모드 비활성화 - 단일 라벨링", 2000)
            # Hide AS-IS/TO-BE container and restore style
            if hasattr(self, 'as_is_tobe_container'):
                self.as_is_tobe_container.setVisible(False)
                self.as_is_tobe_container.setMaximumHeight(0)  # Collapse to 0 height
                self.as_is_tobe_container.setStyleSheet("")
        
        # Refresh label controls to update button style
        self.refresh_label_controls()
        
        # Restore scroll position and current row after layout update
        QtCore.QTimer.singleShot(50, lambda: self._restore_ui_state(current_scroll_pos, current_row_idx))
        QtCore.QTimer.singleShot(100, lambda: self._restore_ui_state(current_scroll_pos, current_row_idx))
        QtCore.QTimer.singleShot(200, lambda: self._restore_ui_state(current_scroll_pos, current_row_idx))
        
        # Focus on first TO-BE combo if mode is activated
        if self.as_is_tobe_mode and hasattr(self, 'tobe_combos') and self.tobe_combos:
            QtCore.QTimer.singleShot(150, lambda: self.tobe_combos[0].setFocus())

    def _restore_scroll_position(self, scroll_pos: int) -> None:
        """Restore scroll position to prevent jumping to top"""
        try:
            controls_scrollbar = self.controls_scroll_area.verticalScrollBar()
            if controls_scrollbar and scroll_pos >= 0:
                # Ensure the scroll position is within valid range
                max_scroll = controls_scrollbar.maximum()
                safe_scroll_pos = min(scroll_pos, max_scroll)
                controls_scrollbar.setValue(safe_scroll_pos)
        except Exception as e:
            print(f"스크롤 위치 복원 오류: {e}")

    def _restore_ui_state(self, scroll_pos: int, current_row_idx: int) -> None:
        """Restore UI state including scroll position and current row"""
        try:
            # Force layout update first
            self.controls_scroll_area.updateGeometry()
            QtWidgets.QApplication.processEvents()
            
            # Restore scroll position with multiple attempts
            for delay in [0, 10, 50, 100]:
                QtCore.QTimer.singleShot(delay, lambda pos=scroll_pos: self._restore_scroll_position(pos))
            
            # Ensure current row is still valid and visible
            if (hasattr(self, 'current_idx') and 
                hasattr(self, 'filtered_indices') and 
                self.filtered_indices and
                current_row_idx < len(self.filtered_indices)):
                
                # If current row changed, restore it
                if self.current_idx != current_row_idx:
                    self.current_idx = current_row_idx
                    # Update the view to show the correct row
                    self._minimal_view_update()
                    # Ensure table shows the current row
                    self._check_table_reload_needed()
                    
        except Exception as e:
            print(f"UI 상태 복원 오류: {e}")

    def _debug_image_matching(self) -> None:
        """Debug image matching issues for current row"""
        if self.df is None or not self.filtered_indices or self.current_idx >= len(self.filtered_indices):
            QtWidgets.QMessageBox.information(self, "디버그", "현재 행이 없습니다.")
            return
        
        row_idx = self.filtered_indices[self.current_idx]
        File_path = self.df.at[row_idx, "File_path"]

        if pd.isna(File_path) or not str(File_path).strip():
            QtWidgets.QMessageBox.information(self, "디버그", "현재 행에 이미지 경로가 없습니다.")
            return

        # Detailed debugging information
        debug_info = f"이미지 매칭 디버그 정보:\n\n"
        debug_info += f"현재 행 인덱스: {row_idx}\n"
        debug_info += f"CSV 이미지 경로: {File_path}\n"
        debug_info += f"이미지 기본 경로: {self.images_base}\n\n"
        
        # Test different resolution strategies
        debug_info += "해결 시도 결과:\n"
        
        # 1. Direct path
        if os.path.isabs(str(File_path)) and os.path.exists(str(File_path)):
            debug_info += f"✓ 절대 경로 존재: {File_path}\n"
        else:
            debug_info += f"✗ 절대 경로 없음: {File_path}\n"
        
        # 2. Normalized relative path
        from create_excel_from_seg_csv import normalize_relative_path
        rel = normalize_relative_path(str(File_path))
        debug_info += f"정규화된 상대 경로: {rel}\n"
        
        # 3. Join with base
        candidate = os.path.join(self.images_base, rel)
        if os.path.exists(candidate):
            debug_info += f"✓ 기본 경로 + 상대 경로 존재: {candidate}\n"
        else:
            debug_info += f"✗ 기본 경로 + 상대 경로 없음: {candidate}\n"
        
        # 4. Try _viz variant
        rel_dir = os.path.dirname(rel)
        rel_base, _ = os.path.splitext(os.path.basename(rel))
        viz_candidate = os.path.join(self.images_base, rel_dir, f"{rel_base}_viz.png")
        if os.path.exists(viz_candidate):
            debug_info += f"✓ _viz 변형 존재: {viz_candidate}\n"
        else:
            debug_info += f"✗ _viz 변형 없음: {viz_candidate}\n"
        
        # 5. Search by basename with improved precision
        basename = os.path.basename(rel)
        base_no_ext, _ = os.path.splitext(basename)
        import glob
        
        debug_info += "\n파일명 검색 결과:\n"
        
        # Exact filename match
        exact_pattern = os.path.join(self.images_base, "**", basename)
        exact_matches = glob.glob(exact_pattern, recursive=True)
        if exact_matches:
            debug_info += f"✓ 정확한 파일명 매치: {len(exact_matches)}개 파일 발견\n"
            for match in exact_matches[:3]:
                debug_info += f"  - {match}\n"
        else:
            debug_info += f"✗ 정확한 파일명 매치: 파일 없음\n"
        
        # Base name with any extension (filtered)
        base_pattern = os.path.join(self.images_base, "**", f"{base_no_ext}.*")
        base_matches = glob.glob(base_pattern, recursive=True)
        if base_matches:
            # Filter out matches with additional suffixes
            filtered_matches = []
            for match in base_matches:
                match_basename = os.path.basename(match)
                match_base_no_ext, _ = os.path.splitext(match_basename)
                if match_base_no_ext == base_no_ext:
                    filtered_matches.append(match)
            
            if filtered_matches:
                debug_info += f"✓ 정확한 기본명 매치: {len(filtered_matches)}개 파일 발견\n"
                for match in filtered_matches[:3]:
                    debug_info += f"  - {match}\n"
            else:
                debug_info += f"✗ 정확한 기본명 매치: 필터링 후 파일 없음\n"
        else:
            debug_info += f"✗ 정확한 기본명 매치: 파일 없음\n"
        
        # _viz variant
        if not base_no_ext.endswith('_viz'):
            viz_pattern = os.path.join(self.images_base, "**", f"{base_no_ext}_viz.*")
            viz_matches = glob.glob(viz_pattern, recursive=True)
            if viz_matches:
                debug_info += f"✓ _viz 변형 매치: {len(viz_matches)}개 파일 발견\n"
                for match in viz_matches[:3]:
                    debug_info += f"  - {match}\n"
            else:
                debug_info += f"✗ _viz 변형 매치: 파일 없음\n"
        
        # Show dialog with debug info
        QtWidgets.QMessageBox.information(self, "이미지 매칭 디버그", debug_info)

    def _fix_image_path_manually(self) -> None:
        """Manually fix image path for current row"""
        if self.df is None or not self.filtered_indices or self.current_idx >= len(self.filtered_indices):
            QtWidgets.QMessageBox.information(self, "수정", "현재 행이 없습니다.")
            return
        
        row_idx = self.filtered_indices[self.current_idx]
        current_File_path = self.df.at[row_idx, "File_path"]

        if pd.isna(current_File_path) or not str(current_File_path).strip():
            QtWidgets.QMessageBox.information(self, "수정", "현재 행에 이미지 경로가 없습니다.")
            return

        # Create dialog for manual path selection
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("이미지 경로 수동 수정")
        dialog.setModal(True)
        dialog.resize(600, 400)

        layout = QtWidgets.QVBoxLayout(dialog)

        # Current path info
        current_info = QtWidgets.QLabel(f"현재 CSV 경로: {current_File_path}")
        current_info.setWordWrap(True)
        layout.addWidget(current_info)
        
        # New path input
        path_label = QtWidgets.QLabel("새 이미지 경로:")
        layout.addWidget(path_label)
        
        path_input = QtWidgets.QLineEdit()
        path_input.setPlaceholderText("절대 경로 또는 images_base 기준 상대 경로 입력")
        layout.addWidget(path_input)
        
        # Browse button
        browse_layout = QtWidgets.QHBoxLayout()
        browse_btn = QtWidgets.QPushButton("파일 찾기")
        browse_btn.clicked.connect(lambda: self._browse_for_image(path_input))
        browse_layout.addWidget(browse_btn)
        browse_layout.addStretch()
        layout.addLayout(browse_layout)
        
        # Preview
        preview_label = QtWidgets.QLabel("미리보기:")
        layout.addWidget(preview_label)
        
        preview_path = QtWidgets.QLabel("")
        preview_path.setWordWrap(True)
        preview_path.setStyleSheet("color: gray;")
        layout.addWidget(preview_path)
        
        # Update preview when path changes
        def update_preview():
            new_path = path_input.text().strip()
            if new_path:
                if os.path.isabs(new_path):
                    resolved = new_path
                else:
                    resolved = os.path.join(self.images_base, new_path)
                
                if os.path.exists(resolved):
                    preview_path.setText(f"해결된 경로: {resolved}\n상태: 파일 존재 ✓")
                    preview_path.setStyleSheet("color: green;")
                else:
                    preview_path.setText(f"해결된 경로: {resolved}\n상태: 파일 없음 ✗")
                    preview_path.setStyleSheet("color: red;")
            else:
                preview_path.setText("경로를 입력하세요")
                preview_path.setStyleSheet("color: gray;")
        
        path_input.textChanged.connect(update_preview)
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        cancel_btn = QtWidgets.QPushButton("취소")
        cancel_btn.clicked.connect(dialog.reject)
        apply_btn = QtWidgets.QPushButton("적용")
        apply_btn.clicked.connect(dialog.accept)
        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(apply_btn)
        layout.addLayout(button_layout)
        
        # Show dialog
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            new_path = path_input.text().strip()
            if new_path:
                # Update DataFrame
                self.df.at[row_idx, "File_path"] = new_path
                
                # Refresh current view
                self._load_image_for_row(row_idx)
                
                # Show success message
                QtWidgets.QMessageBox.information(self, "성공", "이미지 경로가 수정되었습니다.")
                
                # Update status
                self.status.showMessage(f"이미지 경로 수정됨: {os.path.basename(new_path)}", 2000)

    def _browse_for_image(self, path_input: QtWidgets.QLineEdit) -> None:
        """Browse for image file and update path input"""
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, 
            "이미지 파일 선택",
            self.images_base,
            "이미지 파일 (*.png *.jpg *.jpeg *.bmp *.tiff *.tif)"
        )
        
        if file_path:
            # Convert to relative path if possible
            if file_path.startswith(self.images_base):
                rel_path = os.path.relpath(file_path, self.images_base)
                path_input.setText(rel_path)
            else:
                path_input.setText(file_path)

    def _quick_fix_wrong_match(self) -> None:
        """Quick fix for wrong image matches by searching for correct file"""
        if self.df is None or not self.filtered_indices or self.current_idx >= len(self.filtered_indices):
            QtWidgets.QMessageBox.information(self, "빠른 수정", "현재 행이 없습니다.")
            return
        
        row_idx = self.filtered_indices[self.current_idx]
        current_File_path = self.df.at[row_idx, "File_path"]

        if pd.isna(current_File_path) or not str(current_File_path).strip():
            QtWidgets.QMessageBox.information(self, "빠른 수정", "현재 행에 이미지 경로가 없습니다.")
            return
        
        # Extract core identifier from CSV path
        csv_basename = os.path.basename(str(current_File_path))
        core_id = csv_basename.replace('.bmp', '').replace('.jpg', '').replace('.png', '').replace('.jpeg', '')
        
        # Search for files with similar names
        import glob
        search_patterns = [
            os.path.join(self.images_base, "**", f"*{core_id}*.*"),
            os.path.join(self.images_base, "**", f"{core_id}.*"),
            os.path.join(self.images_base, "**", f"{core_id}_viz.*"),
        ]
        
        all_matches = []
        for pattern in search_patterns:
            matches = glob.glob(pattern, recursive=True)
            all_matches.extend(matches)
        
        # Remove duplicates and sort
        all_matches = list(set(all_matches))
        all_matches.sort()
        
        if not all_matches:
            QtWidgets.QMessageBox.information(self, "빠른 수정", f"'{core_id}'와 일치하는 이미지를 찾을 수 없습니다.")
            return
        
        # Create selection dialog
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle(f"올바른 이미지 선택 - {csv_basename}")
        dialog.setModal(True)
        dialog.resize(700, 500)
        
        layout = QtWidgets.QVBoxLayout(dialog)
        
        # Info
        info_label = QtWidgets.QLabel(f"CSV 경로: {current_File_path}\n일치하는 파일들:")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        
        # File list
        list_widget = QtWidgets.QListWidget()
        for match in all_matches:
            rel_path = os.path.relpath(match, self.images_base)
            item_text = f"{rel_path}"
            if os.path.exists(match):
                item_text += " ✓"
            list_widget.addItem(item_text)
        
        layout.addWidget(list_widget)
        
        # Preview selected file
        preview_label = QtWidgets.QLabel("미리보기:")
        layout.addWidget(preview_label)
        
        preview_path = QtWidgets.QLabel("")
        preview_path.setWordWrap(True)
        layout.addWidget(preview_path)
        
        def update_preview():
            current_item = list_widget.currentItem()
            if current_item:
                item_text = current_item.text()
                if " ✓" in item_text:
                    file_path = os.path.join(self.images_base, item_text.replace(" ✓", ""))
                    preview_path.setText(f"선택된 파일: {file_path}")
                    preview_path.setStyleSheet("color: green;")
                else:
                    preview_path.setText("파일을 선택하세요")
                    preview_path.setStyleSheet("color: gray;")
        
        list_widget.currentItemChanged.connect(update_preview)
        
        # Buttons
        button_layout = QtWidgets.QHBoxLayout()
        cancel_btn = QtWidgets.QPushButton("취소")
        cancel_btn.clicked.connect(dialog.reject)
        apply_btn = QtWidgets.QPushButton("적용")
        apply_btn.clicked.connect(dialog.accept)
        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(apply_btn)
        layout.addLayout(button_layout)
        
        # Show dialog
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            current_item = list_widget.currentItem()
            if current_item:
                selected_path = current_item.text().replace(" ✓", "")
                # Update DataFrame
                self.df.at[row_idx, "File_path"] = selected_path
                
                # Refresh current view
                self._load_image_for_row(row_idx)
                
                # Show success message
                QtWidgets.QMessageBox.information(self, "성공", f"이미지 경로가 수정되었습니다:\n{selected_path}")
                
                # Update status
                self.status.showMessage(f"빠른 수정 완료: {os.path.basename(selected_path)}", 2000)

    def _assign_by_index(self, choice_idx: int) -> None:
        """Assign label by choice index"""
        if choice_idx >= len(self.label_choices):
            return
        if self.df is None or self.current_idx >= len(self.filtered_indices):
            return
            
        row_idx = self.filtered_indices[self.current_idx]
        choice = self.label_choices[choice_idx]
        
        if self.as_is_tobe_mode:
            # AS-IS/TO-BE 모드: 기존 라벨을 완전히 덮어쓰기
            new_label = choice
            
            # 즉시 DataFrame 업데이트
            self.df.at[row_idx, self.active_label_col] = new_label
            
            # 즉시 UI 업데이트
            self._update_current_label_display(row_idx, new_label)
            
            # 이미지 상태바 업데이트
            self._update_image_status_bar(row_idx)
            
            # 배치 저장 (지연 없이)
            self._batch_save_json_entry(row_idx, {"values": {self.active_label_col: new_label}})
            
            # 즉시 테이블 새로고침
            self.refresh_table()
            
            # 다음 이미지로 자동 이동
            if self.current_idx < len(self.filtered_indices) - 1:
                self.current_idx += 1
                self.refresh_view()
                self.status.showMessage("다음 이미지로 이동됨", 1000)
        else:
            # 일반 모드: 단일 라벨링 (기존 라벨 덮어쓰기) - 자동 진행
            self._assign_label(row_idx, choice)

    def on_auto_advance_toggled(self, checked: bool) -> None:
        """Handle auto-advance setting change"""
        self.auto_advance_enabled = checked


    def _deferred_ui_update(self) -> None:
        """Deferred UI update to improve performance"""
        if self._pending_ui_update:
            self._pending_ui_update = False
            self.refresh_view()
            self._throttled_table_refresh()

    def _throttled_table_refresh(self) -> None:
        """Throttled table refresh for better performance with large datasets"""
        # Only refresh visible portion of table
        if self.df is None or not self.filtered_indices:
            return
        
        # Update only the current row's display instead of full table refresh
        if self.current_idx < len(self.filtered_indices):
            current_row_idx = self.filtered_indices[self.current_idx]
            # Find the table row that corresponds to current_row_idx
            visible_indices = self.filtered_indices[:self.max_table_rows]
            for table_row in range(min(len(visible_indices), self.table.rowCount())):
                if visible_indices[table_row] == current_row_idx:
                    # Update only the label column for this row
                    label_col_idx = None
                    for col_idx in range(self.table.columnCount()):
                        header = self.table.horizontalHeaderItem(col_idx)
                        if header and header.text() == self.active_label_col:
                            label_col_idx = col_idx
                            break
                    
                    if label_col_idx is not None:
                        current_label = self.df.at[current_row_idx, self.active_label_col]
                        item = QtWidgets.QTableWidgetItem(str(current_label) if not pd.isna(current_label) else "")
                        if current_label and str(current_label).strip():
                            item.setBackground(QtGui.QColor(220, 255, 220))  # Green for labeled
                        else:
                            item.setBackground(QtGui.QColor(255, 255, 200))  # Yellow for unlabeled
                        self.table.setItem(table_row, label_col_idx, item)
                    
                    self.table.selectRow(table_row)
                    break

    def _assign_label(self, row_idx: int, label_value: str) -> None:
        """Assign a label to a specific row - optimized for large datasets"""
        if self.df is None:
            return
            
        # Update DataFrame
        self.df.at[row_idx, self.active_label_col] = label_value
        
        # Update performance metrics
        self._label_count += 1
        
        # Save to JSON (batched) - increased batch delay for better performance
        self._batch_save_json_entry(row_idx, {"values": {self.active_label_col: label_value}})
        
        # Immediately update current label display only
        self._update_current_label_display(row_idx, label_value)
        
        # Update image status bar
        self._update_image_status_bar(row_idx)
        
        # Auto-advance to next item if enabled
        if self.auto_advance_enabled and self.current_idx < len(self.filtered_indices) - 1:
            self.current_idx += 1
            # Minimal view update for auto-advance
            self._minimal_view_update()
            

        else:
            # Deferred UI update for better performance
            self._pending_ui_update = True
            self._ui_update_throttle.start()

    def _update_current_label_display(self, row_idx: int, label_value: str) -> None:
        """Immediately update the current label display without full refresh"""
        if self.df is None or not self.filtered_indices:
            return
        
        # Only update if this is the current row being displayed
        current_row = self.filtered_indices[self.current_idx] if self.current_idx < len(self.filtered_indices) else -1
        if current_row != row_idx:
            return
        
        # Update the current info label
        info_text = f"행 {row_idx + 1}/{len(self.df)} (필터됨: {self.current_idx + 1}/{len(self.filtered_indices)})\n"
        
        # 라벨을 세미콜론으로 구분해서 표시
        if label_value:
            labels = label_value.split(';')
            if len(labels) > 1:
                info_text += f"라벨: {' + '.join(labels)}"
            else:
                info_text += f"라벨: {label_value}"
        else:
            info_text += "라벨: 없음"
        
        # AS-IS/TO-BE 모드 상태 표시
        if self.as_is_tobe_mode:
            info_text += f"\n[AS-IS/TO-BE 모드 활성화 - 다중 라벨링]"

        if "Unique_seg_result" in self.df.columns:
            pred_val = self.df.at[row_idx, "Unique_seg_result"]
            info_text += f"\n예측값: {pred_val}"
        if "Result" in self.df.columns:
            result_val = self.df.at[row_idx, "Result"]
            if pd.notna(result_val):
                info_text += f"\n기본결과: {result_val}"
        if "detail" in self.df.columns:
            detail_val = self.df.at[row_idx, "detail"]
            if pd.notna(detail_val) and str(detail_val).strip():
                detail_str = str(detail_val)[:100]
                if len(str(detail_val)) > 100:
                    detail_str += "..."
                info_text += f"\n상세정보: {detail_str}"
        if "model_name" in self.df.columns:
            model_name = self.df.at[row_idx, "model_name"]
            info_text += f"\n모델: {model_name}"
        
        self._safe_set_text(self.lbl_current_info, info_text)
        
        # Update progress dashboard immediately
        self._update_progress_dashboard()
        
        # Defer table update to avoid blocking UI
        QtCore.QTimer.singleShot(200, self._deferred_table_update)
        
        # Save current position immediately
        if hasattr(self, 'settings'):
            self.settings.setValue("current_idx", self.current_idx)

    def _batch_save_json_entry(self, row_idx: int, updater: Dict[str, object]) -> None:
        """Batch save JSON entries and CSV for better performance"""
        self._pending_ops.append((self.json_path, row_idx, updater, {}))
        self._pending_json_path = self.json_path
        self._save_timer.start()
        
        # Update save status
        self._update_save_status("저장 대기 중", "#FFA500")
        
        # Also update DataFrame for real-time CSV saving
        if self.df is not None:
            if "values" in updater:
                for col, value in updater["values"].items():
                    if col in self.df.columns:
                        self.df.at[row_idx, col] = value
            elif "bookmark" in updater:
                # Add bookmark column if it doesn't exist
                if "bookmark" not in self.df.columns:
                    self.df["bookmark"] = False
                self.df.at[row_idx, "bookmark"] = updater["bookmark"]

    def _flush_pending_ops(self) -> None:
        """Flush pending JSON operations - optimized for large datasets"""
        if not self._pending_ops:
            return
        
        try:
            # Load store only once for all operations
            store = load_label_store(self._pending_json_path)
            
            # Batch update all entries
            saved_count = len(self._pending_ops)
            for _json_path, row_idx, updater, _ in self._pending_ops:
                key = str(row_idx)
                entry = store["labels"].get(key) or {}
                for k, v in updater.items():
                    entry[k] = v
                store["labels"][key] = entry
            
            # Save only once after all updates
            save_label_store(self._pending_json_path, store)
            
            # Also save to CSV in real-time
            if saved_count > 0 and self.df is not None:
                try:
                    # Update save status
                    self._update_save_status("저장 중...", "#FFA500")
                    
                    # Create backup before saving (keep only last 5 backups)
                    backup_dir = os.path.dirname(self.csv_path)
                    base_name = os.path.splitext(os.path.basename(self.csv_path))[0]
                    backup_pattern = os.path.join(backup_dir, f"{base_name}_backup_*.csv")
                    
                    # Clean old backups (keep only last 5)
                    import glob
                    backup_files = glob.glob(backup_pattern)
                    backup_files.sort()
                    if len(backup_files) > 5:
                        for old_backup in backup_files[:-5]:
                            try:
                                os.remove(old_backup)
                            except:
                                pass
                    
                    # Create backup of original CSV (if not already exists)
                    backup_path = self.csv_path.replace('.csv', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv')
                    if os.path.exists(self.csv_path):
                        import shutil
                        shutil.copy2(self.csv_path, backup_path)
                    
                    # Save labeled data to new CSV file (not overwrite original)
                    labeled_csv_path = self.csv_path.replace('.csv', f'_labeled_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv')
                    self.df.to_csv(labeled_csv_path, index=False)
                    print(f"✅ 라벨링 결과가 포함된 CSV 저장됨: {labeled_csv_path}")
                    print(f"   원본 CSV 백업됨: {backup_path}")
                    self._update_save_status("저장 완료", "#4CAF50")
                    self.status.showMessage(f"데이터 저장 완료: {saved_count}개 항목 (JSON + CSV)", 1000)
                except Exception as csv_error:
                    print(f"CSV 저장 오류: {csv_error}")
                    self._update_save_status("CSV 저장 실패", "#F44336")
                    self.status.showMessage(f"JSON 저장 완료: {saved_count}개 항목 (CSV 저장 실패)", 1000)
            else:
                self._update_save_status("저장 완료", "#4CAF50")
                self.status.showMessage(f"데이터 저장 완료: {saved_count}개 항목", 1000)
            
            self._pending_ops.clear()
                
        except Exception as e:
            print(f"JSON 저장 오류: {e}")
            self._update_save_status("저장 실패", "#F44336")
            # Don't clear ops if save failed, will retry on next flush

    def _minimal_view_update(self) -> None:
        """Minimal view update for auto-advance - optimized for performance"""
        if not getattr(self, "_ui_ready", False):
            print("⏸️ _minimal_view_update: UI not ready yet")
            return
        if self.df is None or not self.filtered_indices or self.current_idx >= len(self.filtered_indices):
            return
            
        row_idx = self.filtered_indices[self.current_idx]
        
        # Update only essential info
        current_label = self.df.at[row_idx, self.active_label_col] if self.active_label_col in self.df.columns else ""
        # Handle pandas NA values safely
        current_label_str = str(current_label) if not pd.isna(current_label) else "(라벨없음)"
        info_text = f"행 {row_idx + 1}/{len(self.df)} (필터됨: {self.current_idx + 1}/{len(self.filtered_indices)})\n"
        info_text += f"라벨: {current_label_str}"
        if "Unique_seg_result" in self.df.columns:
            pred_val = self.df.at[row_idx, "Unique_seg_result"]
            pred_val_str = str(pred_val) if not pd.isna(pred_val) else "(없음)"
            info_text += f"\n예측값: {pred_val_str}"
        if "Result" in self.df.columns:
            result_val = self.df.at[row_idx, "Result"]
            result_val_str = str(result_val) if not pd.isna(result_val) else "(없음)"
            info_text += f"\n기본결과: {result_val_str}"
        if "detail" in self.df.columns:
            detail_val = self.df.at[row_idx, "detail"]
            if pd.notna(detail_val) and str(detail_val).strip():
                detail_str = str(detail_val)[:100]
                if len(str(detail_val)) > 100:
                    detail_str += "..."
                info_text += f"\n상세정보: {detail_str}"
        if "model_name" in self.df.columns:
            model_name = self.df.at[row_idx, "model_name"]
            model_name_str = str(model_name) if not pd.isna(model_name) else "(없음)"
            info_text += f"\n모델: {model_name_str}"
        info_text += f"\n\n단축키: 1.OK 2.애매한OK 3.NG 4.애매한NG 5.보류 6.SRLogicOK 7.AS-IS/TO-BE모드"
        info_text += f"\n이동: ←→↑↓ 또는 A/D 또는 Space"
        if self.as_is_tobe_mode:
            info_text += f"\nAS-IS/TO-BE: Tab이동 Enter적용"
        self._safe_set_text(self.lbl_current_info, info_text)
        
        # Update bookmark status only
        entry = get_json_entry(self.json_path, row_idx)
        bookmark_status = entry.get("bookmark", False)
        self._safe_set_text(self.lbl_bookmark_status, f"북마크: {'✅' if bookmark_status else '❌'}")
        
        # Update image status bar
        self._update_image_status_bar(row_idx)
        

        
        # Load image immediately for navigation
        print(f"Loading image for row {row_idx} (current_idx: {self.current_idx})")
        try:
            self._load_image_for_row(row_idx)
            print(f"Image load completed for row {row_idx}")
        except Exception as e:
            print(f"Error in _load_image_for_row: {e}")
            import traceback
            traceback.print_exc()
            # Show error in UI
            self._clear_image_display("이미지 로드 오류", f"오류: {str(e)}")
        
        # Update progress dashboard
        self._update_progress_dashboard()
        
        # Defer AS-IS/TO-BE panel update for better performance
        QtCore.QTimer.singleShot(50, self.refresh_as_is_tobe_panel)

    def _update_image_status_bar(self, row_idx: int) -> None:
        """Update the image status bar with label information"""
        if self.df is None or row_idx >= len(self.df):
            self._safe_set_text(self.image_status_bar, "데이터 오류")
            return
            
        # Get current label
        current_label = self.df.at[row_idx, self.active_label_col] if self.active_label_col in self.df.columns else ""
        current_label_str = str(current_label) if not pd.isna(current_label) else ""
        
        # Get prediction result
        pred_result = ""
        if "Unique_seg_result" in self.df.columns:
            pred_val = self.df.at[row_idx, "Unique_seg_result"]
            pred_result = str(pred_val) if not pd.isna(pred_val) else ""

        # Get Result value
        result_val = ""
        if "Result" in self.df.columns:
            res_val = self.df.at[row_idx, "Result"]
            result_val = str(res_val) if not pd.isna(res_val) else ""
        
        # Create status text
        if current_label_str and current_label_str.strip():
            # Has label
            status_text = f"✅ 라벨됨: {current_label_str}"
            if result_val:
                status_text += f" | 기본결과: {result_val}"
            if pred_result:
                status_text += f" | 예측: {pred_result}"
            self._safe_set_style(self.image_status_bar, """
                QLabel {
                    background-color: #d4edda;
                    border: 1px solid #c3e6cb;
                    border-radius: 4px;
                    padding: 8px;
                    margin: 2px;
                    font-weight: bold;
                    color: #155724;
                }
            """)
        else:
            # No label
            status_text = f"❌ 라벨 없음"
            if result_val:
                status_text += f" | 기본결과: {result_val}"
            if pred_result:
                status_text += f" | 예측: {pred_result}"
            self._safe_set_style(self.image_status_bar, """
                QLabel {
                    background-color: #f8d7da;
                    border: 1px solid #f5c6cb;
                    border-radius: 4px;
                    padding: 8px;
                    margin: 2px;
                    font-weight: bold;
                    color: #721c24;
                }
            """)
        
        self._safe_set_text(self.image_status_bar, status_text)

    def _load_image_if_changed(self, row_idx: int) -> None:
        """Load image only if the path has changed - performance optimization"""
        print(f"🖼️ _load_image_if_changed 호출됨: row_idx={row_idx}")
        if self.df is None or "File_path" not in self.df.columns:
            print("❌ 데이터프레임이 없거나 File_path 컬럼이 없음")
            return

        File_path = self.df.at[row_idx, "File_path"]
        if pd.isna(File_path) or not str(File_path).strip():
            if self._last_image_path != "":
                self.image_label.setText("이미지 경로 없음")
                self.path_label.clear()
                self._last_image_path = ""
            return
        
        resolved_path = resolve_image_path(self.images_base, str(File_path))

        # If resolve_image_path failed, try a more aggressive search
        if not resolved_path:
            # Extract filename from CSV path
            csv_filename = os.path.basename(str(File_path))
            if csv_filename:
                # Search for file with same name anywhere in images_base
                self.status.showMessage(f"🔍 이미지 검색 중: {csv_filename}")
                QtWidgets.QApplication.processEvents()

                for root, dirs, files in os.walk(self.images_base):
                    for file in files:
                        if file == csv_filename:
                            resolved_path = os.path.join(root, file)
                            print(f"✅ 대체 검색으로 이미지 찾음: {resolved_path}")
                            self.status.showMessage(f"✅ 이미지 찾음: {csv_filename}")
                            break
                    if resolved_path:
                        break

        # Only load if path changed
        if resolved_path != self._last_image_path:
            if resolved_path:
                self.status.showMessage(f"🖼️ 이미지 로드: {os.path.basename(resolved_path)}")
                self._load_image_for_row(row_idx)
            else:
                self.status.showMessage(f"❌ 이미지 찾을 수 없음: {os.path.basename(str(File_path))}")
                if self._last_image_path != "":
                    self.image_label.setText("이미지 경로 없음")
                    self.path_label.clear()
                    self._last_image_path = ""
            self._last_image_path = resolved_path or ""

    def _deferred_table_update(self) -> None:
        """Deferred table update with throttling for better performance"""
        current_time = QtCore.QDateTime.currentMSecsSinceEpoch()
        if current_time - self._last_table_update < self._table_update_throttle:
            # Schedule for later
            QtCore.QTimer.singleShot(self._table_update_throttle, self._deferred_table_update)
            return
        
        self._last_table_update = current_time
        self.refresh_table()

    def _update_table_selection(self) -> None:
        """Update table selection to match current image"""
        if self.df is None or not self.filtered_indices or self.current_idx >= len(self.filtered_indices):
            print(f"_update_table_selection: Invalid state - df={self.df is not None}, filtered_indices={len(self.filtered_indices) if self.filtered_indices else 0}, current_idx={self.current_idx}")
            return
            
        # Clear current selection first
        self.table.clearSelection()
        
        # Get the current row's original index
        row_idx = self.filtered_indices[self.current_idx]
        print(f"_update_table_selection: Looking for original_idx {row_idx} (current_idx={self.current_idx})")
        
        # First try: Use smart loading position if available and valid
        if hasattr(self, '_current_row_in_window') and hasattr(self, '_table_start_pos'):
            table_row = self._current_row_in_window
            print(f"_update_table_selection: Smart position {table_row} (table.rowCount()={self.table.rowCount()})")
            if 0 <= table_row < self.table.rowCount():
                # Verify this is the correct row
                item = self.table.item(table_row, 0)
                if item is not None:
                    original_idx = item.data(QtCore.Qt.UserRole)
                    if original_idx == row_idx:
                        self.table.selectRow(table_row)
                        print(f"Selected table row {table_row} using smart position")
                        return
                    else:
                        print(f"Smart position {table_row} has wrong original_idx {original_idx}, expected {row_idx}")
                else:
                    print(f"Smart position {table_row} has no item")
            else:
                print(f"Smart position {table_row} out of range [0, {self.table.rowCount()})")
        
        # Second try: Search by data (fallback)
        print(f"_update_table_selection: Searching all table rows for original_idx {row_idx}")
        for table_row in range(self.table.rowCount()):
            item = self.table.item(table_row, 0)
            if item is not None:
                original_idx = item.data(QtCore.Qt.UserRole)
                if original_idx == row_idx:
                    self.table.selectRow(table_row)
                    print(f"Found and selected table row {table_row} by search")
                    return
        
        print(f"Could not find current row {self.current_idx} (original_idx {row_idx}) in table")
        # Don't trigger refresh here to avoid jumping to last row
        # Just log the issue and continue

    def _check_and_load_more_data(self, current_table_row: int) -> None:
        """Check if we need to load more data when approaching the end of visible data"""
        if self.df is None or not self.filtered_indices:
            return
        
        # Calculate how many rows are currently visible in the table
        current_visible_count = min(self.max_table_rows, len(self.filtered_indices))
        
        # If we're within 5 rows of the end of visible data, load more
        if current_table_row >= current_visible_count - 5:
            # Check if there are more filtered indices to show
            if len(self.filtered_indices) > current_visible_count:
                # Increase the number of visible rows
                self.max_table_rows = min(self.max_table_rows + 50, len(self.filtered_indices))
                
                # Refresh the table to show more data
                QtCore.QTimer.singleShot(100, self.refresh_table)
                
                # Show status message
                self.status.showMessage(f"더 많은 데이터 로드됨: {self.max_table_rows}개 행 표시", 2000)

    def _check_and_load_more_data_for_navigation(self) -> None:
        """Check if we need to load more data during navigation"""
        if self.df is None or not self.filtered_indices:
            return
        
        # Calculate how many rows are currently visible in the table
        current_visible_count = min(self.max_table_rows, len(self.filtered_indices))
        
        # If we're within 10 rows of the end of visible data, load more
        if self.current_idx >= current_visible_count - 10:
            # Check if there are more filtered indices to show
            if len(self.filtered_indices) > current_visible_count:
                # Increase the number of visible rows
                self.max_table_rows = min(self.max_table_rows + 50, len(self.filtered_indices))
                
                # Refresh the table to show more data
                QtCore.QTimer.singleShot(100, self.refresh_table)
                
                # Show status message
                self.status.showMessage(f"더 많은 데이터 로드됨: {self.max_table_rows}개 행 표시", 2000)

    def _on_table_scroll(self, value: int) -> None:
        """Handle table scroll events for smart loading"""
        if self.df is None or not self.filtered_indices:
            return
        
        # Get scroll bar information
        scroll_bar = self.table.verticalScrollBar()
        max_value = scroll_bar.maximum()
        
        # Calculate which row is currently visible at the top
        if max_value > 0:
            scroll_ratio = value / max_value
            total_visible_rows = self.table.rowCount()
            
            # Estimate which filtered index should be at the top
            if hasattr(self, '_table_start_pos') and total_visible_rows > 0:
                estimated_top_index = int(self._table_start_pos + scroll_ratio * total_visible_rows)
                
                # If we're scrolling to a position that's not in our current window, trigger smart reload
                if (estimated_top_index < self._table_start_pos or 
                    estimated_top_index >= self._table_end_pos):
                    # Defer the reload to avoid too many updates
                    if not hasattr(self, '_scroll_reload_timer'):
                        self._scroll_reload_timer = QtCore.QTimer()
                        self._scroll_reload_timer.setSingleShot(True)
                        self._scroll_reload_timer.timeout.connect(self._trigger_smart_table_reload)
                    
                    self._scroll_reload_timer.start(200)  # 200ms delay

    def _update_image_from_table_selection(self) -> None:
        """Update image based on currently selected table row"""
        if self.df is None or not self.filtered_indices:
            return
        
        # Get the currently selected row in the table
        selected_rows = self.table.selectedItems()
        if not selected_rows:
            return
        
        # Get the first selected row (should be only one)
        selected_row = selected_rows[0].row()
        item = self.table.item(selected_row, 0)
        if item is None:
            return
        
        # Get the original dataframe index from the selected row
        original_idx = item.data(QtCore.Qt.UserRole)
        if original_idx is None:
            return
        
        # Update current_idx to match the selected row
        if original_idx in self.filtered_indices:
            new_current_idx = self.filtered_indices.index(original_idx)
            if new_current_idx != self.current_idx:
                self.current_idx = new_current_idx
        
        # Update the view based on the selected row
        self._minimal_view_update()

        # Check if we've reached the end
        if self.current_idx >= len(self.filtered_indices) - 1:
            # If we're at the very last item, show completion message
            self.status.showMessage("모든 항목 라벨링 완료!", 3000)

    def toggle_bookmark(self) -> None:
        """Toggle bookmark for current row"""
        if self.df is None or self.current_idx >= len(self.filtered_indices):
            return
            
        row_idx = self.filtered_indices[self.current_idx]
        entry = get_json_entry(self.json_path, row_idx)
        current_bookmark = entry.get("bookmark", False)
        
        self._batch_save_json_entry(row_idx, {"bookmark": not current_bookmark})
        
        # Update UI immediately
        self._safe_set_text(self.lbl_bookmark_status, f"북마크: {'✅' if not current_bookmark else '❌'}")
        self.status.showMessage(f"행 {row_idx + 1} 북마크: {'켜짐' if not current_bookmark else '꺼짐'}")
        
        # Update image status bar
        self._update_image_status_bar(row_idx)
        
        # Refresh table to show bookmark status
        self.refresh_table()

    def _get_filter_hash(self) -> str:
        """Generate hash of current filter settings for caching"""
        filter_state = (
            self.cmb_label_state.currentText(),
            self.cmb_label_value.currentText(),
            self.cmb_result_filter.currentText(),
            self.cmb_background_filter.currentText(),
            self.chk_bookmarks.isChecked(),
            tuple(sorted(self.selected_pred_filters))
        )
        return str(hash(filter_state))

    def apply_filters(self) -> None:
        """Apply various filters to determine which rows to show - optimized for large datasets"""
        if not getattr(self, "_ui_ready", False):
            print("⏸️ apply_filters: UI not ready yet")
            return

        print("🔍 apply_filters 호출됨")

        if self.df is None:
            print("❌ self.df가 None입니다")
            self.status.showMessage("❌ 데이터가 로드되지 않았습니다", 2000)
            return

        print(f"📝 데이터프레임 크기: {len(self.df)} 행")
        
        # 버튼 비활성화로 사용자 피드백 제공
        if hasattr(self, 'btn_apply_filters'):
            self.btn_apply_filters.setEnabled(False)
            self.btn_apply_filters.setText("⏳ 적용 중...")

        try:
            # Check if filters have changed to avoid unnecessary recalculation
            current_filter_hash = self._get_filter_hash()
            if self._last_filter_hash == current_filter_hash and self._filter_cache is not None:
                self.filtered_indices = self.df[self._filter_cache].index.tolist()
                self._update_filter_results()
                return
                
            # Start with all rows
            mask = pd.Series([True] * len(self.df), index=self.df.index)
            
            # Label state filter
            label_state = self.cmb_label_state.currentText()
            if label_state == "라벨됨":
                mask &= ~(self.df[self.active_label_col].isna() | (self.df[self.active_label_col] == ""))
            elif label_state == "라벨안됨":
                mask &= (self.df[self.active_label_col].isna() | (self.df[self.active_label_col] == ""))
            
            # Label value filter
            label_value = self.cmb_label_value.currentText()
            if label_value and label_value != "전체":
                mask &= (self.df[self.active_label_col] == label_value)

            # Result filter
            result_value = self.cmb_result_filter.currentText()
            print(f"🔍 Result 필터: '{result_value}' (컬럼 존재: {'Result' in self.df.columns})")
            if result_value and result_value != "전체" and "Result" in self.df.columns:
                before_count = mask.sum()
                mask &= (self.df["Result"] == result_value)
                after_count = mask.sum()
                print(f"   Result 필터 적용: {before_count} → {after_count} 개 행")
            
            # Background_result filter (use actual values from CSV)
            background_value = self.cmb_background_filter.currentText()
            print(f"🖼️ Background_result 필터: '{background_value}' (컬럼 존재: {'Background_result' in self.df.columns})")
            if background_value and background_value != "전체" and "Background_result" in self.df.columns:
                before_count = mask.sum()
                mask &= (self.df["Background_result"] == background_value)
                after_count = mask.sum()
                print(f"   Background_result 필터 적용: {before_count} → {after_count} 개 행")
            
            # Unique_seg_result filter
            if self.selected_pred_filters and "Unique_seg_result" in self.df.columns:
                pred_mask = pd.Series([False] * len(self.df), index=self.df.index)
                for idx, row in self.df.iterrows():
                    pred_list = parse_pred_list(row["Unique_seg_result"])
                    if any(pred_val in self.selected_pred_filters for pred_val in pred_list):
                        pred_mask.at[idx] = True
                mask &= pred_mask
            
            # Bookmarks filter
            if self.chk_bookmarks.isChecked():
                store = load_label_store(self.json_path)
                bookmarked_rows = [int(k) for k, v in store.get("labels", {}).items() if v.get("bookmark", False)]
                mask &= self.df.index.isin(bookmarked_rows)
            
            # Cache filter results and update indices
            self._filter_cache = mask
            self._last_filter_hash = current_filter_hash
            self.filtered_indices = self.df[mask].index.tolist()
            
            self._update_filter_results()

        except Exception as e:
            print(f"❌ apply_filters 오류: {e}")
            self.status.showMessage(f"필터 적용 중 오류: {str(e)}", 5000)

        finally:
            # 버튼 복원
            if hasattr(self, 'btn_apply_filters'):
                self.btn_apply_filters.setEnabled(True)
                self.btn_apply_filters.setText("🔍 필터 적용")

    def _update_filter_results(self) -> None:
        """Update UI after filter results are ready"""
        if not getattr(self, "_ui_ready", False):
            print("⏸️ _update_filter_results: UI not ready yet")
            return

        print("📊 _update_filter_results 호출됨")
        print(f"🔍 필터된 인덱스 수: {len(self.filtered_indices) if hasattr(self, 'filtered_indices') else 'None'}")
        
        # Ensure current index is valid
        if self.current_idx >= len(self.filtered_indices):
            self.current_idx = max(0, len(self.filtered_indices) - 1)
        
        try:
            # Update UI efficiently with forced refresh
            print("🖼️ refresh_view 호출 시작")
            self.refresh_view()
            print("✅ refresh_view 호출 완료")

            # Always refresh table for better user experience
            print("📋 테이블 갱신 시작")
            self.refresh_table()
            print("✅ 테이블 갱신 완료")
            
            # Update progress dashboard
            self._update_progress_dashboard()
            
            # Force UI update
            QtWidgets.QApplication.processEvents()
            
            # Show filter status with more details
            total_items = len(self.df) if self.df is not None else 0
            if len(self.filtered_indices) > 0:
                self.status.showMessage(f"✅ 필터 적용됨: {len(self.filtered_indices):,}개 / {total_items:,}개 행", 3000)
            else:
                self.status.showMessage("⚠️ 필터 결과가 없습니다", 3000)

        except Exception as e:
            print(f"❌ _update_filter_results 오류: {e}")
            self.status.showMessage(f"필터 적용 중 오류 발생: {str(e)}", 5000)
            # Retry table refresh on error
            QtCore.QTimer.singleShot(100, self.refresh_table)

    def refresh_view(self) -> None:
        """Refresh the current view (image and info)"""
        if not getattr(self, "_ui_ready", False):
            print("⏸️ refresh_view: UI not ready yet")
            return
        print("🖼️ refresh_view 시작")
        print(f"📊 self.df는 {'존재' if self.df is not None else '없음'}")
        print(f"🔍 filtered_indices: {len(self.filtered_indices) if hasattr(self, 'filtered_indices') and self.filtered_indices else '없음'}")
        
        if self.df is None or not self.filtered_indices:
            print("❌ 데이터가 없어서 빈 화면 표시")
            self._safe_set_text(self.lbl_current_info, "표시할 데이터 없음")
            if hasattr(self, 'image_label') and self.image_label is not None:
                self.image_label.clear()
            if hasattr(self, 'path_label') and self.path_label is not None:
                self.path_label.clear()
            self._update_progress_dashboard()
            return
        
        if self.current_idx >= len(self.filtered_indices):
            return
            
        row_idx = self.filtered_indices[self.current_idx]
        
        # Update current info
        current_label = self.df.at[row_idx, self.active_label_col] if self.active_label_col in self.df.columns else ""
        info_text = f"행 {row_idx + 1}/{len(self.df)} (필터됨: {self.current_idx + 1}/{len(self.filtered_indices)})\n"
        info_text += f"라벨: {current_label or '(라벨없음)'}"
        if "Unique_seg_result" in self.df.columns:
            pred_val = self.df.at[row_idx, "Unique_seg_result"]
            info_text += f"\n예측값: {pred_val}"
        if "model_name" in self.df.columns:
            model_name = self.df.at[row_idx, "model_name"]
            info_text += f"\n모델: {model_name}"
        info_text += f"\n\n단축키: 1.OK 2.애매한OK 3.NG 4.애매한NG 5.보류 6.SRLogicOK 7.AS-IS/TO-BE모드"
        self._safe_set_text(self.lbl_current_info, info_text)
        
        # Update bookmark status
        entry = get_json_entry(self.json_path, row_idx)
        bookmark_status = entry.get("bookmark", False)
        
        self._safe_set_text(self.lbl_bookmark_status, f"북마크: {'✅' if bookmark_status else '❌'}")
        
        # Load and display image (optimized for speed)
        self._load_image_if_changed(row_idx)
        
        # Refresh AS-IS/TO-BE panel
        self.refresh_as_is_tobe_panel()
        
        # Update progress dashboard
        self._update_progress_dashboard()

    def _prepare_overlay_info(self, row_idx: int) -> dict:
        """JSON 파일에서 오버레이 정보를 준비합니다."""
        overlay_info = {
            'json_found': False,
            'json_path': '',
            'details': [],
            'result': '',
            'has_overlay': False,
            'annotations': [],  # 런랭스 마스크 정보 추가
            'image_size': None,
            'bboxes': []  # bbox 정보 추가
        }

        try:
            # Result_path에서 JSON 파일 경로 추출
            if "Result_path" in self.df.columns:
                result_path = self.df.at[row_idx, "Result_path"]
                if pd.notna(result_path) and str(result_path).strip():
                    result_path_str = str(result_path).strip()

                    # JSON 파일 경로를 찾는 여러 방법 시도
                    json_file_path = None

                    # 1. 절대 경로로 존재하는지 확인
                    if os.path.isabs(result_path_str):
                        if result_path_str.endswith('.json'):
                            json_file_path = result_path_str
                        else:
                            json_file_path = result_path_str + '.json'
                        if not os.path.exists(json_file_path):
                            json_file_path = None

                    # 2. JSON 기본 경로에서 파일명으로 찾기 (개선된 알고리즘)
                    if json_file_path is None and self.json_base:
                        # File_path에서 파일명만 추출
                        if "File_path" in self.df.columns:
                            file_path = str(self.df.at[row_idx, "File_path"])
                            filename = os.path.basename(file_path)
                            name_without_ext = os.path.splitext(filename)[0]

                            print(f"🔍 JSON 검색 시작: {filename} -> {name_without_ext}.json")

                            # 개선된 JSON 후보 경로들
                            json_candidates = []

                            # 기본 패턴들
                            json_candidates.extend([
                                os.path.join(self.json_base, name_without_ext + '.json'),
                                os.path.join(self.json_base, filename + '.json'),
                                os.path.join(self.json_base, filename),
                                os.path.join(self.json_base, name_without_ext, name_without_ext + '.json'),
                            ])

                            # 초유연 구조 기반 오버레이 JSON 검색
                            import re

                            # CSV File_path 구조 분석
                            csv_number_pattern = re.search(r'/(\d+)/', file_path)
                            csv_structure = {}

                            if csv_number_pattern:
                                csv_number = csv_number_pattern.group(1)
                                csv_structure['number'] = csv_number

                                # /숫자/ 이후 경로 분석
                                after_number = file_path.split(f'/{csv_number}/', 1)[1]
                                path_parts = after_number.split('/')

                                if len(path_parts) >= 4:
                                    csv_structure.update({
                                        'part1': path_parts[0],  # 0001
                                        'part2': path_parts[1],  # Unit
                                        'part3': path_parts[2],  # U12, U70 등
                                        'part4': path_parts[3],  # BC, FC 등
                                    })

                            # JSON 기본 경로 구조 분석
                            base_number_pattern = re.search(r'/test/(\d+)/', self.json_base)
                            base_structure = {}

                            if base_number_pattern:
                                base_number = base_number_pattern.group(1)
                                base_structure['number'] = base_number

                                # 기본 경로의 나머지 부분 분석
                                after_base_number = self.json_base.split(f'/test/{base_number}/', 1)[1]
                                base_path_parts = after_base_number.split('/')

                                if len(base_path_parts) >= 4:
                                    base_structure.update({
                                        'part1': base_path_parts[0],
                                        'part2': base_path_parts[1],
                                        'part3': base_path_parts[2],
                                        'part4': base_path_parts[3],
                                    })

                            # 구조 기반 JSON 패턴 생성
                            if csv_structure and base_structure:
                                # 다양한 Unit 폴더 조합 생성
                                unit_folders = ['U0', 'U1', 'U2', 'U6', 'U7', 'U8', 'U9', 'U10', 'U11', 'U12', 'U13', 'U14', 'U15', 'U16', 'U19']
                                type_folders = ['BC', 'FC', 'DC']

                                for unit in unit_folders:
                                    for type_folder in type_folders:
                                        # JSON 구조 기반 경로 생성
                                        json_struct_path = f"{base_number}/{base_structure.get('part2', 'Unit')}/{unit}/{type_folder}/{name_without_ext}.json"
                                        json_candidates.append(os.path.join(self.json_base.split(f'/test/{base_number}/')[0], "test", json_struct_path))

                                        # /Unit 폴더에서도
                                        unit_base = os.path.join(self.json_base.split(f'/test/{base_number}/')[0], "test", "Unit", f"{name_without_ext}.json")
                                        json_candidates.append(unit_base)

                                        # /img 폴더에서도
                                        img_base = os.path.join(self.json_base.split(f'/test/{base_number}/')[0], "test", "img", json_struct_path)
                                        json_candidates.append(img_base)

                            # 기존 방식도 유지 (폴백)
                            number_pattern_in_base = re.search(r'/test/(\d+)/', self.json_base)
                            if number_pattern_in_base:
                                number_part = number_pattern_in_base.group(1)
                                unit_base = self.json_base.replace(f'/test/{number_part}/', '/test/Unit/')
                                json_candidates.extend([
                                    os.path.join(unit_base, name_without_ext + '.json'),
                                    os.path.join(unit_base, filename),
                                ])

                            # Unit 폴더 간 검색 (U0 -> U9, U12 등)
                            if '/Unit/' in self.json_base:
                                # Unit 폴더 목록
                                unit_folders = ['U0', 'U1', 'U2', 'U6', 'U7', 'U8', 'U9', 'U10', 'U11', 'U12', 'U13', 'U14', 'U15', 'U16', 'U19']

                                for unit_folder in unit_folders:
                                    # 기본 Unit 경로에서
                                    unit_path = self.json_base.replace('/Unit/U0/', f'/Unit/{unit_folder}/')
                                    json_candidates.extend([
                                        os.path.join(unit_path, name_without_ext + '.json'),
                                        os.path.join(unit_path, filename),
                                    ])

                                    # /img Unit 경로에서도
                                    if number_pattern_in_base:
                                        number_part = number_pattern_in_base.group(1)
                                        img_unit_path = self.json_base.replace(f'/test/{number_part}/', '/test/img/1/').replace('/Unit/U0/', f'/Unit/{unit_folder}/')
                                        json_candidates.extend([
                                            os.path.join(img_unit_path, name_without_ext + '.json'),
                                            os.path.join(img_unit_path, filename),
                                        ])

                            # 기존 Unit 기반 폴백
                            if "Unit" in self.json_base:
                                unit_parent = os.path.dirname(self.json_base)
                                json_candidates.extend([
                                    os.path.join(unit_parent, "**", name_without_ext + '.json'),
                                    os.path.join(unit_parent, "**", filename),
                                ])

                            # Unit 폴더의 모든 하위 폴더에서 찾기
                            for root, dirs, files in os.walk(self.json_base):
                                for file in files:
                                    if file == name_without_ext + '.json' or file == filename:
                                        json_candidates.append(os.path.join(root, file))

                            # 모든 후보 경로에서 검색
                            for candidate in json_candidates:
                                if os.path.exists(candidate):
                                    json_file_path = candidate
                                    print(f"✅ JSON 파일 발견: {candidate}")
                                    break

                                # glob 패턴으로 검색 (디렉토리인 경우)
                                if os.path.isdir(os.path.dirname(candidate)):
                                    import glob
                                    pattern = os.path.join(os.path.dirname(candidate), "**", os.path.basename(candidate))
                                    matches = glob.glob(pattern, recursive=True)
                                    if matches:
                                        json_file_path = matches[0]
                                        print(f"✅ JSON 파일 발견 (재귀 검색): {json_file_path}")
                                        break

                    # 3. 기존 방식으로도 시도 (fallback)
                    if json_file_path is None and self.json_base:
                        # 상대 경로를 JSON 기본 경로와 결합
                        combined_path = os.path.join(self.json_base, result_path_str.lstrip('/'))
                        if combined_path.endswith('.json'):
                            json_file_path = combined_path
                        else:
                            json_file_path = combined_path + '.json'

                        if not os.path.exists(json_file_path):
                            json_file_path = None

                    # 3. 다양한 변형 시도
                    if json_file_path is None:
                        # .json 확장자 없이도 시도
                        base_path = os.path.join(self.json_base, result_path_str.lstrip('/'))
                        if os.path.exists(base_path):
                            json_file_path = base_path
                        elif os.path.exists(base_path + '.json'):
                            json_file_path = base_path + '.json'

                    if json_file_path and os.path.exists(json_file_path):
                        overlay_info['json_path'] = json_file_path
                        overlay_info['json_found'] = True
                        details = extract_detail_from_json(json_file_path)
                        overlay_info['details'] = details

                        # bbox 정보 추출
                        bboxes = extract_bbox_from_json(json_file_path)
                        overlay_info['bboxes'] = bboxes
                        print(f"📦 bbox 추출 완료: {len(bboxes)}개 (JSON: {json_file_path})")
                        if bboxes:
                            print(f"🎯 bbox 데이터 발견!")
                            for i, bbox in enumerate(bboxes[:3]):  # 처음 3개만 출력
                                print(f"   bbox[{i}]: {bbox['label']} at {bbox['bbox']} (score: {bbox['score']:.3f})")
                        else:
                            print(f"⚠️ bbox 데이터 없음")

                        # 런랭스 마스크 정보 추출
                        annotations, image_size = self._extract_run_length_data(json_file_path)
                        overlay_info['annotations'] = annotations
                        overlay_info['image_size'] = image_size
                        overlay_info['has_overlay'] = len(annotations) > 0 or len(details) > 0 or len(bboxes) > 0
                        print(f"✅ JSON 파일 발견: {json_file_path}")
                        print(f"   - bbox: {len(bboxes)}개")
                        print(f"   - annotations: {len(annotations)}개")
                        print(f"   - details: {len(details)}개")
                        print(f"   - has_overlay: {overlay_info['has_overlay']}")
                        print(f"   - show_overlay: {self.show_overlay}")
                    else:
                        print(f"JSON 파일을 찾을 수 없음: {result_path_str} (기본 경로: {self.json_base})")

            # Result 값도 포함
            if "Result" in self.df.columns:
                result_val = self.df.at[row_idx, "Result"]
                if pd.notna(result_val):
                    overlay_info['result'] = str(result_val)

            # 현재 라벨 정보도 포함
            current_label = self.df.at[row_idx, self.active_label_col] if self.active_label_col in self.df.columns else ""
            if pd.notna(current_label) and str(current_label).strip():
                overlay_info['current_label'] = str(current_label)

        except Exception as e:
            print(f"오버레이 정보 준비 중 오류: {e}")

        return overlay_info

    def _extract_run_length_data(self, json_path: str) -> tuple:
        """JSON 파일에서 런랭스 마스크 정보를 추출합니다."""
        annotations = []
        image_size = None

        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            if isinstance(data, dict):
                # 이미지 크기 정보 추출
                if 'imageWidth' in data and 'imageHeight' in data:
                    image_size = (data['imageWidth'], data['imageHeight'])

                # 어노테이션 정보 추출
                if 'annotations' in data and isinstance(data['annotations'], list):
                    for ann in data['annotations']:
                        if isinstance(ann, dict):
                            annotation = {
                                'type': ann.get('type', ''),
                                'label': ann.get('label', ''),
                                'bbox': ann.get('bbox', []),
                                'score': ann.get('score', 0.0),
                                'mask': data.get('mask', [])  # 런랭스 마스크 데이터
                            }
                            annotations.append(annotation)

        except Exception as e:
            print(f"런랭스 데이터 추출 중 오류: {e}")

        return annotations, image_size

    def _load_image_for_row(self, row_idx: int) -> None:
        """Load and display image for the given row - safe and error-free"""
        try:
            if not getattr(self, "_ui_ready", False):
                print("⏸️ _load_image_for_row: UI not ready yet")
                return
            if self.df is None or "File_path" not in self.df.columns:
                self._clear_image_display("데이터가 로드되지 않음")
                return

            # Safe data access to handle pandas NA values
            try:
                File_path = self.df.at[row_idx, "File_path"]
                if pd.isna(File_path):
                    self._clear_image_display("이미지 경로 없음")
                    return
                File_path_str = str(File_path).strip()
                if not File_path_str:
                    self._clear_image_display("이미지 경로 없음")
                    return
            except Exception as e:
                print(f"이미지 경로 접근 오류: {e}")
                self._clear_image_display("이미지 경로 접근 오류")
                return

            # JSON 정보 추출 및 오버레이 데이터 준비
            overlay_info = self._prepare_overlay_info(row_idx)
            print(f"📋 오버레이 정보: JSON 찾음={overlay_info['json_found']}, 오버레이={overlay_info['has_overlay']}")
            
            # 상태창에 JSON 매칭 정보 표시
            if overlay_info['json_found']:
                bbox_count = len(overlay_info.get('bboxes', []))
                annotation_count = len(overlay_info.get('annotations', []))
                if bbox_count > 0 or annotation_count > 0:
                    status_msg = f"📄 JSON 매칭됨: bbox {bbox_count}개, 어노테이션 {annotation_count}개"
                else:
                    status_msg = "📄 JSON 파일 있음 (데이터 없음)"
                self.status.showMessage(status_msg, 5000)
                print(f"   JSON 경로: {overlay_info['json_path']}")
                print(f"   bbox 개수: {bbox_count}")
                print(f"   어노테이션 개수: {annotation_count}")
            else:
                self.status.showMessage("❌ JSON 파일 매칭 실패", 3000)
        
            print(f"이미지 로드 시도: 행 {row_idx}, 경로: {File_path_str}")
            print(f"  기본 경로: {self.images_base}")
            
            # Resolve image path with detailed debugging
            resolved_path = resolve_image_path(self.images_base, File_path_str)
            print(f"  해결된 경로: {resolved_path}")
            
            # Additional validation to prevent wrong matches
            if resolved_path:
                expected_path = os.path.join(self.images_base, File_path_str)
                if resolved_path != expected_path:
                    # Check if the resolved path is significantly different from expected
                    resolved_basename = os.path.basename(resolved_path)
                    expected_basename = os.path.basename(File_path_str)
                    
                    # Extract core identifiers from filenames for comparison
                    def extract_core_id(filename):
                        # Remove extensions and common suffixes
                        core = filename.replace('.bmp', '').replace('.jpg', '').replace('.png', '').replace('.jpeg', '')
                        # Remove _viz suffix if present
                        if core.endswith('_viz'):
                            core = core[:-4]
                        return core
                    
                    expected_core = extract_core_id(expected_basename)
                    resolved_core = extract_core_id(resolved_basename)
                    
                    # If core identifiers are completely different, this is likely a wrong match
                    if expected_core != resolved_core:
                        print(f"⚠️  잘못된 매칭 감지!")
                        print(f"  예상: {expected_basename} (코어: {expected_core})")
                        print(f"  실제: {resolved_basename} (코어: {resolved_core})")
                        print(f"  이 매칭이 올바른지 확인하세요.")
                        
                        # Show warning to user
                        self.status.showMessage(f"⚠️ 잘못된 매칭: {expected_basename} → {resolved_basename}", 5000)
                        
                        # Prevent loading wrong image by setting resolved_path to None
                        resolved_path = None
            
                    if not resolved_path or not os.path.exists(resolved_path):
                        # Enhanced error message with debugging info
                        error_msg = f"이미지를 찾을 수 없음\nCSV 경로: {File_path_str}\n기본 경로: {self.images_base}"
                        if resolved_path:
                            error_msg += f"\n해결된 경로: {resolved_path}"
                            if not os.path.exists(resolved_path):
                                error_msg += f"\n파일 존재 여부: ❌"
                        else:
                            error_msg += f"\n해결된 경로: None"
                        
                        self._clear_image_display("이미지를 찾을 수 없음", error_msg)
                        
                        # Log detailed debugging info
                        print(f"이미지 매칭 실패:")
                        print(f"  CSV 경로: {File_path_str}")
                        print(f"  기본 경로: {self.images_base}")
                        print(f"  해결된 경로: {resolved_path}")
                        if resolved_path:
                            print(f"  파일 존재 여부: {os.path.exists(resolved_path)}")
                        
                        # Show warning about potential wrong matches
                        if resolved_path and resolved_path != os.path.join(self.images_base, File_path_str):
                            print(f"  ⚠️  경고: 잘못된 매칭 가능성 - CSV와 실제 파일이 다를 수 있습니다!")
                        
                        # Show status message
                        self.status.showMessage(f"이미지를 찾을 수 없음: {os.path.basename(File_path_str)}", 3000)
                        return
            
            # Load image with caching - optimized for speed
            cache_key = resolved_path
            if cache_key in self._image_cache:
                pixmap = self._image_cache[cache_key]
            else:
                # Load image directly without thumbnail building for maximum speed
                pixmap = QtGui.QPixmap(resolved_path)
                
                # Cache management - optimized for speed
                if len(self._image_cache) >= self.image_cache_size * 3:  # Triple cache size for better performance
                    # Remove oldest entry
                    self._image_cache.pop(next(iter(self._image_cache)))
                self._image_cache[cache_key] = pixmap
            
            if pixmap.isNull():
                self._clear_image_display("이미지 로드 실패", f"로드 오류: {resolved_path}")
                return
            
            # Display image - optimized for speed
            if self.fit_to_window and hasattr(self, 'scroll_area') and self.scroll_area is not None:
                try:
                    scroll_size = self.scroll_area.viewport().size()
                    # Use FastTransformation for speed instead of SmoothTransformation
                    scaled_pixmap = pixmap.scaled(scroll_size, QtCore.Qt.KeepAspectRatio, QtCore.Qt.FastTransformation)
                    display_pixmap = scaled_pixmap
                except RuntimeError:
                    # Widget has been deleted, fall back to original size
                    display_pixmap = pixmap
            else:
                display_pixmap = pixmap

            # 오버레이 정보가 있으면 추가
            if overlay_info.get('has_overlay', False):
                print(f"🎨 오버레이 적용 시작 (show_overlay: {self.show_overlay})")
                original_cache_key = display_pixmap.cacheKey()
                display_pixmap = self._add_overlay_to_pixmap(display_pixmap, overlay_info)
                new_cache_key = display_pixmap.cacheKey()
                print(f"🎨 오버레이 적용 완료 (pixmap 변경됨: {original_cache_key != new_cache_key})")
                
                # 오버레이 적용 성공 시 상태창 업데이트
                if self.show_overlay:
                    bbox_count = len(overlay_info.get('bboxes', []))
                    annotation_count = len(overlay_info.get('annotations', []))
                    if bbox_count > 0 or annotation_count > 0:
                        self.status.showMessage(f"✅ 오버레이 적용됨: bbox {bbox_count}개, 어노테이션 {annotation_count}개", 3000)
                    else:
                        self.status.showMessage("⚠️ 오버레이 적용됨 (데이터 없음)", 3000)

            if hasattr(self, 'image_label') and self.image_label is not None:
                try:
                    self.image_label.setPixmap(display_pixmap)
                    print(f"🖼️ 이미지 레이블에 픽스맵 설정 완료: {display_pixmap.width()}x{display_pixmap.height()}")
                except RuntimeError:
                    pass  # Widget has been deleted
            
            if hasattr(self, 'path_label') and self.path_label is not None:
                try:
                    self.path_label.setText(resolved_path)
                except RuntimeError:
                    pass  # Widget has been deleted
            print(f"이미지 로드 성공: {resolved_path}")
            print(f"  이미지 크기: {pixmap.width()}x{pixmap.height()}")
            print(f"  표시 모드: {'fit_to_window' if self.fit_to_window else 'original_size'}")
            
        except Exception as e:
            print(f"이미지 로드 중 오류 발생: {e}")
            import traceback
            traceback.print_exc()
            self._clear_image_display("이미지 로드 오류", f"오류: {str(e)}")
            self.status.showMessage(f"이미지 로드 오류: {str(e)}", 3000)

    def _add_overlay_to_pixmap(self, pixmap: QtGui.QPixmap, overlay_info: dict) -> QtGui.QPixmap:
        """픽스맵에 오버레이 정보를 추가합니다."""
        print(f"🎨 _add_overlay_to_pixmap 시작")
        print(f"   - has_overlay: {overlay_info.get('has_overlay', False)}")
        print(f"   - show_overlay: {self.show_overlay}")
        print(f"   - bboxes: {len(overlay_info.get('bboxes', []))}개")
        print(f"   - annotations: {len(overlay_info.get('annotations', []))}개")

        if not overlay_info.get('has_overlay', False):
            print("⚠️ has_overlay가 False이므로 오버레이 생략")
            print(f"   현재 상태: bboxes={len(overlay_info.get('bboxes', []))}, annotations={len(overlay_info.get('annotations', []))}, details={len(overlay_info.get('details', []))}")
            return pixmap

        # 원본 픽스맵을 복사하여 수정
        overlay_pixmap = pixmap.copy()
        print(f"✅ 픽스맵 복사 완료: {overlay_pixmap.width()}x{overlay_pixmap.height()}")

        # QPainter를 사용하여 텍스트 그리기
        painter = QtGui.QPainter(overlay_pixmap)
        try:
            painter.setRenderHint(QtGui.QPainter.Antialiasing)
            print("✅ QPainter 생성 완료")

            # bbox 오버레이 먼저 그리기 (가장 아래에 표시되도록)
            if overlay_info.get('bboxes') and self.show_overlay:
                print(f"🎨 bbox 오버레이 그리기: {len(overlay_info['bboxes'])}개")
                self._draw_bbox_overlay(painter, overlay_info, pixmap.width(), pixmap.height())
            elif overlay_info.get('bboxes'):
                print(f"⚠️ bbox 데이터 있지만 오버레이 꺼짐: {len(overlay_info['bboxes'])}개")

            # 런랭스 마스크 오버레이 그리기 (bbox 위에 표시되도록)
            if overlay_info.get('annotations') and self.show_overlay:
                print(f"🎨 런랭스 오버레이 그리기: {len(overlay_info['annotations'])}개")
                self._draw_run_length_overlay(painter, overlay_info, pixmap.width(), pixmap.height())
            elif overlay_info.get('annotations'):
                print(f"⚠️ 런랭스 데이터 있지만 오버레이 꺼짐: {len(overlay_info['annotations'])}개")

            # 폰트 설정
            font = QtGui.QFont("Arial", 12, QtGui.QFont.Bold)
            painter.setFont(font)

            # 배경색 설정 (반투명 검은색)
            bg_color = QtGui.QColor(0, 0, 0, 180)  # 검은색, 70% 투명도
            painter.setBrush(bg_color)
            painter.setPen(QtCore.Qt.NoPen)

            # 텍스트 색상 설정
            text_color = QtGui.QColor(255, 255, 255)  # 흰색
            painter.setPen(text_color)

            # 오버레이 정보 구성
            overlay_lines = []

            if overlay_info.get('result'):
                overlay_lines.append(f"결과: {overlay_info['result']}")

            if overlay_info.get('current_label'):
                overlay_lines.append(f"라벨: {overlay_info['current_label']}")

            # 어노테이션 정보 추가
            if overlay_info.get('annotations'):
                for i, ann in enumerate(overlay_info['annotations'][:3]):  # 최대 3개 어노테이션 표시
                    label = ann.get('label', 'Unknown')
                    score = ann.get('score', 0.0)
                    overlay_lines.append(f"• {label}: {score:.1f}")

            if overlay_info.get('details') and not overlay_info.get('annotations'):
                for i, detail in enumerate(overlay_info['details'][:3]):  # 최대 3개까지만 표시
                    overlay_lines.append(f"• {detail}")

            if overlay_info.get('json_found'):
                overlay_lines.append("📄 JSON 파일 있음")

            # 각 라인의 높이 계산
            font_metrics = QtGui.QFontMetrics(font)
            line_height = font_metrics.height()
            padding = 10

            # 오버레이 배경 영역 계산
            max_text_width = 0
            for line in overlay_lines:
                max_text_width = max(max_text_width, font_metrics.horizontalAdvance(line))

            overlay_width = max_text_width + (padding * 2)
            overlay_height = (line_height * len(overlay_lines)) + (padding * 2)

            # 오버레이 위치 (우하단)
            overlay_x = pixmap.width() - overlay_width - 20
            overlay_y = pixmap.height() - overlay_height - 20

            # 배경 사각형 그리기
            overlay_rect = QtCore.QRect(overlay_x, overlay_y, overlay_width, overlay_height)
            painter.drawRoundedRect(overlay_rect, 8, 8)

            # 텍스트 그리기
            text_y = overlay_y + padding + font_metrics.ascent()
            for line in overlay_lines:
                painter.drawText(overlay_x + padding, text_y, line)
                text_y += line_height

            print(f"🎨 오버레이 픽스맵 생성 완료: {overlay_pixmap.width()}x{overlay_pixmap.height()}")
            print(f"   원본과 다른가?: {overlay_pixmap.cacheKey() != pixmap.cacheKey()}")

        except Exception as e:
            print(f"오버레이 추가 중 오류: {e}")
            import traceback
            traceback.print_exc()
        finally:
            # QPainter 안전하게 종료
            if painter.isActive():
                painter.end()
                print("✅ QPainter 종료됨")

        return overlay_pixmap

    def _draw_bbox_overlay(self, painter: QtGui.QPainter, overlay_info: dict, img_width: int, img_height: int):
        """bbox 정보를 이미지 위에 사각형 오버레이로 그립니다."""
        try:
            bboxes = overlay_info.get('bboxes', [])
            print(f"🎨 _draw_bbox_overlay 시작: {len(bboxes)}개 bbox")
            if not bboxes:
                print("⚠️ bbox 데이터가 없음")
                return

            # label별 색상 캐시 (중복 계산 방지)
            if not hasattr(self, '_label_color_cache'):
                self._label_color_cache = {}

            drawn_count = 0
            for i, bbox_info in enumerate(bboxes):
                bbox = bbox_info.get('bbox', [])
                label = bbox_info.get('label', 'Unknown')
                score = bbox_info.get('score', 0.0)
                print(f"   bbox[{i}] 처리 중: {label} (score: {score:.2f}), coords: {bbox}")

                if len(bbox) != 4:
                    print(f"⚠️ bbox[{i}] 형식이 잘못됨: {bbox}")
                    continue

                x1, y1, x2, y2 = bbox
                
                # JSON 이미지 크기와 실제 표시 이미지 크기 간 스케일링 적용
                json_img_width = bbox_info.get('json_img_width')
                json_img_height = bbox_info.get('json_img_height')
                
                if json_img_width and json_img_height:
                    # 스케일 팩터 계산
                    scale_x = img_width / json_img_width
                    scale_y = img_height / json_img_height
                    
                    # bbox 좌표 스케일링
                    x1 = int(x1 * scale_x)
                    y1 = int(y1 * scale_y) 
                    x2 = int(x2 * scale_x)
                    y2 = int(y2 * scale_y)
                    
                    print(f"📦 bbox[{i}] 스케일링 적용: {label}")
                    print(f"   JSON 크기: {json_img_width}x{json_img_height} → 표시 크기: {img_width}x{img_height}")
                    print(f"   스케일: {scale_x:.3f}x{scale_y:.3f}")
                    print(f"   좌표: {bbox} → [{x1},{y1},{x2},{y2}]")
                else:
                    print(f"📦 bbox[{i}]: {label} at ({x1},{y1})-({x2},{y2}) score={score:.3f} (스케일링 없음)")

                # label별 유니크 색상 생성/캐시
                if label not in self._label_color_cache:
                    self._label_color_cache[label] = generate_label_color(label)
                    print(f"🎨 새 색상 생성: {label} -> {self._label_color_cache[label].name()}")

                color = self._label_color_cache[label]

                # 사각형 테두리 그리기
                pen = QtGui.QPen(color, 3)  # 3px 두께의 테두리
                painter.setPen(pen)
                rect = QtCore.QRectF(x1, y1, x2 - x1, y2 - y1)
                painter.drawRect(rect)
                print(f"✅ 사각형 그리기 완료: {rect} (펜: {pen.color().name()}, 두께: {pen.width()})")

                # 텍스트 배경을 위한 사각형 (반투명)
                text_bg_color = QtGui.QColor(color)
                text_bg_color.setAlpha(180)  # 70% 불투명
                text_rect = QtCore.QRectF(x1, y1 - 25, x2 - x1, 20)
                painter.fillRect(text_rect, text_bg_color)
                print(f"✅ 텍스트 배경 그리기 완료: {text_rect}")

                # 텍스트 그리기 (흰색)
                painter.setPen(QtGui.QPen(QtGui.QColor(255, 255, 255), 2))
                font = QtGui.QFont("Arial", 10, QtGui.QFont.Bold)
                painter.setFont(font)

                # label과 score 표시
                text = f"{label}: {score:.3f}"
                painter.drawText(int(x1 + 5), int(y1 - 8), text)
                print(f"✅ 텍스트 그리기 완료: '{text}'")

                drawn_count += 1

            print(f"🎉 bbox 오버레이 완료: {drawn_count}/{len(bboxes)}개 그리기 성공")

        except Exception as e:
            print(f"❌ bbox 오버레이 그리기 중 오류: {e}")
            import traceback
            traceback.print_exc()

    def _draw_run_length_overlay(self, painter: QtGui.QPainter, overlay_info: dict, img_width: int, img_height: int):
        """런랭스 마스크를 이미지 위에 오버레이로 그립니다."""
        try:
            # JSON에서 추출한 이미지 크기 사용 (있는 경우)
            json_size = overlay_info.get('image_size')
            if json_size:
                mask_width, mask_height = json_size
            else:
                mask_width, mask_height = img_width, img_height

            print(f"마스크 크기: {mask_width}x{mask_height}, 이미지 크기: {img_width}x{img_height}")

            for i, annotation in enumerate(overlay_info.get('annotations', [])):
                mask_data = annotation.get('mask', [])
                if not mask_data:
                    continue

                # 런랭스 디코딩
                mask_image = self._decode_run_length(mask_data, mask_width, mask_height)
                if mask_image is None:
                    continue

                # 마스크를 QImage로 변환
                mask_qimage = self._mask_to_qimage(mask_image)

                # 어노테이션 색상 설정 (객체마다 다른 색상)
                colors = [
                    QtGui.QColor(255, 0, 0, 100),    # 빨강 (SR-이물 등)
                    QtGui.QColor(0, 255, 0, 100),    # 초록
                    QtGui.QColor(0, 0, 255, 100),    # 파랑
                    QtGui.QColor(255, 255, 0, 100),  # 노랑
                    QtGui.QColor(255, 0, 255, 100),  # 마젠타
                ]
                color = colors[i % len(colors)]

                # 마스크 오버레이 그리기 (이미지 크기에 맞게 스케일링)
                if mask_qimage:
                    mask_pixmap = QtGui.QPixmap.fromImage(mask_qimage)

                    # 마스크가 이미지 크기와 다르면 스케일링
                    if mask_width != img_width or mask_height != img_height:
                        scaled_mask = mask_pixmap.scaled(img_width, img_height, QtCore.Qt.IgnoreAspectRatio, QtCore.Qt.FastTransformation)
                        painter.setOpacity(0.4)  # 40% 투명도
                        painter.drawPixmap(0, 0, scaled_mask)
                    else:
                        painter.setOpacity(0.4)  # 40% 투명도
                        painter.drawPixmap(0, 0, mask_pixmap)

                    painter.setOpacity(1.0)  # 투명도 리셋

                # 바운딩 박스 그리기 (이미지 크기에 맞게 스케일링)
                bbox = annotation.get('bbox', [])
                if len(bbox) == 4:
                    x1, y1, x2, y2 = bbox

                    # 마스크 크기와 이미지 크기가 다르면 바운딩 박스도 스케일링
                    if mask_width != img_width or mask_height != img_height:
                        scale_x = img_width / mask_width
                        scale_y = img_height / mask_height
                        x1, x2 = x1 * scale_x, x2 * scale_x
                        y1, y2 = y1 * scale_y, y2 * scale_y

                    # 바운딩 박스 선 설정
                    pen = QtGui.QPen(color)
                    pen.setWidth(3)
                    painter.setPen(pen)
                    painter.setBrush(QtCore.Qt.NoBrush)

                    # 바운딩 박스 그리기
                    painter.drawRect(int(x1), int(y1), int(x2 - x1), int(y2 - y1))

                    # 라벨 텍스트 표시
                    label = annotation.get('label', 'Unknown')
                    score = annotation.get('score', 0.0)

                    label_text = f"{label} ({score:.2f})"
                    painter.setFont(QtGui.QFont("Arial", 10, QtGui.QFont.Bold))

                    # 텍스트 배경
                    text_rect = QtCore.QRect(int(x1), int(y1 - 25), 200, 20)
                    painter.fillRect(text_rect, QtGui.QColor(0, 0, 0, 150))

                    # 텍스트
                    painter.setPen(QtGui.QColor(255, 255, 255))
                    painter.drawText(int(x1 + 5), int(y1 - 10), label_text)

        except Exception as e:
            print(f"런랭스 오버레이 그리기 중 오류: {e}")
            import traceback
            traceback.print_exc()

    def _decode_run_length(self, mask_data: list, width: int, height: int) -> list:
        """런랭스 인코딩된 마스크를 디코딩합니다."""
        try:
            if not mask_data or not isinstance(mask_data, list):
                return None

            # 1차원 마스크 생성
            flat_mask = []
            for pair in mask_data:
                if isinstance(pair, list) and len(pair) == 2:
                    value, count = pair
                    flat_mask.extend([value] * count)

            # 2차원 마스크로 변환
            total_pixels = width * height
            if len(flat_mask) != total_pixels:
                print(f"마스크 크기 불일치: 예상 {total_pixels}, 실제 {len(flat_mask)}")
                return None

            mask_2d = []
            for i in range(height):
                row_start = i * width
                row_end = (i + 1) * width
                mask_2d.append(flat_mask[row_start:row_end])

            return mask_2d

        except Exception as e:
            print(f"런랭스 디코딩 중 오류: {e}")
            return None

    def _mask_to_qimage(self, mask_2d: list) -> QtGui.QImage:
        """2차원 마스크를 QImage로 변환합니다."""
        try:
            if not mask_2d or not mask_2d[0]:
                return None

            height = len(mask_2d)
            width = len(mask_2d[0])

            # RGBA 형식의 QImage 생성
            image = QtGui.QImage(width, height, QtGui.QImage.Format_ARGB32)
            image.fill(QtCore.Qt.transparent)  # 투명으로 초기화

            # 마스크 데이터 적용
            for y in range(height):
                for x in range(width):
                    if mask_2d[y][x] == 1:  # 객체 픽셀
                        # 반투명 빨강으로 설정
                        color = QtGui.QColor(255, 0, 0, 100)
                        image.setPixelColor(x, y, color)

            return image

        except Exception as e:
            print(f"마스크를 QImage로 변환 중 오류: {e}")
            return None

    def _safe_set_text(self, widget, text: str) -> None:
        """Safely set text on a widget, handling deleted C++ objects"""
        try:
            if widget is not None:
                widget.setText(text)
        except (RuntimeError, AttributeError) as e:
            print(f"⚠️ 위젯 텍스트 설정 실패: {e}")

    def _safe_set_style(self, widget, style: str) -> None:
        """Safely set style on a widget, handling deleted C++ objects"""
        try:
            if widget is not None:
                widget.setStyleSheet(style)
        except (RuntimeError, AttributeError) as e:
            print(f"⚠️ 위젯 스타일 설정 실패: {e}")

    def _safe_widget_operation(self, widget, operation, *args, **kwargs) -> bool:
        """Safely perform an operation on a widget, handling deleted C++ objects"""
        try:
            if widget is not None:
                operation(*args, **kwargs)
                return True
        except (RuntimeError, AttributeError) as e:
            print(f"⚠️ 위젯 작업 실패: {e}")
        return False

    def _safe_clear_layout(self, layout) -> None:
        """Safely clear a layout, handling deleted C++ objects.
        CRITICAL: Only clears dynamic inner layouts, NEVER core UI widgets.
        """
        if not layout:
            return
        
        # UI must be ready to perform safe operations
        if not getattr(self, "_ui_ready", False):
            print("⚠️ _safe_clear_layout: UI not ready, skipping layout clear")
            return
            
        try:
            # Define ONLY allowed layouts - core UI widgets are NEVER cleared
            allowed_layouts = set()
            try:
                if hasattr(self, 'choice_buttons_layout') and self.choice_buttons_layout:
                    allowed_layouts.add(self.choice_buttons_layout)
                if hasattr(self, 'pred_filter_checkboxes_layout') and self.pred_filter_checkboxes_layout:
                    allowed_layouts.add(self.pred_filter_checkboxes_layout)
                if hasattr(self, 'as_is_tobe_layout') and self.as_is_tobe_layout:
                    allowed_layouts.add(self.as_is_tobe_layout)
            except Exception:
                pass

            # STRICT: Only proceed if layout is explicitly allowed
            if layout not in allowed_layouts:
                print(f"⚠️ _safe_clear_layout: 허용되지 않은 레이아웃 요청 - 건너뜀 (layout: {type(layout)})")
                return

            print(f"🧹 Safe layout clear: {type(layout)} with {layout.count()} items")
            
            # Clear only child widgets, not the layout itself
            while layout.count() > 0:
                item = layout.takeAt(0)
                if not item:
                    continue
                w = item.widget()
                if w is not None:
                    try:
                        # Only delete dynamic widgets, not core UI components
                        w.setParent(None)
                        w.deleteLater()
                    except RuntimeError:
                        pass  # Widget already deleted
                else:
                    nested = item.layout()
                    if nested is not None and nested in allowed_layouts:
                        self._safe_clear_layout(nested)
                        
            print(f"✅ Layout cleared successfully")
        except (RuntimeError, AttributeError) as e:
            print(f"❌ Error clearing layout: {e}")
            import traceback
            traceback.print_exc()

    def _clear_image_display(self, message: str, path_info: str = "") -> None:
        """Safely clear image display with error message"""
        try:
            if hasattr(self, 'image_label') and self.image_label is not None:
                try:
                    self.image_label.setText(message)
                except RuntimeError:
                    pass  # Widget has been deleted
            if hasattr(self, 'path_label') and self.path_label is not None:
                try:
                    if path_info:
                        self.path_label.setText(path_info)
                    else:
                        self.path_label.clear()
                except RuntimeError:
                    pass  # Widget has been deleted
        except Exception as e:
            print(f"이미지 표시 지우기 중 오류: {e}")

    def _restore_scroll_position(self) -> None:
        """Restore scroll position to maintain user's view"""
        try:
            if hasattr(self, '_saved_scroll_pos'):
                # Calculate the target scroll position to keep current row visible
                if self.current_idx < len(self.filtered_indices):
                    # Calculate where the current row should be in the table
                    if hasattr(self, '_table_start_pos') and hasattr(self, '_current_row_in_window'):
                        relative_pos = self._current_row_in_window
                        if 0 <= relative_pos < self.table.rowCount():
                            # Calculate scroll position to center the current row
                            row_height = self.table.rowHeight(0) if self.table.rowCount() > 0 else 20
                            target_scroll = max(0, relative_pos * row_height - self.table.viewport().height() // 2)
                            
                            # Set scroll position
                            self.table.verticalScrollBar().setValue(int(target_scroll))
                            print(f"스크롤 위치 복원: {target_scroll} (현재 행: {self.current_idx}, 상대 위치: {relative_pos})")
                        else:
                            # Fallback to saved position
                            self.table.verticalScrollBar().setValue(self._saved_scroll_pos)
                            print(f"스크롤 위치 복원 (fallback): {self._saved_scroll_pos}")
                    else:
                        # Fallback to saved position
                        self.table.verticalScrollBar().setValue(self._saved_scroll_pos)
                        print(f"스크롤 위치 복원 (fallback): {self._saved_scroll_pos}")
        except Exception as e:
            print(f"스크롤 위치 복원 중 오류: {e}")

    def refresh_table(self) -> None:
        """Refresh the data table with smart loading - optimized for large datasets"""
        if not getattr(self, "_ui_ready", False):
            print("⏸️ refresh_table: UI not ready yet")
            return
        if self.df is None or not hasattr(self, 'table') or self.table is None:
            return
        
        # Get visible data (filtered rows only)
        if not self.filtered_indices:
            try:
                self.table.setRowCount(0)
            except RuntimeError:
                return  # Table widget has been deleted
            return
        
        # Smart table loading: ensure current row is always visible
        print(f"refresh_table: current_idx={self.current_idx}, total_filtered={len(self.filtered_indices)}")
        visible_indices = self._get_smart_visible_indices()
        print(f"visible_indices range: {self._table_start_pos} to {self._table_end_pos}, current_row_in_window={self._current_row_in_window}")
        visible_df = self.df.iloc[visible_indices]
        
        # Set up table - add model_name if available
        base_cols = ["File_path", "Result", "Background_result", "Unique_seg_result"]
        display_cols = base_cols.copy()

        # Only add active_label_col if it's not already in the list (avoid duplicates)
        if self.active_label_col not in display_cols:
            display_cols.append(self.active_label_col)

        if "model_name" in visible_df.columns:
            # Insert before the last column (which should be the label column)
            display_cols.insert(-1, "model_name")
        display_cols = [col for col in display_cols if col in visible_df.columns]
        
        try:
            self.table.setRowCount(len(visible_df))
            self.table.setColumnCount(len(display_cols))
            self.table.setHorizontalHeaderLabels(display_cols)
        except RuntimeError:
            return  # Table widget has been deleted
        
        # Debug: Check if data is available
        if len(visible_df) == 0:
            print("Warning: visible_df is empty!")
            return

        print(f"Table will show {len(visible_df)} rows with columns: {display_cols}")

        # Fill table with visible data
        for i, (original_idx, row) in enumerate(visible_df.iterrows()):
            # Check if this row is bookmarked
            entry = get_json_entry(self.json_path, original_idx)
            is_bookmarked = entry.get("bookmark", False)
            
            for j, col in enumerate(display_cols):
                cell_value = str(row[col]) if not pd.isna(row[col]) else ""
                # Truncate long values for better display
                if col == "Unique_seg_result" and len(cell_value) > 50:
                    cell_value = cell_value[:47] + "..."
                elif col == "detail" and len(cell_value) > 100:
                    cell_value = cell_value[:97] + "..."
                elif col == "File_path" and len(cell_value) > 80:
                    # File path는 파일명만 표시
                    import os
                    basename = os.path.basename(cell_value)
                    if len(basename) < len(cell_value):
                        cell_value = "..." + basename
                
                # Add bookmark indicator to first column
                if j == 0 and is_bookmarked:
                    cell_value = "🔖 " + cell_value
                
                item = QtWidgets.QTableWidgetItem(cell_value)
                # Store the original dataframe index in the item for reference
                if j == 0:  # Store in first column
                    item.setData(QtCore.Qt.UserRole, original_idx)
                
                # Color code based on label status
                if col == self.active_label_col:
                    if cell_value and cell_value.strip():
                        # Has label - light green background
                        item.setBackground(QtGui.QColor(220, 255, 220))
                    else:
                        # No label - light yellow background
                        item.setBackground(QtGui.QColor(255, 255, 200))
                elif is_bookmarked:
                    # Bookmark row - light blue background
                    item.setBackground(QtGui.QColor(220, 235, 255))
                
                self.table.setItem(i, j, item)
        
        # Enable column resizing by user
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        
        # Set reasonable initial column widths
        for j, col in enumerate(display_cols):
            if col == "File_path":
                self.table.setColumnWidth(j, 200)
            elif col == "Unique_seg_result":
                self.table.setColumnWidth(j, 150)
            elif col == "Result":
                self.table.setColumnWidth(j, 80)
            elif col == "Background_result":
                self.table.setColumnWidth(j, 120)
            elif col == "model_name":
                self.table.setColumnWidth(j, 100)
            elif col == self.active_label_col:
                self.table.setColumnWidth(j, 100)
        
        # Highlight current row in table
        if self.current_idx < len(self.filtered_indices):
            current_row_idx = self.filtered_indices[self.current_idx]
            print(f"refresh_table: current_idx={self.current_idx}, current_row_idx={current_row_idx}")
            # Calculate the table row position based on smart loading
            if hasattr(self, '_table_start_pos'):
                table_row = self.current_idx - self._table_start_pos
                print(f"Smart loading: table_row={table_row}, start_pos={self._table_start_pos}, table.rowCount()={self.table.rowCount()}")
                if 0 <= table_row < self.table.rowCount():
                    print(f"Selecting table row {table_row}")
                    self.table.selectRow(table_row)
                    # Update _current_row_in_window to match the actual selection
                    self._current_row_in_window = table_row
                    print(f"Updated _current_row_in_window to {table_row}")
                    # Ensure the row is visible (no centering, just make it visible)
                    item = self.table.item(table_row, 0)
                    if item:
                        print(f"Ensuring row {table_row} is visible")
                        self.table.scrollToItem(item, QtWidgets.QAbstractItemView.EnsureVisible)
                else:
                    print(f"Row {table_row} out of range ({self._table_start_pos}-{self._table_end_pos}), but we're already in refresh_table")
                    # Already in refresh_table, don't trigger another reload
            else:
                # Fallback: search for the row
                found = False
                for table_row in range(self.table.rowCount()):
                    item = self.table.item(table_row, 0)
                    if item is not None:
                        original_idx = item.data(QtCore.Qt.UserRole)
                        if original_idx == current_row_idx:
                            self.table.selectRow(table_row)
                            # Update _current_row_in_window to match the actual selection
                            self._current_row_in_window = table_row
                            print(f"Updated _current_row_in_window to {table_row} (fallback)")
                            # Ensure the row is visible (no centering, just make it visible)
                            self.table.scrollToItem(item, QtWidgets.QAbstractItemView.EnsureVisible)
                            found = True
                            break
                
                if not found:
                    print(f"Current row {current_row_idx} not found in fallback search, but we're already in refresh_table")
                    # Don't select anything if we can't find the current row
                    # This prevents jumping to the last row
        
        # Don't restore scroll position during navigation - let the current row centering handle it
        if hasattr(self, '_saved_scroll_pos'):
            # Clear saved position without restoring to avoid conflicts with row centering
            delattr(self, '_saved_scroll_pos')

    def _trigger_smart_table_reload(self) -> None:
        """Trigger smart table reload with loading indicator"""
        # Prevent multiple simultaneous reloads
        if hasattr(self, '_reload_in_progress') and self._reload_in_progress:
            return
        
        self._reload_in_progress = True
        
        # Save current scroll position
        self._saved_scroll_pos = self.table.verticalScrollBar().value()
        
        # Show loading indicator
        self.status.showMessage("테이블 데이터 로딩 중...", 1000)
        
        # Defer the reload to avoid blocking UI
        QtCore.QTimer.singleShot(50, self._smart_table_reload)

    def _smart_table_reload(self) -> None:
        """Perform smart table reload with current row centering"""
        try:
            # Show loading indicator in table
            self.table.setRowCount(1)
            self.table.setColumnCount(1)
            self.table.setHorizontalHeaderLabels(["로딩 중..."])
            
            loading_item = QtWidgets.QTableWidgetItem("🔄 테이블 데이터 로딩 중...")
            loading_item.setTextAlignment(QtCore.Qt.AlignCenter)
            self.table.setItem(0, 0, loading_item)
            
            # Defer actual reload
            QtCore.QTimer.singleShot(100, lambda: self._complete_smart_reload())
            
        except Exception as e:
            print(f"스마트 테이블 리로드 오류: {e}")
            # Fallback to normal refresh
            self._complete_smart_reload()
    
    def _complete_smart_reload(self) -> None:
        """Complete the smart table reload process"""
        try:
            self.refresh_table()
        finally:
            # Always clear the reload flag
            if hasattr(self, '_reload_in_progress'):
                delattr(self, '_reload_in_progress')

    def _check_table_reload_needed(self) -> None:
        """Check if table needs to be reloaded to show current row"""
        print(f"_check_table_reload_needed called with current_idx={self.current_idx}")
        
        if not hasattr(self, '_table_start_pos') or not hasattr(self, '_table_end_pos'):
            # First time, do a full refresh
            print("First time table load, doing full refresh")
            self.refresh_table()
            print("First time refresh completed")
            return
        
        # Check if current row is outside the visible range
        if (self.current_idx < self._table_start_pos or 
            self.current_idx >= self._table_end_pos):
            print(f"Current row {self.current_idx} outside range [{self._table_start_pos}, {self._table_end_pos}), doing full refresh")
            # Current row is not visible, do a simple full refresh instead of complex smart reload
            self.refresh_table()
            print("Range-based refresh completed")
        else:
            print(f"Current row {self.current_idx} is within range [{self._table_start_pos}, {self._table_end_pos}), no reload needed")
            # Even if within range, ensure table selection is correct
            # But first verify that the current row is actually visible in the table
            if hasattr(self, '_current_row_in_window'):
                expected_table_row = self._current_row_in_window
                if 0 <= expected_table_row < self.table.rowCount():
                    # Verify the row contains the expected data
                    item = self.table.item(expected_table_row, 0)
                    if item is not None:
                        original_idx = item.data(QtCore.Qt.UserRole)
                        expected_original_idx = self.filtered_indices[self.current_idx]
                        if original_idx == expected_original_idx:
                            print(f"Table selection is correct at row {expected_table_row}")
                            self.table.selectRow(expected_table_row)
                            return
                        else:
                            print(f"Table row {expected_table_row} has wrong data: {original_idx} vs expected {expected_original_idx}")
                    else:
                        print(f"Table row {expected_table_row} has no item")
                else:
                    print(f"Expected table row {expected_table_row} out of range [0, {self.table.rowCount()})")
            
            # If verification failed, update table selection
            print("Verification failed, updating table selection")
            self._update_table_selection()

    def on_table_click(self, row: int, _column: int) -> None:
        """Handle single click on table"""
        self._handle_table_selection(row)
    
    def on_table_double_click(self, row: int, _column: int) -> None:
        """Handle double click on table"""
        self._handle_table_selection(row)
    
    def _handle_table_selection(self, row: int) -> None:
        """Handle table row selection (both single and double click) - safe version"""
        try:
            if row >= self.table.rowCount():
                return
                    
            print(f"테이블 클릭: 행 {row}")
                
            # Get the original dataframe index from the clicked row
            item = self.table.item(row, 0)
            if item is not None:
                original_idx = item.data(QtCore.Qt.UserRole)
                if original_idx is not None and original_idx in self.filtered_indices:
                    # Find the position in filtered_indices
                    new_current_idx = self.filtered_indices.index(original_idx)
                    if new_current_idx != self.current_idx:
                        print(f"테이블 클릭으로 인덱스 변경: {self.current_idx} → {new_current_idx}")
                        self.current_idx = new_current_idx
                        
                        # Update image based on the selected row
                        print(f"테이블 클릭으로 이미지 로드: original_idx={original_idx}")
                        self._load_image_for_row(original_idx)
                        
                        # Update minimal view to refresh all UI elements
                        self._minimal_view_update()
                        
                        # Note: Don't call refresh_table() here to avoid recursion
            
            # Check if we need to load more data (if we're near the end of visible data)
            self._check_and_load_more_data(row)
                
        except Exception as e:
            print(f"테이블 선택 처리 중 오류: {e}")
            self.status.showMessage(f"테이블 선택 오류: {str(e)}", 3000)

    def _on_table_selection_changed(self) -> None:
        """Handle table selection change events (for all table interactions) - safe version"""
        try:
            if self.df is None or not self.filtered_indices:
                return
            
            # Get current selection
            selected_items = self.table.selectedItems()
            if not selected_items:
                return
            
            # Get the row of the first selected item
            selected_row = selected_items[0].row()
            
            # Get the original dataframe index for this table row
            item = self.table.item(selected_row, 0)
            if item is None:
                return
            
            original_idx = item.data(QtCore.Qt.UserRole)
            if original_idx is None:
                return
            
            # Find the index in filtered_indices
            try:
                new_current_idx = self.filtered_indices.index(original_idx)
                if new_current_idx != self.current_idx:
                    print(f"테이블 선택 변경: {self.current_idx} → {new_current_idx} (행: {selected_row})")
                    
                    # Only update if this is a user-initiated change, not programmatic
                    # Check if this is likely a user click (not from navigation)
                    if not hasattr(self, '_navigation_in_progress') or not self._navigation_in_progress:
                        self.current_idx = new_current_idx
                        
                        # Load image for the new selection
                        print(f"테이블 선택으로 이미지 로드: original_idx={original_idx}")
                        self._load_image_for_row(original_idx)
                        
                        # Update minimal view to refresh all UI elements
                        self._minimal_view_update()
                        
                        self.settings.setValue("current_idx", self.current_idx)
                        
                        # Update status
                        self.status.showMessage(f"테이블 선택: {self.current_idx + 1}/{len(self.filtered_indices)}", 1000)
                    else:
                        print("Ignoring table selection change during navigation")
                    
                    # Don't call _update_table_selection here to avoid circular calls
                    # The table selection is already correct from the user's click
            except ValueError:
                pass  # Index not found in filtered indices
                
        except Exception as e:
            print(f"테이블 선택 변경 처리 중 오류: {e}")
            self.status.showMessage(f"테이블 선택 변경 오류: {str(e)}", 3000)

    def on_prev(self) -> None:
        """Navigate to previous item - optimized for speed"""
        print("on_prev called")  # Debug log
        if not self.filtered_indices:
            print("No filtered indices available")
            return
            
        if self.current_idx > 0:
            print(f"Moving from {self.current_idx} to {self.current_idx - 1}")
            self.current_idx -= 1
            print(f"Current idx now: {self.current_idx}")
            
            # Validate current_idx is within bounds
            if self.current_idx >= len(self.filtered_indices):
                print(f"ERROR: current_idx {self.current_idx} out of bounds [0, {len(self.filtered_indices)})")
                self.current_idx = len(self.filtered_indices) - 1
                print(f"Corrected current_idx to {self.current_idx}")
            
            try:
                # Set navigation flag to prevent table selection interference
                self._navigation_in_progress = True
                
                print("About to call _check_table_reload_needed")
                self._check_table_reload_needed()
                print("_check_table_reload_needed completed")
                
                print("About to call _minimal_view_update")
                self._minimal_view_update()
                print("_minimal_view_update completed")
                
                print("About to call _update_table_selection")
                self._update_table_selection()
                print("_update_table_selection completed")
                
                # Save position change
                self.settings.setValue("current_idx", self.current_idx)
                # Show navigation feedback
                self.status.showMessage(f"이전 항목으로 이동: {self.current_idx + 1}/{len(self.filtered_indices)}", 1000)
                print("on_prev completed successfully")
            except Exception as e:
                print(f"Error in on_prev: {e}")
                import traceback
                traceback.print_exc()
            finally:
                # Clear navigation flag
                self._navigation_in_progress = False
        else:
            self.status.showMessage("첫 번째 항목입니다", 1000)

    def on_next(self) -> None:
        """Navigate to next item - optimized for speed"""
        print("on_next called")  # Debug log
        if not self.filtered_indices:
            print("No filtered indices available")
            return
            
        if self.current_idx < len(self.filtered_indices) - 1:
            print(f"Moving from {self.current_idx} to {self.current_idx + 1}")
            self.current_idx += 1
            print(f"Current idx now: {self.current_idx}")
            
            # Validate current_idx is within bounds
            if self.current_idx >= len(self.filtered_indices):
                print(f"ERROR: current_idx {self.current_idx} out of bounds [0, {len(self.filtered_indices)})")
                self.current_idx = len(self.filtered_indices) - 1
                print(f"Corrected current_idx to {self.current_idx}")
            
            try:
                # Set navigation flag to prevent table selection interference
                self._navigation_in_progress = True
                
                print("About to call _check_table_reload_needed")
                self._check_table_reload_needed()
                print("_check_table_reload_needed completed")
                
                print("About to call _minimal_view_update")
                self._minimal_view_update()
                print("_minimal_view_update completed")
                
                print("About to call _update_table_selection")
                self._update_table_selection()
                print("_update_table_selection completed")
                
                # Save position change
                self.settings.setValue("current_idx", self.current_idx)
                # Show navigation feedback
                self.status.showMessage(f"다음 항목으로 이동: {self.current_idx + 1}/{len(self.filtered_indices)}", 1000)
                print("on_next completed successfully")
            except Exception as e:
                print(f"Error in on_next: {e}")
                import traceback
                traceback.print_exc()
            finally:
                # Clear navigation flag
                self._navigation_in_progress = False
        else:
            self.status.showMessage("마지막 항목입니다", 1000)

    def on_export_labels(self) -> None:
        """Export labels to Excel"""
        if self.df is None:
            return
            
        output_path = self.csv_path.replace('.csv', '_labeled.xlsx')
        try:
            # Export to Excel
            self.df.to_excel(output_path, index=False, sheet_name="labeled_results")
            QtWidgets.QMessageBox.information(self, "내보내기 완료", f"라벨이 다음 경로로 내보내짐: {output_path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "내보내기 오류", f"내보내기 실패: {str(e)}")

    # Memory management methods
    def _clear_image_cache(self) -> None:
        """Clear the image cache"""
        self._image_cache.clear()
        force_garbage_collection()
        self.status.showMessage("이미지 캐시 삭제됨")

    def _show_memory_info(self) -> None:
        """Show memory usage information"""
        memory_mb = get_memory_usage()
        system_mb = get_system_memory()
        cache_size = len(self._image_cache)
        
        msg = f"메모리 사용량: {memory_mb:.1f} MB\n"
        msg += f"시스템 메모리: {system_mb:.1f} MB\n"
        msg += f"이미지 캐시: {cache_size} 개\n"
        msg += f"메모리 제한: {self.max_memory_mb:.1f} MB"
        
        QtWidgets.QMessageBox.information(self, "메모리 정보", msg)

    def _force_memory_cleanup(self) -> None:
        """Force memory cleanup - enhanced for large datasets"""
        self._clear_image_cache()
        self._flush_pending_ops()
        
        # Clear filter cache to save memory
        self._filter_cache = None
        self._last_filter_hash = None
        
        # Force garbage collection multiple times for better cleanup
        for _ in range(3):
            force_garbage_collection()
            QtWidgets.QApplication.processEvents()
        
        memory_after = get_memory_usage()
        self.status.showMessage(f"메모리 정리 완료. 사용량: {memory_after:.1f} MB")

    def _show_performance_stats(self) -> None:
        """Show performance statistics"""
        if self.df is None:
            return
        
        session_duration = datetime.now() - self._session_start_time
        hours = session_duration.total_seconds() / 3600
        
        total_rows = len(self.df)
        labeled_rows = len(self.df[~(self.df[self.active_label_col].isna() | (self.df[self.active_label_col] == ""))])
        progress = (labeled_rows / total_rows * 100) if total_rows > 0 else 0
        
        labels_per_hour = self._label_count / hours if hours > 0 else 0
        
        msg = f"성능 통계:\n"
        msg += f"세션 시간: {session_duration}\n"
        msg += f"이번 세션 라벨링: {self._label_count:,}개\n"
        msg += f"시간당 라벨링 속도: {labels_per_hour:.1f}개/시간\n"
        msg += f"전체 진행률: {labeled_rows:,}/{total_rows:,} ({progress:.1f}%)\n"
        msg += f"남은 예상 시간: {((total_rows - labeled_rows) / labels_per_hour):.1f}시간" if labels_per_hour > 0 else "남은 시간: 계산 불가"
        
        QtWidgets.QMessageBox.information(self, "성능 통계", msg)

    def save_session_state(self) -> None:
        """Save current session state to settings"""
        if self.df is None:
            return
            
        try:
            # Save current position and filter state
            self.settings.setValue("current_idx", self.current_idx)
            self.settings.setValue("auto_advance_enabled", self.auto_advance_enabled)
            
            # Save filter settings
            self.settings.setValue("label_state", self.cmb_label_state.currentText())
            self.settings.setValue("label_value", self.cmb_label_value.currentText())
            self.settings.setValue("model_name", self.cmb_model_name.currentText())
            self.settings.setValue("result_filter", self.cmb_result_filter.currentText())
            self.settings.setValue("bookmarks_only", self.chk_bookmarks.isChecked())
            
            # Save pred filter selections
            self.settings.setValue("selected_pred_filters", list(self.selected_pred_filters))
            
            # Save UI section states (collapsed/expanded)
            self.settings.setValue("quick_labeling_expanded", self.quick_labeling_container.isVisible())
            self.settings.setValue("as_is_tobe_expanded", self.as_is_tobe_container.isVisible())
            self.settings.setValue("basic_filters_expanded", self.basic_filters_widget.isVisible())
            self.settings.setValue("pred_filters_expanded", self.pred_filters_container.isVisible())
            # Bookmark memo container no longer exists - skip this setting
            
            # Save window geometry
            self.settings.setValue("geometry", self.saveGeometry())
            self.settings.setValue("window_state", self.saveState())
            
            print("세션 상태 저장 완료")
            
        except Exception as e:
            print(f"세션 저장 오류: {e}")

    def restore_session_state(self) -> None:
        """Restore session state from settings"""
        if self.df is None:
            return
            
        try:
            # Restore current position
            saved_idx = self.settings.value("current_idx", 0, type=int)
            if 0 <= saved_idx < len(self.filtered_indices):
                self.current_idx = saved_idx
            
            # Restore auto-advance setting
            auto_advance = self.settings.value("auto_advance_enabled", True, type=bool)
            self.auto_advance_enabled = auto_advance
            self.chk_auto_advance.setChecked(auto_advance)
            
            # Restore filter settings
            label_state = self.settings.value("label_state", "전체", type=str)
            if label_state in ["전체", "라벨됨", "라벨안됨"]:
                self.cmb_label_state.setCurrentText(label_state)
            
            label_value = self.settings.value("label_value", "전체", type=str)
            idx = self.cmb_label_value.findText(label_value)
            if idx >= 0:
                self.cmb_label_value.setCurrentIndex(idx)
            
            model_name = self.settings.value("model_name", "전체", type=str)
            idx = self.cmb_model_name.findText(model_name)
            if idx >= 0:
                self.cmb_model_name.setCurrentIndex(idx)

            result_filter = self.settings.value("result_filter", "전체", type=str)
            idx = self.cmb_result_filter.findText(result_filter)
            if idx >= 0:
                self.cmb_result_filter.setCurrentIndex(idx)
            
            bookmarks_only = self.settings.value("bookmarks_only", False, type=bool)
            self.chk_bookmarks.setChecked(bookmarks_only)
            
            # Restore pred filter selections
            saved_pred_filters = self.settings.value("selected_pred_filters", [], type=list)
            if saved_pred_filters:
                self.selected_pred_filters = set(saved_pred_filters)
                # Update checkboxes
                for choice, checkbox in self.pred_filter_checkboxes.items():
                    checkbox.setChecked(choice in self.selected_pred_filters)
            
            # UI sections are now always visible - no need to restore
            
            # Restore window geometry
            geometry = self.settings.value("geometry")
            if geometry:
                self.restoreGeometry(geometry)
            
            window_state = self.settings.value("window_state")
            if window_state:
                self.restoreState(window_state)
            
            # Apply filters and refresh view
            self.apply_filters()
            self.refresh_view()
            
            print(f"세션 상태 복원 완료 - 위치: {self.current_idx}")
            
        except Exception as e:
            print(f"세션 복원 오류: {e}")

    def _cleanup_ui_objects(self) -> None:
        """Clean up UI objects safely to prevent memory issues"""
        try:
            # Clear layout objects safely
            if hasattr(self, 'choice_buttons_layout') and self.choice_buttons_layout:
                self._safe_clear_layout(self.choice_buttons_layout)

            if hasattr(self, 'pred_filter_checkboxes_layout') and self.pred_filter_checkboxes_layout:
                self._safe_clear_layout(self.pred_filter_checkboxes_layout)

            if hasattr(self, 'as_is_tobe_layout') and self.as_is_tobe_layout:
                self._safe_clear_layout(self.as_is_tobe_layout)

            # Clear image cache
            if hasattr(self, 'image_cache'):
                self.image_cache.clear()

        except Exception as e:
            print(f"UI cleanup error: {e}")


    def closeEvent(self, event) -> None:
        """Handle application close event"""
        try:
            # Save session state before closing
            self.save_session_state()

            # Flush any pending JSON operations
            self._flush_pending_ops()

            # Clean up UI objects safely
            self._cleanup_ui_objects()

            event.accept()

        except Exception as e:
            print(f"종료 시 저장 오류: {e}")
            event.accept()


# 전역 상태 바 변수
global_status_bar = None

def _get_saved_settings_from_qsettings() -> dict:
    """Load last used paths from QSettings for direct UI testing."""
    try:
        settings = QtCore.QSettings("rtm", "inference_labeler")
        csv_path = settings.value("last_csv_path", "", type=str)
        images_base = settings.value("last_images_base", "", type=str)
        json_base = settings.value("last_json_base", "", type=str)
        csv_type = settings.value("last_csv_type", "report", type=str)
        return {
            "csv_path": csv_path,
            "images_base": images_base,
            "json_base": json_base,
            "csv_type": csv_type,
        }
    except Exception as e:
        print(f"❌ QSettings 로드 오류: {e}")
        return {
            "csv_path": "",
            "images_base": "",
            "json_base": "",
            "csv_type": "report",
        }

def main():
    global global_status_bar
    # If launched with --use-saved, open main UI directly with saved settings (bypass SetupWindow)
    if "--use-saved" in sys.argv:
        print("🚀 --use-saved 플래그 감지: 저장된 설정으로 바로 메인 UI 실행")
        app = QtWidgets.QApplication(sys.argv)
        saved = _get_saved_settings_from_qsettings()
        print(f"📊 저장된 설정: CSV={saved['csv_path']}, 이미지={saved['images_base']}, JSON={saved['json_base']}, 타입={saved['csv_type']}")
        if not saved["csv_path"]:
            print("❌ 저장된 CSV 경로가 없습니다. 기본 흐름으로 진행합니다.")
        else:
            window = InferenceLabelerWindow(saved)
            global_status_bar = window.status
            window.show()
            return sys.exit(app.exec())
    print("🚀 애플리케이션 시작")
    app = QtWidgets.QApplication(sys.argv)
    print("✅ QApplication 생성 완료")

    # 설정 창 표시
    print("📋 설정 창 생성 중...")
    setup_window = SetupWindow()
    print("✅ 설정 창 생성 완료")

    print("🔍 설정 창 실행...")
    result = setup_window.exec_()
    print(f"📊 설정 창 결과: {result} (Accepted={QtWidgets.QDialog.Accepted})")

    if result != QtWidgets.QDialog.Accepted:
        # 사용자가 취소한 경우 종료
        print("❌ 사용자가 취소함")
        setup_window.deleteLater()
        return

    # 설정값 가져오기
    print("📋 설정값 가져오기...")
    settings = setup_window.get_settings()
    print(f"📊 설정값: CSV={settings['csv_path']}, 이미지={settings['images_base']}, JSON={settings['json_base']}, 타입={settings['csv_type']}")

    # 설정된 경로 확인
    print("🔍 경로 존재 여부 확인...")
    if not os.path.exists(settings["csv_path"]):
        print(f"❌ CSV 파일 없음: {settings['csv_path']}")
        QtWidgets.QMessageBox.critical(None, "오류", f"CSV 파일을 찾을 수 없음: {settings['csv_path']}")
        setup_window.deleteLater()
        return

    # 전역 상태 바에 메시지 표시
    if global_status_bar:
        global_status_bar.showMessage("🔍 경로 검증 중...")

    if not os.path.exists(settings["images_base"]):
        print(f"⚠️ 이미지 디렉토리 없음: {settings['images_base']}")
        QtWidgets.QMessageBox.warning(None, "경고", f"이미지 디렉토리를 찾을 수 없음: {settings['images_base']}")

    if not os.path.exists(settings["json_base"]):
        print(f"⚠️ JSON 디렉토리 없음: {settings['json_base']}")
        QtWidgets.QMessageBox.warning(None, "경고", f"JSON 디렉토리를 찾을 수 없음: {settings['json_base']}")

    print("💾 설정값 저장 중...")
    # 설정 창에서 경로 설정을 QSettings에 저장
    setup_window.save_paths_to_settings()

    # 설정 창 정리
    print("🧹 설정 창 정리 중...")
    setup_window.deleteLater()

    # 라벨링 창 표시
    print("🏠 메인 라벨링 창 생성 중...")
    window = InferenceLabelerWindow(settings)
    print("✅ 메인 라벨링 창 생성 완료")

    # 전역 상태 바 설정
    global_status_bar = window.status

    # 초기 상태 메시지
    if global_status_bar:
        global_status_bar.showMessage("메인 창 초기화 완료 - 데이터를 로드하는 중...")

    print("🖥️ 메인 창 표시...")
    window.show()
    print("✅ 메인 창 표시 완료")

    print("🎯 이벤트 루프 시작...")
    sys.exit(app.exec())


if __name__ == "__main__":
    main()