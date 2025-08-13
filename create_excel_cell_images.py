#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import os
import locale

# 한글 자소 분리 문제 해결을 위한 인코딩 설정
if sys.platform.startswith('darwin'):  # macOS
    os.environ['LC_ALL'] = 'en_US.UTF-8'
    os.environ['LANG'] = 'en_US.UTF-8'
try:
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_ALL, 'C.UTF-8')
    except:
        pass

import glob
import csv
import unicodedata
from openpyxl import Workbook
from openpyxl.drawing import image as openpyxl_image
from openpyxl.styles import Alignment
from PIL import Image as PILImage
import uuid

def load_inference_results(base_path):
    """Load all inference_results.csv files and create a lookup dictionary"""
    print("=== 추론 결과 파일 검색 시작 ===")
    csv_files = glob.glob(os.path.join(base_path, "**", "inference_results.csv"), recursive=True)
    print(f"발견된 CSV 파일 수: {len(csv_files)}")
    
    results_dict = {}
    
    for i, csv_file in enumerate(csv_files):
        try:
            print(f"[{i+1}/{len(csv_files)}] CSV 파일 로딩 중: {csv_file}")
            row_count = 0
            with open(csv_file, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    filename = row.get('filename', '').replace('.bmp', '')
                    results_dict[filename] = {
                        'gt_status': row.get('gt_status', ''),
                        'pred_status': row.get('pred_status', ''),
                        'dominant_class': row.get('dominant_class', ''),
                        'csv_source': unicodedata.normalize('NFC', os.path.basename(os.path.dirname(csv_file)))
                    }
                    row_count += 1
            print(f"   └── {row_count}개 결과 로딩 완료")
        except Exception as e:
            print(f"   └── 오류 발생: {e}")
    
    print(f"=== 총 {len(results_dict)}개 파일의 추론 결과 로딩 완료 ===\n")
    return results_dict

def find_image_pairs(base_path):
    """Find image pairs (viz.png and .bmp files) in the given path"""
    print("=== 이미지 파일 검색 시작 ===")
    print("PNG 파일 (viz) 검색 중...")
    png_files = glob.glob(os.path.join(base_path, "**", "images", "**", "*_viz.png"), recursive=True)
    print(f"발견된 PNG 파일 수: {len(png_files)}")
    
    print("BMP 파일 검색 중...")
    bmp_files = glob.glob(os.path.join(base_path, "**", "images", "**", "*.bmp"), recursive=True)
    print(f"발견된 BMP 파일 수: {len(bmp_files)}")
    
    print("이미지 쌍 매칭 중...")
    pairs = []
    
    for i, png_file in enumerate(png_files):
        if i % 100 == 0:
            print(f"   진행률: {i}/{len(png_files)} PNG 파일 처리 중...")
        
        # Extract base name by removing "_viz.png"
        base_name = os.path.basename(png_file).replace("_viz.png", "")
        
        # Find corresponding .bmp file
        bmp_file = None
        for bmp in bmp_files:
            if os.path.basename(bmp).replace(".bmp", "") == base_name:
                bmp_file = bmp
                break
        
        if bmp_file:
            pairs.append({
                'filename': base_name,
                'viz_img_path': png_file,
                'img_path': bmp_file
            })
    
    print(f"=== 총 {len(pairs)}개 이미지 쌍 매칭 완료 ===\n")
    return pairs

def create_excel_with_cell_images(pairs, results_dict, output_file):
    """Create Excel file with images positioned within cells for filtering/sorting"""
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Image Pairs with Results"
        
        # Set headers
        ws['A1'] = 'filename'
        ws['B1'] = 'img'
        ws['C1'] = 'viz_img'
        ws['D1'] = 'gt_status'
        ws['E1'] = 'pred_status'
        ws['F1'] = 'dominant_class'
        ws['G1'] = 'csv_source'
        
        # Set column widths to fit images properly
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 30  # Width for BMP image
        ws.column_dimensions['C'].width = 30  # Width for PNG image
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 30
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        current_row = 2
        temp_files = []  # Track temporary files for cleanup
        
        print("=== 엑셀 파일에 이미지 및 결과 삽입 시작 ===")
        total_pairs = len(pairs)
        
        for i, pair in enumerate(pairs):
                
            try:
                print(f"[{i+1}/{total_pairs}] 처리 중: {pair['filename']}")
                
                # Add filename with proper Unicode normalization
                normalized_filename = unicodedata.normalize('NFC', pair['filename'])
                ws[f'A{current_row}'] = normalized_filename
                print(f"   └── 파일명 추가: {normalized_filename}")
                
                # Add inference results if available
                if pair['filename'] in results_dict:
                    result = results_dict[pair['filename']]
                    ws[f'D{current_row}'] = unicodedata.normalize('NFC', result['gt_status'])
                    ws[f'E{current_row}'] = unicodedata.normalize('NFC', result['pred_status'])
                    ws[f'F{current_row}'] = unicodedata.normalize('NFC', result['dominant_class'])
                    ws[f'G{current_row}'] = unicodedata.normalize('NFC', result['csv_source'])
                    print(f"   └── 추론 결과 추가: GT={result['gt_status']}, Pred={result['pred_status']}, Class={result['dominant_class']}")
                else:
                    ws[f'D{current_row}'] = 'Not found'
                    ws[f'E{current_row}'] = 'Not found'
                    ws[f'F{current_row}'] = 'Not found'
                    ws[f'G{current_row}'] = 'Not found'
                    print(f"   └── 추론 결과 없음")
                
                # Set row height for images (200 pixels for better cell fit)
                ws.row_dimensions[current_row].height = 150
                
                # Add BMP image (img column) - positioned to fit within cell
                print(f"   └── BMP 이미지 처리 중: {pair['img_path']}")
                if os.path.exists(pair['img_path']):
                    try:
                        # Create unique temporary filename in current directory
                        temp_bmp_path = f"temp_bmp_{uuid.uuid4().hex[:8]}.png"
                        temp_files.append(temp_bmp_path)
                        
                        # Resize BMP image to fit in cell
                        with PILImage.open(pair['img_path']) as pil_img:
                            original_size = pil_img.size
                            # Convert to RGB if necessary
                            if pil_img.mode != 'RGB':
                                pil_img = pil_img.convert('RGB')
                            # Resize to fit cell dimensions (200x150 pixels)
                            pil_img.thumbnail((200, 150), PILImage.Resampling.LANCZOS)
                            new_size = pil_img.size
                            pil_img.save(temp_bmp_path, "PNG")
                        
                        # Add to Excel with positioning to fit within cell bounds
                        img = openpyxl_image.Image(temp_bmp_path)
                        img.width = new_size[0]
                        img.height = new_size[1]
                        
                        # Position image within the cell (with small margin)
                        cell_ref = f'B{current_row}'
                        img.anchor = cell_ref
                        
                        # Add some offset to position within cell
                        ws.add_image(img)
                        
                        print(f"       └── BMP 이미지 삽입 완료 ({original_size[0]}x{original_size[1]} → {new_size[0]}x{new_size[1]})")
                        
                    except Exception as e:
                        ws[f'B{current_row}'] = f"Error: {str(e)}"
                        print(f"       └── BMP 이미지 처리 오류: {e}")
                else:
                    ws[f'B{current_row}'] = "BMP not found"
                    print(f"       └── BMP 파일 없음")
                
                # Add PNG image (viz_img column) - positioned to fit within cell
                print(f"   └── PNG 이미지 처리 중: {pair['viz_img_path']}")
                if os.path.exists(pair['viz_img_path']):
                    try:
                        # Create unique temporary filename in current directory
                        temp_png_path = f"temp_png_{uuid.uuid4().hex[:8]}.png"
                        temp_files.append(temp_png_path)
                        
                        # Resize PNG image to fit in cell
                        with PILImage.open(pair['viz_img_path']) as pil_img:
                            original_size = pil_img.size
                            # Convert to RGB if necessary
                            if pil_img.mode != 'RGB':
                                pil_img = pil_img.convert('RGB')
                            # Resize to fit cell dimensions (200x150 pixels)
                            pil_img.thumbnail((200, 150), PILImage.Resampling.LANCZOS)
                            new_size = pil_img.size
                            pil_img.save(temp_png_path, "PNG")
                        
                        # Add to Excel with positioning to fit within cell bounds
                        img = openpyxl_image.Image(temp_png_path)
                        img.width = new_size[0]
                        img.height = new_size[1]
                        
                        # Position image within the cell
                        cell_ref = f'C{current_row}'
                        img.anchor = cell_ref
                        
                        ws.add_image(img)
                        
                        print(f"       └── PNG 이미지 삽입 완료 ({original_size[0]}x{original_size[1]} → {new_size[0]}x{new_size[1]})")
                        
                    except Exception as e:
                        ws[f'C{current_row}'] = f"Error: {str(e)}"
                        print(f"       └── PNG 이미지 처리 오류: {e}")
                else:
                    ws[f'C{current_row}'] = "PNG not found"
                    print(f"       └── PNG 파일 없음")
                
                current_row += 1
                print(f"   └── 행 {current_row-1} 처리 완료\n")
                
            except Exception as e:
                print(f"   └── 쌍 처리 오류: {e}\n")
                continue
        
        # Enable AutoFilter for the data range
        ws.auto_filter.ref = f"A1:G{current_row-1}"
        print("=== 자동 필터 설정 완료 ===")
        
        # Save workbook
        print("엑셀 파일 저장 중...")
        wb.save(output_file)
        
        # Clean up temporary files
        print("임시 파일 정리 중...")
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except Exception as e:
                print(f"임시 파일 정리 오류 {temp_file}: {e}")
        
        print(f"Excel file created successfully: {output_file}")
        print(f"Total pairs processed: {min(len(pairs), 50)}")
        print("필터 및 정렬 기능이 활성화되었습니다!")
        
        return True
        
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        # Clean up temporary files even on error
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except:
                pass
        return False

def main():
    base_path = "/Users/rtm/Downloads/v0.3"
    output_file = "/Users/rtm/Downloads/v0.3/image_pairs_with_filter.xlsx"
    
    print("추론 결과 로딩 중...")
    results_dict = load_inference_results(base_path)
    
    print("이미지 쌍 검색 중...")
    pairs = find_image_pairs(base_path)
    
    print(f"발견된 이미지 쌍: {len(pairs)}개")
    
    if pairs:
        print("필터링 가능한 엑셀 파일 생성 중...")
        success = create_excel_with_cell_images(pairs, results_dict, output_file)
        
        if success:
            print(f"Excel file created: {output_file}")
            print(f"파일이 성공적으로 생성되었습니다: {output_file}")
            print("이제 Excel에서 필터와 정렬 기능을 사용할 수 있습니다!")
        else:
            print("Failed to create Excel file")
    else:
        print("No image pairs found")

if __name__ == "__main__":
    main()