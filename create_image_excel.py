#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import glob
import sys

def find_image_pairs(base_path):
    """Find image pairs (viz.png and .bmp files) in the given path"""
    png_files = glob.glob(os.path.join(base_path, "**", "images", "**", "*_viz.png"), recursive=True)
    bmp_files = glob.glob(os.path.join(base_path, "**", "images", "**", "*.bmp"), recursive=True)
    
    pairs = []
    
    for png_file in png_files:
        # Extract base name by removing "_viz.png"
        base_name = os.path.basename(png_file).replace("_viz.png", "")
        
        # Find corresponding .bmp file
        bmp_file = None
        for bmp in bmp_files:
            if os.path.basename(bmp).replace(".bmp", "") == base_name:
                bmp_file = bmp
                break
        
        if bmp_file:
            pairs.append({
                'filename': base_name,
                'viz_img_path': png_file,
                'img_path': bmp_file
            })
    
    return pairs

def create_excel_with_images(pairs, output_file):
    """Create Excel file with images"""
    try:
        from openpyxl import Workbook
        from openpyxl.drawing import image as openpyxl_image
        from openpyxl.utils import get_column_letter
        from PIL import Image as PILImage
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Image Pairs"
        
        # Set headers
        ws['A1'] = 'filename'
        ws['B1'] = 'img'
        ws['C1'] = 'viz_img'
        
        # Set column widths for images
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40
        
        # Set row height for header
        ws.row_dimensions[1].height = 20
        
        current_row = 2
        
        for pair in pairs:
            try:
                # Add filename
                ws[f'A{current_row}'] = pair['filename']
                
                # Set row height for images
                ws.row_dimensions[current_row].height = 200
                
                # Add BMP image (img column)
                if os.path.exists(pair['img_path']):
                    try:
                        # Try to load and resize image
                        pil_img = PILImage.open(pair['img_path'])
                        # Resize to fit in cell (approximately 300px width)
                        pil_img.thumbnail((300, 300))
                        
                        # Save temporary resized image
                        temp_img_path = f"./temp_img_{current_row}.png"
                        pil_img.save(temp_img_path, "PNG")
                        
                        img = openpyxl_image.Image(temp_img_path)
                        img.anchor = f'B{current_row}'
                        ws.add_image(img)
                        
                        # Clean up temp file
                        os.remove(temp_img_path)
                    except Exception as e:
                        ws[f'B{current_row}'] = f"Image load error: {str(e)}"
                else:
                    ws[f'B{current_row}'] = "Image not found"
                
                # Add PNG image (viz_img column)
                if os.path.exists(pair['viz_img_path']):
                    try:
                        # Try to load and resize image
                        pil_img = PILImage.open(pair['viz_img_path'])
                        # Resize to fit in cell (approximately 300px width)
                        pil_img.thumbnail((300, 300))
                        
                        # Save temporary resized image
                        temp_img_path = f"./temp_viz_{current_row}.png"
                        pil_img.save(temp_img_path, "PNG")
                        
                        img = openpyxl_image.Image(temp_img_path)
                        img.anchor = f'C{current_row}'
                        ws.add_image(img)
                        
                        # Clean up temp file
                        os.remove(temp_img_path)
                    except Exception as e:
                        ws[f'C{current_row}'] = f"Image load error: {str(e)}"
                else:
                    ws[f'C{current_row}'] = "Image not found"
                
                current_row += 1
                
            except Exception as e:
                print(f"Error processing pair {pair['filename']}: {e}")
                continue
        
        # Save workbook
        wb.save(output_file)
        print(f"Excel file created successfully: {output_file}")
        print(f"Total pairs processed: {len(pairs)}")
        
    except ImportError as e:
        print(f"Required library not found: {e}")
        print("Please install required packages: pip install openpyxl pillow")
        return False
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return False
    
    return True

def main():
    base_path = "/Users/rtm/Downloads/v0.3"
    output_file = "/Users/rtm/Downloads/v0.3/image_pairs.xlsx"
    
    print("Searching for image pairs...")
    pairs = find_image_pairs(base_path)
    
    print(f"Found {len(pairs)} image pairs")
    
    if pairs:
        print("Creating Excel file with images...")
        success = create_excel_with_images(pairs, output_file)
        
        if success:
            print(f"Excel file created: {output_file}")
        else:
            print("Failed to create Excel file")
    else:
        print("No image pairs found")

if __name__ == "__main__":
    main()