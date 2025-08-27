#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import json
import hashlib
import re
from datetime import datetime
try:
    # Python 3.11+
    from datetime import UTC as _UTC
except Exception:
    from datetime import timezone as _tz
    _UTC = _tz.utc
from typing import Dict, List, Optional, Tuple
import argparse
import gc
import psutil

import pandas as pd
from PySide6 import QtCore, QtGui, QtWidgets

from openpyxl import load_workbook

# Reuse path resolution from the existing module
from create_excel_from_seg_csv import resolve_image_path, normalize_relative_path
import glob


# Memory management utilities
def get_memory_usage():
    """Get current memory usage in MB"""
    try:
        process = psutil.Process(os.getpid())
        return process.memory_info().rss / 1024 / 1024
    except:
        return 0

def check_memory_limit(limit_mb=1024):  # Reduced from 2048 to 1024
    """Check if memory usage exceeds limit"""
    return get_memory_usage() > limit_mb

def force_garbage_collection():
    """Force garbage collection to free memory"""
    gc.collect()

def get_system_memory():
    """Get total system memory in MB"""
    try:
        return psutil.virtual_memory().total / 1024 / 1024
    except:
        return 8192  # Default to 8GB if can't detect


def parse_pred_list(value) -> List[str]:
    """Parse pred_seg_results value into a list of strings.
    Handles JSON arrays, python-like list strings, or comma-separated strings.
    """
    try:
        if isinstance(value, (list, tuple, set)):
            return [str(x).strip() for x in value]
        s = str(value).strip()
        if not s:
            return []
        # Try JSON first
        if s.startswith("[") or s.startswith("{"):
            try:
                data = json.loads(s)
                if isinstance(data, (list, tuple, set)):
                    return [str(x).strip() for x in data]
            except Exception:
                pass
        # Fallback: strip brackets and split by semicolon/comma variants
        s2 = s.strip().strip('[](){}')
        parts = [p.strip().strip("'\"") for p in re.split(r"[;,\uFF1B\uFF0C]+", s2) if p.strip()]
        return parts
    except Exception:
        return []


def ensure_object_dtype(df: pd.DataFrame, column: str) -> None:
    try:
        df[column] = df[column].astype("object")
    except Exception:
        pass


def default_json_path(xlsx_path: str) -> str:
    base = os.path.basename(xlsx_path)
    root, _ = os.path.splitext(base)
    parent = os.path.dirname(xlsx_path) or os.getcwd()
    # Preferred naming: if Excel is *_labeled.xlsx → JSON is *_labels.json
    if root.endswith("_labeled"):
        new_root = root[: -len("_labeled")] + "_labels"
    else:
        new_root = root + "_labels"
    new_path = os.path.join(parent, f"{new_root}.json")
    # Legacy path compatibility: previously always appended _labels
    legacy_path = os.path.join(parent, f"{root}_labels.json")
    # If legacy exists and new doesn't, keep using legacy for seamless migration
    if os.path.exists(legacy_path) and not os.path.exists(new_path):
        return legacy_path
    return new_path


def load_label_store(json_path: str) -> dict:
    if not json_path or not os.path.exists(json_path):
        return {"version": 1, "updated_at": None, "labels": {}}
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, dict) and "labels" in data:
                return data
    except Exception:
        pass
    return {"version": 1, "updated_at": None, "labels": {}}


def save_label_store(json_path: str, store: dict) -> bool:
    """Atomically persist the label store. Returns True on success."""
    try:
        store["updated_at"] = datetime.now(_UTC).isoformat()
        tmp = json_path + ".tmp"
        os.makedirs(os.path.dirname(json_path) or ".", exist_ok=True)
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(store, f, ensure_ascii=False, indent=2)
        os.replace(tmp, json_path)
        return True
    except Exception:
        return False


def apply_json_to_excel(json_path: str, xlsx_path: str, sheet_name: str, col_indices: Dict[str, int], df: pd.DataFrame) -> int:
    store = load_label_store(json_path)
    labels = store.get("labels", {})
    applied = 0
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]
    for key, entry in labels.items():
        try:
            row_idx = int(key)
        except Exception:
            continue
        excel_row = row_idx + 2
        values = entry.get("values", {})
        for col_name, val in values.items():
            idx = col_indices.get(col_name)
            if idx is None:
                continue
            try:
                ws.cell(row=excel_row, column=idx, value=val)
                applied += 1
                if col_name in df.columns:
                    df.at[row_idx, col_name] = val
            except Exception:
                pass
    wb.save(xlsx_path)
    wb.close()
    return applied


def get_json_entry(json_path: str, row_idx: int) -> dict:
    store = load_label_store(json_path)
    key = str(row_idx)
    return store.get("labels", {}).get(key) or {}


def upsert_json_entry(json_path: str, row_idx: int, updater: Dict[str, object]) -> None:
    store = load_label_store(json_path)
    key = str(row_idx)
    entry = store["labels"].get(key) or {}
    for k, v in updater.items():
        entry[k] = v
    store["labels"][key] = entry
    save_label_store(json_path, store)


def merge_json_into_df(json_path: str, df: pd.DataFrame, label_columns: List[str]) -> None:
    """Load existing JSON labels and reflect into DataFrame so work can resume after restart.

    - Fills only empty cells to avoid overwriting values already present in Excel/DF
    - Supports legacy column alias mapping: review_label -> review_label_inf
    """
    store = load_label_store(json_path)
    labels = store.get("labels", {})
    for key, entry in labels.items():
        try:
            ridx = int(key)
        except Exception:
            continue
        if ridx not in df.index:
            continue
        values = entry.get("values", {})
        for col, val in values.items():
            target_col = col
            if target_col not in label_columns and col == "review_label" and "review_label_inf" in label_columns:
                target_col = "review_label_inf"
            if target_col in label_columns:
                try:
                    curr = None
                    try:
                        curr = df.at[ridx, target_col]
                    except Exception:
                        curr = None
                    is_empty = (curr is None) or (str(curr) == "")
                    try:
                        if not is_empty:
                            is_empty = bool(pd.isna(curr))  # type: ignore
                    except Exception:
                        pass
                    if is_empty:
                        df.at[ridx, target_col] = val
                except Exception:
                    pass

def is_xlsx(path: str) -> bool:
    try:
        return os.path.isfile(path) and path.lower().endswith(".xlsx")
    except Exception:
        return False


def thumb_cache_path(images_base: str, resolved_path: str, target_edge: int) -> str:
    rel = os.path.relpath(resolved_path, images_base)
    key = hashlib.md5(f"{rel}|{target_edge}".encode("utf-8")).hexdigest()
    cache_dir = os.path.join(images_base, ".thumb_cache")
    os.makedirs(cache_dir, exist_ok=True)
    return os.path.join(cache_dir, f"{key}.png")


def build_thumb_if_needed(images_base: str, resolved_path: str, target_edge: int) -> str:
    thumb = thumb_cache_path(images_base, resolved_path, target_edge)
    try:
        src_mtime = os.path.getmtime(resolved_path)
        if os.path.exists(thumb) and os.path.getmtime(thumb) >= src_mtime:
            return thumb
        # Use Qt to scale and save
        img = QtGui.QImage(resolved_path)
        if img.isNull():
            return resolved_path
        w, h = img.width(), img.height()
        scale = target_edge / float(max(w, h))
        if scale < 1.0:
            img = img.scaled(int(w * scale), int(h * scale), QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        img.save(thumb, "PNG")
        return thumb
    except Exception:
        return resolved_path


class LabelerWindow(QtWidgets.QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("PySide6 Local Labeler")
        self.resize(1400, 900)

        # State
        self.images_base: str = ""  # inference/viz base
        self.images_base_orig: str = ""  # original images base with same sub-structure
        self.images_base_extra: str = ""  # optional extra images base
        self.excel_path: str = ""
        self.output_excel_path: str = ""
        self.json_path: str = ""
        self.df: Optional[pd.DataFrame] = None
        self.sheet_name: str = "inference_results"
        self.col_indices: Dict[str, int] = {}
        
        # Memory management and performance settings
        system_memory = get_system_memory()
        # Adaptive memory limits based on system memory
        self.max_memory_mb = min(1024, system_memory * 0.25)  # 25% of system memory, max 1GB
        self.chunk_size = 500  # Reduced from 1000 for smaller chunks
        self.max_table_rows = 2000  # Reduced from 5000 for better performance
        self.image_cache_size = 5  # Reduced from 10 for less memory usage
        self._image_cache: Dict[str, QtGui.QPixmap] = {}
        self._lazy_loading = True  # Enable lazy loading for large datasets
        
        # Default labeling columns split by mode (INF/EXT)
        self.label_map: Dict[str, List[str]] = {
            "review_label_inf": [
                "OK",
                "NG",
                "NG_BUT",
                "보류",
                # Custom relabel choices requested
                "SR-이물->OK",
                "SR-이물->도금-찍힘",
            ],
            "review_label_ext": ["OK_V3","OK_V4", "NG", "NG_BUT", "보류","매칭","매칭_안됨"],
        }
        self.active_label_col: str = "review_label_inf"
        self.current_idx: int = 0
        self.filtered_indices: List[int] = []
        self.fit_to_window: bool = True
        # Dynamic TO-BE choices extracted from CSV predictions
        self.tobe_choices: List[str] = []
        # Persist settings
        self.settings = QtCore.QSettings("rtm", "pyside_labeler")
        # Internal navigation guard
        self._navigating: bool = False
        # Batched JSON save
        self._pending_ops: List[Tuple[str, int, Dict[str, object], Dict[str, str]]] = []
        self._pending_json_path: str = ""
        self._save_timer = QtCore.QTimer(self)
        self._save_timer.setSingleShot(True)
        self._save_timer.setInterval(150)  # ms
        self._save_timer.timeout.connect(self._flush_pending_ops)

        # UI
        self._build_ui()
        self._connect_shortcuts()
        # Try restore last session
        try:
            self.restore_last_session()
        except Exception:
            pass

    def _build_ui(self) -> None:
        self.status = self.statusBar()
        # Permanent status widgets (INF/ORG indicator)
        self.lbl_status_io = QtWidgets.QLabel("")
        self.lbl_status_io.setStyleSheet("font-weight:600; padding-left:8px;")
        self.status.addPermanentWidget(self.lbl_status_io)

        # Menus
        file_menu = self.menuBar().addMenu("File")
        act_open = file_menu.addAction("Open Excel/CSV…")
        act_set_images = file_menu.addAction("Set Images Base…")
        act_export = file_menu.addAction("Apply JSON → Excel…")
        act_set_images_orig = file_menu.addAction("Set Original Images Base…")
        act_set_images_extra = file_menu.addAction("Set Extra Images Base…")
        act_quit = file_menu.addAction("Quit")
        act_quit.triggered.connect(self.close)
        act_open.triggered.connect(self.on_open_excel)
        act_set_images.triggered.connect(self.on_set_images_base)
        act_export.triggered.connect(self.on_apply_json)
        act_set_images_orig.triggered.connect(self.on_set_images_base_orig)
        act_set_images_extra.triggered.connect(self.on_set_images_base_extra)

        config_menu = self.menuBar().addMenu("Config")
        act_labels = config_menu.addAction("Configure Labels…")
        act_labels.triggered.connect(self.on_configure_labels)

        tools_menu = self.menuBar().addMenu("Tools")
        act_test = tools_menu.addAction("Matching Test…")
        act_test.triggered.connect(self.on_matching_test)
        
        # Memory management menu
        memory_menu = self.menuBar().addMenu("Memory")
        act_clear_cache = memory_menu.addAction("Clear Image Cache")
        act_clear_cache.triggered.connect(self._clear_image_cache)
        act_memory_info = memory_menu.addAction("Memory Info")
        act_memory_info.triggered.connect(self._show_memory_info)
        act_load_more = memory_menu.addAction("Load More Data")
        act_load_more.triggered.connect(self._load_more_data)
        act_force_cleanup = memory_menu.addAction("Force Memory Cleanup")
        act_force_cleanup.triggered.connect(self._force_memory_cleanup)
        act_optimize_settings = memory_menu.addAction("Optimize Memory Settings")
        act_optimize_settings.triggered.connect(self._optimize_memory_settings)

        # Central splitter
        splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        self.setCentralWidget(splitter)

        # Left: three image previews (inference/viz | original | extra) side-by-side with a status banner
        images_split = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        # Inference/Viz panel
        self.scroll_infer = QtWidgets.QScrollArea()
        self.scroll_infer.setWidgetResizable(True)
        self.image_label_infer = QtWidgets.QLabel(alignment=QtCore.Qt.AlignCenter)
        self.image_label_infer.setScaledContents(False)
        self.image_label_infer.setBackgroundRole(QtGui.QPalette.Base)
        self.scroll_infer.setWidget(self.image_label_infer)
        self.path_label_infer = QtWidgets.QLabel("")
        self.path_label_infer.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse)
        self.path_label_infer.setWordWrap(True)
        self.path_label_infer.setStyleSheet("color:#666; font-size:11px;")
        infer_panel = QtWidgets.QWidget()
        infer_layout = QtWidgets.QVBoxLayout(infer_panel)
        infer_layout.setContentsMargins(0, 0, 0, 0)
        infer_layout.setSpacing(2)
        infer_layout.addWidget(self.scroll_infer)
        infer_layout.addWidget(self.path_label_infer)
        # Original panel
        self.scroll_orig = QtWidgets.QScrollArea()
        self.scroll_orig.setWidgetResizable(True)
        self.image_label_orig = QtWidgets.QLabel(alignment=QtCore.Qt.AlignCenter)
        self.image_label_orig.setScaledContents(False)
        self.image_label_orig.setBackgroundRole(QtGui.QPalette.Base)
        self.scroll_orig.setWidget(self.image_label_orig)
        self.path_label_orig = QtWidgets.QLabel("")
        self.path_label_orig.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse)
        self.path_label_orig.setWordWrap(True)
        self.path_label_orig.setStyleSheet("color:#666; font-size:11px;")
        orig_panel = QtWidgets.QWidget()
        orig_layout = QtWidgets.QVBoxLayout(orig_panel)
        orig_layout.setContentsMargins(0, 0, 0, 0)
        orig_layout.setSpacing(2)
        orig_layout.addWidget(self.scroll_orig)
        orig_layout.addWidget(self.path_label_orig)
        # Extra panel (optional)
        self.scroll_extra = QtWidgets.QScrollArea()
        self.scroll_extra.setWidgetResizable(True)
        self.image_label_extra = QtWidgets.QLabel(alignment=QtCore.Qt.AlignCenter)
        self.image_label_extra.setScaledContents(False)
        self.image_label_extra.setBackgroundRole(QtGui.QPalette.Base)
        self.scroll_extra.setWidget(self.image_label_extra)
        self.path_label_extra = QtWidgets.QLabel("")
        self.path_label_extra.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse)
        self.path_label_extra.setWordWrap(True)
        self.path_label_extra.setStyleSheet("color:#666; font-size:11px;")
        extra_panel = QtWidgets.QWidget()
        extra_layout = QtWidgets.QVBoxLayout(extra_panel)
        extra_layout.setContentsMargins(0, 0, 0, 0)
        extra_layout.setSpacing(2)
        extra_layout.addWidget(self.scroll_extra)
        extra_layout.addWidget(self.path_label_extra)
        # Assemble side-by-side
        images_split.addWidget(infer_panel)
        images_split.addWidget(orig_panel)
        images_split.addWidget(extra_panel)
        images_split.splitterMoved.connect(lambda *_: self.refresh_view())

        # Right: controls
        right = QtWidgets.QWidget()
        right_layout = QtWidgets.QVBoxLayout(right)

        # Mode tabs (INF / EXT) → controls which label column is active by default
        self.tab_mode = QtWidgets.QTabWidget()
        self.tab_mode.addTab(QtWidgets.QWidget(), "INF")
        self.tab_mode.addTab(QtWidgets.QWidget(), "EXT")
        self.tab_mode.currentChanged.connect(self.on_change_mode_tab)
        right_layout.addWidget(self.tab_mode)

        # Active label column
        self.cmb_label_col = QtWidgets.QComboBox()
        self.cmb_label_col.currentTextChanged.connect(self.on_change_label_col)
        right_layout.addWidget(QtWidgets.QLabel("Active Label Column"))
        right_layout.addWidget(self.cmb_label_col)

        # Buttons for choices
        self.choice_buttons_container = QtWidgets.QWidget()
        self.choice_buttons_layout = QtWidgets.QGridLayout(self.choice_buttons_container)
        right_layout.addWidget(self.choice_buttons_container)

        # Dropdown for current row value
        right_layout.addWidget(QtWidgets.QLabel("Set label (dropdown)"))
        self.cmb_choice = QtWidgets.QComboBox()
        self.cmb_choice.currentTextChanged.connect(self.on_select_choice)
        right_layout.addWidget(self.cmb_choice)
        # Bulk from predictions → review_label_inf
        self.btn_from_preds_inf = QtWidgets.QPushButton("From preds → review_label_inf")
        self.btn_from_preds_inf.clicked.connect(self.on_bulk_label_from_preds_inf)
        right_layout.addWidget(self.btn_from_preds_inf)

        # AS-IS / TO-BE mapping panel
        self.grp_as_is_tobe = QtWidgets.QGroupBox("AS-IS / TO-BE")
        self.as_is_tobe_layout = QtWidgets.QGridLayout(self.grp_as_is_tobe)
        self.btn_apply_tobe = QtWidgets.QPushButton("Apply TO-BE → review_label_inf")
        self.btn_apply_tobe.clicked.connect(self.on_apply_tobe_to_review_inf)
        self.as_is_tobe_layout.addWidget(self.btn_apply_tobe, 0, 0, 1, 2)
        right_layout.addWidget(self.grp_as_is_tobe)

        # Add new label column UI
        grp_add = QtWidgets.QGroupBox("Add Label Column")
        add_layout = QtWidgets.QFormLayout(grp_add)
        self.edt_new_col = QtWidgets.QLineEdit()
        self.edt_new_opts = QtWidgets.QLineEdit()
        self.btn_add_col = QtWidgets.QPushButton("Add")
        self.btn_add_col.clicked.connect(self.on_add_column)
        add_layout.addRow("Column name", self.edt_new_col)
        add_layout.addRow("Options (comma)", self.edt_new_opts)
        add_layout.addRow(self.btn_add_col)
        right_layout.addWidget(grp_add)

        # Filter / Sort
        grp_filter = QtWidgets.QGroupBox("Filter / Sort")
        fl = QtWidgets.QGridLayout(grp_filter)
        self.cmb_origin = QtWidgets.QComboBox()
        self.edt_text = QtWidgets.QLineEdit()
        self.chk_unlabeled = QtWidgets.QCheckBox("Only unlabeled (active col)")
        self.cmb_label_state = QtWidgets.QComboBox()
        self.cmb_label_state.addItems(["All", "Labeled", "Unlabeled"])  # stronger label filter
        # Value filter (specific option in active label column)
        self.cmb_label_value = QtWidgets.QComboBox()
        self.cmb_sort_col = QtWidgets.QComboBox()
        self.chk_sort_desc = QtWidgets.QCheckBox("Desc")
        self.btn_clear_sort = QtWidgets.QPushButton("Clear sort")
        self.btn_clear_sort.clicked.connect(self.on_clear_sort)
        self.chk_bookmarks = QtWidgets.QCheckBox("Only bookmarks")
        # pred_seg_results filters
        self.grp_pred = QtWidgets.QGroupBox("pred_seg_results contains")
        gl_pred = QtWidgets.QGridLayout(self.grp_pred)
        self.chk_pred_exclusive = QtWidgets.QCheckBox("Exclusive (only selected)")
        self.chk_pred_exclude = QtWidgets.QCheckBox("Exclude selected")
        self.pred_checks_container = QtWidgets.QWidget()
        self.pred_checks_layout = QtWidgets.QGridLayout(self.pred_checks_container)
        gl_pred.addWidget(self.chk_pred_exclusive, 0, 0)
        gl_pred.addWidget(self.chk_pred_exclude, 0, 1)
        gl_pred.addWidget(self.pred_checks_container, 1, 0, 1, 2)
        self.btn_apply_filter = QtWidgets.QPushButton("Apply")
        self.btn_reset_filter = QtWidgets.QPushButton("Reset")
        self.btn_apply_filter.clicked.connect(self.apply_filters)
        self.btn_reset_filter.clicked.connect(self.reset_filters)
        fl.addWidget(QtWidgets.QLabel("origin_class"), 0, 0)
        fl.addWidget(self.cmb_origin, 0, 1)
        fl.addWidget(QtWidgets.QLabel("Text contains"), 1, 0)
        fl.addWidget(self.edt_text, 1, 1)
        fl.addWidget(QtWidgets.QLabel("Label state"), 2, 0)
        fl.addWidget(self.cmb_label_state, 2, 1)
        fl.addWidget(QtWidgets.QLabel("Value (active col)"), 3, 0)
        fl.addWidget(self.cmb_label_value, 3, 1)
        fl.addWidget(self.chk_unlabeled, 4, 0, 1, 2)
        fl.addWidget(self.chk_bookmarks, 4, 2)
        fl.addWidget(QtWidgets.QLabel("Sort by"), 5, 0)
        fl.addWidget(self.cmb_sort_col, 5, 1)
        fl.addWidget(self.chk_sort_desc, 5, 2)
        fl.addWidget(self.btn_clear_sort, 5, 3)
        # pred filters row
        fl.addWidget(self.grp_pred, 6, 0, 1, 4)
        fl.addWidget(self.btn_apply_filter, 7, 1)
        fl.addWidget(self.btn_reset_filter, 7, 2)
        right_layout.addWidget(grp_filter)

        # Preview table of filtered items (sortable columns)
        self.table_preview = QtWidgets.QTableWidget()
        self.table_preview.setColumnCount(5)
        self.table_preview.setHorizontalHeaderLabels(["idx", "label", "path", "INF", "EXT"]) 
        self.table_preview.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table_preview.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table_preview.setSortingEnabled(True)
        self.table_preview.itemSelectionChanged.connect(self.on_table_select)
        self.table_preview.horizontalHeader().setStretchLastSection(True)
        right_layout.addWidget(self.table_preview)

        # View options
        self.chk_fit = QtWidgets.QCheckBox("Fit to window")
        self.chk_fit.setChecked(True)
        self.chk_fit.toggled.connect(lambda *_: self.on_fit_toggle())
        right_layout.addWidget(self.chk_fit)

        # Live stats
        self.lbl_stats = QtWidgets.QLabel("")
        right_layout.addWidget(self.lbl_stats)

        # Summary (overall dataset)
        grp_sum = QtWidgets.QGroupBox("Summary")
        sum_layout = QtWidgets.QVBoxLayout(grp_sum)
        self.txt_summary = QtWidgets.QPlainTextEdit(readOnly=True)
        self.txt_summary.setMinimumHeight(140)
        sum_layout.addWidget(self.txt_summary)
        right_layout.addWidget(grp_sum)

        # Log bar at bottom
        grp_log = QtWidgets.QGroupBox("Log")
        log_layout = QtWidgets.QVBoxLayout(grp_log)
        self.txt_log = QtWidgets.QPlainTextEdit(readOnly=True)
        self.txt_log.setMinimumHeight(80)
        log_layout.addWidget(self.txt_log)
        right_layout.addWidget(grp_log)

        # Bookmark + Memo
        grp_bm = QtWidgets.QGroupBox("Bookmark / Memo")
        bm_layout = QtWidgets.QGridLayout(grp_bm)
        self.btn_toggle_bm = QtWidgets.QPushButton("★ Toggle Bookmark")
        self.btn_toggle_bm.clicked.connect(self.on_toggle_bookmark)
        self.edt_memo = QtWidgets.QPlainTextEdit()
        self.edt_memo.setPlaceholderText("Enter memo for this row…")
        self.btn_save_memo = QtWidgets.QPushButton("Save Memo")
        self.btn_save_memo.clicked.connect(self.on_save_memo)
        bm_layout.addWidget(self.btn_toggle_bm, 0, 0)
        bm_layout.addWidget(self.btn_save_memo, 0, 1)
        bm_layout.addWidget(self.edt_memo, 1, 0, 1, 2)
        right_layout.addWidget(grp_bm)

        # Navigation
        nav_layout = QtWidgets.QHBoxLayout()
        self.btn_prev = QtWidgets.QPushButton("◀ Prev")
        self.btn_next = QtWidgets.QPushButton("Next ▶")
        self.btn_prev.clicked.connect(self.on_prev)
        self.btn_next.clicked.connect(self.on_next)
        nav_layout.addWidget(self.btn_prev)
        nav_layout.addWidget(self.btn_next)
        right_layout.addLayout(nav_layout)

        # Info
        self.lbl_info = QtWidgets.QLabel("")
        self.lbl_info.setWordWrap(True)
        right_layout.addWidget(self.lbl_info)
        right_layout.addStretch()

        # Left container with banner + images
        left_container = QtWidgets.QWidget()
        left_v = QtWidgets.QVBoxLayout(left_container)
        self.lbl_banner = QtWidgets.QLabel("Status: -")
        self.lbl_banner.setAlignment(QtCore.Qt.AlignCenter)
        self.lbl_banner.setMinimumHeight(28)
        self.lbl_banner.setMaximumHeight(36)
        self.lbl_banner.setStyleSheet(
            "font-size: 24px; font-weight: 800; padding: 4px; border-radius: 4px;"
        )
        left_v.addWidget(self.lbl_banner)
        left_v.addWidget(images_split)

        # Wrap right panel with scroll area to avoid overflow
        right_scroll = QtWidgets.QScrollArea()
        right_scroll.setWidgetResizable(True)
        right_scroll.setWidget(right)
        splitter.addWidget(left_container)
        splitter.addWidget(right_scroll)
        splitter.setSizes([1200, 400])

        # Update on viewport resize for responsive fit
        self.scroll_infer.viewport().installEventFilter(self)
        self.scroll_orig.viewport().installEventFilter(self)
        self.scroll_extra.viewport().installEventFilter(self)

    def _connect_shortcuts(self) -> None:
        QtGui.QShortcut(QtGui.QKeySequence(QtCore.Qt.Key_Left), self, activated=self.on_prev)
        QtGui.QShortcut(QtGui.QKeySequence(QtCore.Qt.Key_Right), self, activated=self.on_next)
        # Number keys to assign labels
        for i in range(1, 10):
            QtGui.QShortcut(QtGui.QKeySequence(str(i)), self, activated=lambda i=i: self.on_assign_index(i - 1))

    def eventFilter(self, obj: QtCore.QObject, event: QtCore.QEvent) -> bool:
        if event.type() == QtCore.QEvent.Resize and getattr(self, 'fit_to_window', True):
            self.refresh_view()
        return super().eventFilter(obj, event)

    def on_fit_toggle(self) -> None:
        self.fit_to_window = self.chk_fit.isChecked()
        self.refresh_view()

    # Data loading / configuration
    def compute_tobe_choices(self) -> None:
        self.tobe_choices = []
        try:
            if self.df is None or self.df.empty or "pred_seg_results" not in self.df.columns:
                return
            uniq: List[str] = []
            seen = set()
            for v in self.df["pred_seg_results"].fillna(""):
                for item in parse_pred_list(v):
                    s = str(item).strip()
                    if not s:
                        continue
                    if s not in seen:
                        seen.add(s)
                        uniq.append(s)
            # Keep order of first appearance; exclude 'OK' here (we'll pin it at front in UI)
            self.tobe_choices = [c for c in uniq if c != "OK"]
        except Exception:
            self.tobe_choices = []
    def load_excel_from_path(self, path: str) -> None:
        if not path:
            return
        try:
            # Clear existing data and force garbage collection
            self.df = None
            self._image_cache.clear()
            force_garbage_collection()
            
            # Check file size first
            file_size_mb = os.path.getsize(path) / (1024 * 1024)
            self.log(f"File size: {file_size_mb:.1f}MB")
            
            if path.lower().endswith(".csv"):
                # For large CSV files, read in chunks
                if file_size_mb > 50:  # Reduced threshold from 100MB to 50MB
                    self.log(f"Large file detected ({file_size_mb:.1f}MB), loading in chunks...")
                    # Read first chunk to get column info
                    chunk_df = pd.read_csv(path, encoding="utf-8-sig", nrows=self.chunk_size)
                    self.df = chunk_df
                    self.log(f"Loaded first {self.chunk_size} rows. Use 'Load More' to load additional data.")
                else:
                    # Check memory before loading
                    if check_memory_limit(self.max_memory_mb):
                        QtWidgets.QMessageBox.warning(self, "Memory Warning", 
                            f"Memory usage is high ({get_memory_usage():.1f}MB). Loading in chunks.")
                        chunk_df = pd.read_csv(path, encoding="utf-8-sig", nrows=self.chunk_size)
                        self.df = chunk_df
                    else:
                        self.df = pd.read_csv(path, encoding="utf-8-sig")
                self.sheet_name = "inference_results"
                # Create a working xlsx path next to csv
                xlsx_path = os.path.splitext(path)[0] + ".xlsx"
                with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
                    self.df.to_excel(writer, index=False, sheet_name=self.sheet_name)
                self.excel_path = xlsx_path
            else:
                # For Excel files, check memory usage
                if check_memory_limit(self.max_memory_mb):
                    QtWidgets.QMessageBox.warning(self, "Memory Warning", 
                        f"Memory usage is high ({get_memory_usage():.1f}MB). Consider closing other applications.")
                
                # Read first sheet name
                xl = pd.ExcelFile(path)
                self.sheet_name = xl.sheet_names[0]
                
                # For large Excel files, read in chunks
                try:
                    # Try to get row count without loading everything
                    wb = load_workbook(path, read_only=True)
                    ws = wb[self.sheet_name]
                    row_count = ws.max_row
                    wb.close()
                    
                    if row_count > 5000:  # Reduced from 10000 to 5000
                        self.log(f"Large Excel file detected ({row_count} rows), loading first {self.chunk_size} rows...")
                        self.df = pd.read_excel(path, sheet_name=self.sheet_name, nrows=self.chunk_size)
                        self.log(f"Loaded first {self.chunk_size} rows. Use 'Load More' to load additional data.")
                    else:
                        # Check memory before loading
                        if check_memory_limit(self.max_memory_mb):
                            self.log("Memory usage high, loading in chunks...")
                            self.df = pd.read_excel(path, sheet_name=self.sheet_name, nrows=self.chunk_size)
                        else:
                            self.df = xl.parse(self.sheet_name)
                except Exception:
                    # Fallback to normal loading
                    self.df = xl.parse(self.sheet_name)
                
                self.excel_path = path
            
            # Check memory usage after loading
            memory_usage = get_memory_usage()
            self.log(f"Memory usage after loading: {memory_usage:.1f}MB")
            
            if memory_usage > self.max_memory_mb * 0.8:  # 80% of limit
                QtWidgets.QMessageBox.warning(self, "Memory Warning", 
                    f"High memory usage detected ({memory_usage:.1f}MB). Consider reducing data size.")
            
            # Defaults
            self.output_excel_path = os.path.splitext(self.excel_path)[0] + "_labeled.xlsx"
            self.json_path = default_json_path(self.output_excel_path)
            # Ensure label columns
            if self.df is not None:
                for col in self.label_map.keys():
                    if col not in self.df.columns:
                        self.df[col] = ""
                    ensure_object_dtype(self.df, col)
                # Merge previous JSON labels into DataFrame (resume work)
                try:
                    merge_json_into_df(self.json_path, self.df, list(self.label_map.keys()))
                except Exception:
                    pass
                # Build dynamic TO-BE choices from CSV predictions
                try:
                    self.compute_tobe_choices()
                except Exception:
                    pass
            self.filtered_indices = list(self.df.index) if self.df is not None else []
            self.current_idx = 0
            # Build label controls and filter controls
            self.refresh_label_controls()
            self.populate_filter_controls()
            # Default filter: origin_class=(all), label state=Unlabeled, sort by img_path if exists
            try:
                idx = self.cmb_label_state.findText("Unlabeled")
                if idx >= 0:
                    self.cmb_label_state.setCurrentIndex(idx)
            except Exception:
                pass
            # Set default sort column
            if self.cmb_sort_col.count() > 0:
                pref = "img_path" if (self.df is not None and "img_path" in self.df.columns) else ("filename" if (self.df is not None and "filename" in self.df.columns) else None)
                if pref is not None:
                    i2 = self.cmb_sort_col.findText(pref)
                    if i2 >= 0:
                        self.cmb_sort_col.setCurrentIndex(i2)
            # Apply filters to drive the list and navigation
            self.apply_filters()
            self.refresh_view()
            self.status.showMessage(f"Loaded: {self.excel_path}")
            self.log(f"Loaded file: {self.excel_path}")
            # persist
            self.settings.setValue("excel_path", path)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Open failed", str(e))
            self.log(f"Error loading file: {str(e)}")

    def on_open_excel(self) -> None:
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Open Excel/CSV", os.getcwd(), "Excel/CSV (*.xlsx *.csv)")
        if not path:
            return
        self.load_excel_from_path(path)

    def on_set_images_base(self) -> None:
        path = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Images Base", os.getcwd())
        if path:
            self.images_base = path
            self.refresh_view()
            self.log(f"Set Images Base: {path}")
            self.settings.setValue("images_base", path)

    def on_set_images_base_orig(self) -> None:
        path = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Original Images Base", os.getcwd())
        if path:
            self.images_base_orig = path
            self.refresh_view()
            self.log(f"Set Original Images Base: {path}")
            self.settings.setValue("images_base_orig", path)

    def on_set_images_base_extra(self) -> None:
        path = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Extra Images Base", os.getcwd())
        if path:
            self.images_base_extra = path
            self.refresh_view()
            self.log(f"Set Extra Images Base: {path}")
            self.settings.setValue("images_base_extra", path)

    def restore_last_session(self) -> None:
        excel = self.settings.value("excel_path", "", str)
        img_base = self.settings.value("images_base", "", str)
        img_base_orig = self.settings.value("images_base_orig", "", str)
        img_base_extra = self.settings.value("images_base_extra", "", str)
        if excel and os.path.exists(excel):
            # Reuse same loading routine
            try:
                self.load_excel_from_path(excel)
            except Exception:
                pass
        if img_base and os.path.isdir(img_base):
            self.images_base = img_base
        if img_base_orig and os.path.isdir(img_base_orig):
            self.images_base_orig = img_base_orig
        if img_base or img_base_orig:
            self.refresh_view()
        if img_base_extra and os.path.isdir(img_base_extra):
            self.images_base_extra = img_base_extra
            self.refresh_view()

    def on_configure_labels(self) -> None:
        text, ok = QtWidgets.QInputDialog.getMultiLineText(
            self,
            "Configure Labels",
            "One line per column: name: opt1, opt2, ...",
            "review_label: OK, NG, 보류\npriority: High, Medium, Low",
        )
        if not ok:
            return
        mapping: Dict[str, List[str]] = {}
        for line in text.splitlines():
            if ":" in line:
                name, vals = line.split(":", 1)
                name = name.strip()
                opts = [v.strip() for v in vals.split(",") if v.strip()]
                if name:
                    mapping[name] = opts
        if mapping:
            self.label_map = mapping
            self.active_label_col = list(self.label_map.keys())[0]
            if self.df is not None:
                for col in self.label_map.keys():
                    if col not in self.df.columns:
                        self.df[col] = ""
                    ensure_object_dtype(self.df, col)
            self.refresh_label_controls()
            self.refresh_view()

    def on_add_column(self) -> None:
        name = self.edt_new_col.text().strip()
        opts = [v.strip() for v in self.edt_new_opts.text().split(',') if v.strip()]
        if not name or not opts:
            QtWidgets.QMessageBox.information(self, "Add Column", "Enter column name and at least one option.")
            return
        self.label_map[name] = opts
        self.active_label_col = name
        if self.df is not None:
            if name not in self.df.columns:
                self.df[name] = ""
            ensure_object_dtype(self.df, name)
        self.refresh_label_controls()
        self.refresh_view()
        self.status.showMessage(f"Added label column: {name}")

    def on_matching_test(self) -> None:
        if self.df is None or self.df.empty:
            QtWidgets.QMessageBox.information(self, "Matching Test", "Open Excel/CSV first.")
            return
        # Determine test target by current mode tab (INF vs EXT)
        mode_idx = 0
        try:
            mode_idx = self.tab_mode.currentIndex() if hasattr(self, 'tab_mode') else 0
        except Exception:
            mode_idx = 0
        testing_ext = (mode_idx == 1)
        base_dir = self.images_base_extra if testing_ext else self.images_base
        if not base_dir:
            QtWidgets.QMessageBox.information(self, "Matching Test", "Set {} Images Base first.".format("Extra" if testing_ext else "Images"))
            return
        total_available = len(self.filtered_indices) if self.filtered_indices else len(self.df)
        default_n = min(200, total_available) if total_available > 0 else 0
        n, ok = QtWidgets.QInputDialog.getInt(self, "Matching Test", "Sample size", default_n, 1, max(1, total_available), 1)
        if not ok:
            return
        if n <= 0 or total_available == 0:
            QtWidgets.QMessageBox.information(self, "Matching Test", "No rows to test.")
            return
        indices = self.filtered_indices if self.filtered_indices else list(self.df.index)
        sample = indices[:n]
        ok_count = 0
        misses: List[str] = []
        for ridx in sample:
            r = self.df.loc[ridx]
            p = str(r.get("img_path", "")) or str(r.get("filename", ""))
            if testing_ext:
                # Use same strategy as extra/original resolution: join rel; else search by filename patterns
                rel = normalize_relative_path(p)
                rp = os.path.join(base_dir, rel)
                if not os.path.exists(rp):
                    basename = os.path.basename(rel)
                    base_no_ext, _ = os.path.splitext(basename)
                    patterns = [
                        os.path.join(base_dir, "**", basename),
                        os.path.join(base_dir, "**", f"{base_no_ext}.*"),
                        os.path.join(base_dir, "**", f"*{base_no_ext}*.*"),
                    ]
                    rp = None
                    for pat in patterns:
                        m = glob.glob(pat, recursive=True)
                        if m:
                            rp = m[0]
                            break
            else:
                rp = resolve_image_path(base_dir, p)
            if rp and os.path.exists(rp):
                ok_count += 1
            else:
                if len(misses) < 10:
                    misses.append(p)
        rate = (ok_count / float(len(sample))) * 100.0
        mode_label = "EXT" if testing_ext else "INF"
        msg = f"[{mode_label}] Matched {ok_count}/{len(sample)} ({rate:.1f}%)\nBase: {base_dir}"
        if misses:
            msg += "\n\nExamples not found:" + "\n- " + "\n- ".join(misses)
        QtWidgets.QMessageBox.information(self, "Matching Test", msg)

    # Export JSON → Excel
    def on_apply_json(self) -> None:
        if not (self.excel_path and is_xlsx(self.excel_path)):
            QtWidgets.QMessageBox.warning(self, "Export", "Open a valid Excel/CSV first.")
            return
        out, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save labeled Excel", self.output_excel_path, "Excel (*.xlsx)")
        if not out:
            return
        try:
            # Ensure pending JSON updates are flushed before exporting
            self._flush_pending_ops()
            # Ensure workbook has all label columns and get indices
            wb = load_workbook(self.excel_path)
            ws = wb[self.sheet_name]
            headers = [c.value for c in ws[1]]
            if headers is None:
                headers = []
            col_indices: Dict[str, int] = {}
            for col in self.label_map.keys():
                if col in headers:
                    col_indices[col] = headers.index(col) + 1
                else:
                    headers.append(col)
                    idx = len(headers)
                    ws.cell(row=1, column=idx, value=col)
                    col_indices[col] = idx
            wb.save(out)
            wb.close()
            # Apply JSON into the new file
            applied = apply_json_to_excel(self.json_path or default_json_path(out), out, self.sheet_name, col_indices, self.df)
            self.output_excel_path = out
            self.status.showMessage(f"Applied {applied} cells → {out}")
            self.log(f"Applied JSON to Excel: {applied} cells → {out}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export failed", str(e))

    # Navigation / labeling
    def on_prev(self) -> None:
        if self._navigating:
            return
        if not self.filtered_indices:
            return
        self.current_idx = max(0, self.current_idx - 1)
        self.refresh_view()

    def on_next(self) -> None:
        if self._navigating:
            return
        if not self.filtered_indices:
            return
        self.current_idx = min(len(self.filtered_indices) - 1, self.current_idx + 1)
        self.refresh_view()

    def on_assign_index(self, choice_index: int) -> None:
        if not (self.df is not None and self.filtered_indices):
            return
        opts = self.label_map.get(self.active_label_col, [])
        if not (0 <= choice_index < len(opts)):
            return
        value = opts[choice_index]
        row_idx = self.filtered_indices[self.current_idx]
        # Reflect in DataFrame immediately for UI updates
        try:
            self.df.at[row_idx, self.active_label_col] = value
        except Exception:
            pass
        # Queue JSON save (batched)
        row = self.df.loc[row_idx]
        keys_for_row = {"img_path": str(row.get("img_path", "")), "filename": str(row.get("filename", ""))}
        self._queue_set_values(row_idx, {self.active_label_col: value}, keys_for_row)
        self.status.showMessage(f"Queued save: {self.active_label_col}={value}")
        self.log(f"Label saved: row {row_idx} {self.active_label_col}={value}")
        # Keep working order stable; update list/stats without re-sorting
        self._after_label_saved(row_idx)

    def on_select_choice(self, text: str) -> None:
        if not text or text == "Select…":
            return
        # Map dropdown selection to save
        if self.df is None or not self.filtered_indices:
            return
        row_idx = self.filtered_indices[self.current_idx]
        # Reflect in DataFrame immediately
        try:
            self.df.at[row_idx, self.active_label_col] = text
        except Exception:
            pass
        # Queue JSON save (batched)
        row = self.df.loc[row_idx]
        keys_for_row = {"img_path": str(row.get("img_path", "")), "filename": str(row.get("filename", ""))}
        self._queue_set_values(row_idx, {self.active_label_col: text}, keys_for_row)
        self.status.showMessage(f"Queued save: {self.active_label_col}={text}")
        self.log(f"Label saved: row {row_idx} {self.active_label_col}={text}")
        # Keep working order stable; update list/stats without re-sorting
        self._after_label_saved(row_idx)

    def on_bulk_label_from_preds_inf(self) -> None:
        # Build labels for each predicted item and join into review_label_inf
        if self.df is None or not self.filtered_indices:
            QtWidgets.QMessageBox.information(self, "Bulk from preds", "Open Excel/CSV first.")
            return
        row_idx = self.filtered_indices[self.current_idx]
        row = self.df.loc[row_idx]
        preds_raw = str(row.get("pred_seg_results", ""))
        preds = parse_pred_list(preds_raw)
        if not preds:
            QtWidgets.QMessageBox.information(self, "Bulk from preds", "pred_seg_results 가 비어있습니다.")
            return
        choices = self.label_map.get("review_label_inf", [])
        if not choices:
            QtWidgets.QMessageBox.information(self, "Bulk from preds", "review_label_inf 선택지가 없습니다.")
            return

        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("From preds → review_label_inf")
        lay = QtWidgets.QVBoxLayout(dlg)
        form = QtWidgets.QFormLayout()
        combo_boxes: List[QtWidgets.QComboBox] = []
        for i, pred in enumerate(preds):
            cb = QtWidgets.QComboBox(dlg)
            cb.addItems(["(skip)"] + choices)
            # Try to preselect something matching the pred prefix
            pre_idx = 0
            for j, opt in enumerate(choices, start=1):
                if str(pred) and opt.startswith(str(pred)):
                    pre_idx = j
                    break
            cb.setCurrentIndex(pre_idx)
            combo_boxes.append(cb)
            form.addRow(QtWidgets.QLabel(f"{i+1}. {pred}"), cb)
        lay.addLayout(form)
        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        lay.addWidget(btns)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        if dlg.exec() != QtWidgets.QDialog.Accepted:
            return
        selected: List[str] = []
        for cb in combo_boxes:
            val = cb.currentText().strip()
            if val and val != "(skip)":
                selected.append(val)
        if not selected:
            QtWidgets.QMessageBox.information(self, "Bulk from preds", "선택된 라벨이 없습니다.")
            return
        final_text = ". ".join(selected)
        # Persist into DF and JSON
        try:
            self.df.at[row_idx, "review_label_inf"] = final_text
        except Exception:
            pass
        keys_for_row = {"img_path": str(row.get("img_path", "")), "filename": str(row.get("filename", ""))}
        self._queue_set_values(row_idx, {"review_label_inf": final_text}, keys_for_row)
        # Ensure active column and refresh
        self.active_label_col = "review_label_inf"
        self.refresh_label_controls()
        self.status.showMessage("Bulk labeled review_label_inf from preds")
        self.log(f"Bulk from preds → review_label_inf: {final_text}")
        self._after_label_saved(row_idx)

    def _clear_as_is_tobe_panel(self) -> None:
        try:
            # Remove all widgets except the apply button at (0, 0)-(0,1)
            for i in reversed(range(self.as_is_tobe_layout.count())):
                item = self.as_is_tobe_layout.itemAt(i)
                w = item.widget()
                # Keep the first row (apply button)
                if w is not None and w is not self.btn_apply_tobe:
                    self.as_is_tobe_layout.removeWidget(w)
                    w.setParent(None)
        except Exception:
            pass

    def _refresh_as_is_tobe_panel(self) -> None:
        # Build per-pred row: [AS-IS label]  [TO-BE dropdown]
        if not hasattr(self, 'grp_as_is_tobe'):
            return
        self._tobe_combos: List[QtWidgets.QComboBox] = []
        self._clear_as_is_tobe_panel()
        if self.df is None or not self.filtered_indices:
            return
        row_idx = self.filtered_indices[self.current_idx]
        row = self.df.loc[row_idx]
        preds = parse_pred_list(str(row.get("pred_seg_results", "")))
        # TO-BE dropdown: '(skip)' + OK + unique classes from CSV in order
        base_choices = self.tobe_choices if hasattr(self, 'tobe_choices') and self.tobe_choices else []
        choices = ["(skip)", "OK", *base_choices]
        # Try to preselect based on existing review_label_inf split by '. '
        existing = str(row.get("review_label_inf", ""))
        existing_items: List[str] = [x.strip() for x in existing.split(". ") if x.strip()]
        used = [False] * len(existing_items)
        base_row = 1
        for i, pred in enumerate(preds):
            lbl = QtWidgets.QLabel(pred)
            cb = QtWidgets.QComboBox()
            cb.addItems(choices)
            # Heuristic preselect: exact or prefix match to existing items
            pre_idx = 0
            for j, ex in enumerate(existing_items):
                if not used[j] and (ex == pred or ex.startswith(pred)):
                    try:
                        k = choices.index(ex) if ex in choices else 0
                    except Exception:
                        k = 0
                    if k > 0:
                        pre_idx = k
                        used[j] = True
                        break
            if pre_idx == 0:
                # Fallback: choose first option that startswith pred
                for k, opt in enumerate(choices):
                    if k > 0 and opt.startswith(str(pred)):
                        pre_idx = k
                        break
            cb.setCurrentIndex(pre_idx)
            self.as_is_tobe_layout.addWidget(QtWidgets.QLabel("AS-IS"), base_row + i, 0)
            self.as_is_tobe_layout.addWidget(lbl, base_row + i, 1)
            self.as_is_tobe_layout.addWidget(QtWidgets.QLabel("TO-BE"), base_row + i, 2)
            self.as_is_tobe_layout.addWidget(cb, base_row + i, 3)
            self._tobe_combos.append(cb)

    def on_apply_tobe_to_review_inf(self) -> None:
        if self.df is None or not self.filtered_indices:
            return
        row_idx = self.filtered_indices[self.current_idx]
        row = self.df.loc[row_idx]
        selected: List[str] = []
        try:
            for cb in getattr(self, '_tobe_combos', []):
                val = cb.currentText().strip()
                if val and val != "(skip)":
                    selected.append(val)
        except Exception:
            selected = []
        if not selected:
            QtWidgets.QMessageBox.information(self, "AS-IS / TO-BE", "선택된 TO-BE 라벨이 없습니다.")
            return
        final_text = ". ".join(selected)
        try:
            self.df.at[row_idx, "review_label_inf"] = final_text
        except Exception:
            pass
        keys_for_row = {"img_path": str(row.get("img_path", "")), "filename": str(row.get("filename", ""))}
        self._queue_set_values(row_idx, {"review_label_inf": final_text}, keys_for_row)
        self.status.showMessage("Applied TO-BE → review_label_inf")
        self.log(f"Apply TO-BE: {final_text}")
        self._after_label_saved(row_idx)

    def _find_list_row_by_index(self, idx: int) -> int:
        try:
            # lookup first column in table
            for i in range(self.table_preview.rowCount()):
                it = self.table_preview.item(i, 0)
                if it and it.text() == str(idx):
                    return i
        except Exception:
            pass
        return -1

    def _select_current_in_list(self) -> None:
        try:
            if self.filtered_indices and 0 <= self.current_idx < len(self.filtered_indices):
                self.table_preview.blockSignals(True)
                self.table_preview.clearSelection()
                self.table_preview.selectRow(self.current_idx)
                self.table_preview.blockSignals(False)
        except Exception:
            pass

    def _update_stats_quick(self) -> None:
        try:
            total = len(self.df) if self.df is not None else 0
            overall_unlabeled = 0
            overall_labeled = 0
            if self.df is not None and self.active_label_col in self.df.columns:
                overall_unlabeled = int(((self.df[self.active_label_col].isna()) | (self.df[self.active_label_col] == "")).sum())
                overall_labeled = total - overall_unlabeled
            f_total = len(self.filtered_indices)
            f_labeled = 0
            f_unlabeled = 0
            if f_total > 0 and self.active_label_col in self.df.columns:
                sub = self.df.loc[self.filtered_indices, self.active_label_col]
                f_unlabeled = int(((sub.isna()) | (sub == "")).sum())
                f_labeled = f_total - f_unlabeled
            self.lbl_stats.setText(
                f"Filtered: {f_total} | Labeled: {f_labeled} | Unlabeled: {f_unlabeled}  ||  Overall: {total} (L:{overall_labeled} U:{overall_unlabeled})"
            )
        except Exception:
            pass

    def _after_label_saved(self, row_idx: int) -> None:
        # Update current list item or remove it if filtered out
        state = self.cmb_label_state.currentText()
        unlabeled_only = (state == "Unlabeled") or self.chk_unlabeled.isChecked()
        label_val = str(self.df.loc[row_idx].get(self.active_label_col, "")) if (self.df is not None and self.active_label_col in self.df.columns) else ""
        removed = False
        self._navigating = True
        # Update list item text/icon
        li = self._find_list_row_by_index(row_idx)
        if 0 <= li:
            # Always keep the row visible; just update columns
            self.table_preview.setItem(li, 1, QtWidgets.QTableWidgetItem("1" if label_val else "0"))
            # Refresh INF/EXT columns from DF so active label goes to correct column
            try:
                inf_val = str(self.df.loc[row_idx].get("review_label_inf", "")) if "review_label_inf" in self.df.columns else ""
            except Exception:
                inf_val = ""
            try:
                ext_val = str(self.df.loc[row_idx].get("review_label_ext", "")) if "review_label_ext" in self.df.columns else ""
            except Exception:
                ext_val = ""
            self.table_preview.setItem(li, 3, QtWidgets.QTableWidgetItem(inf_val))
            self.table_preview.setItem(li, 4, QtWidgets.QTableWidgetItem(ext_val))
        # Force auto-advance to the immediate next row within current filtered order
        if self.current_idx < len(self.filtered_indices) - 1:
            self.current_idx += 1
        # Clamp if at end
        if self.current_idx >= len(self.filtered_indices):
            self.current_idx = max(0, len(self.filtered_indices) - 1)
        # Update stats/summary and view
        self._update_stats_quick()
        self.update_summary()
        self.refresh_view()
        self._select_current_in_list()
        self._navigating = False

    def on_change_label_col(self, name: str) -> None:
        if name:
            self.active_label_col = name
            self.refresh_label_controls()
            self.populate_filter_controls()
            self._populate_value_filter()
            # Reflect into mode tab if it's one of predefined modes
            try:
                if name == "review_label_inf":
                    self.tab_mode.blockSignals(True)
                    self.tab_mode.setCurrentIndex(0)
                    self.tab_mode.blockSignals(False)
                elif name == "review_label_ext":
                    self.tab_mode.blockSignals(True)
                    self.tab_mode.setCurrentIndex(1)
                    self.tab_mode.blockSignals(False)
            except Exception:
                pass

    def on_change_mode_tab(self, idx: int) -> None:
        # 0: INF → review_label_inf, 1: EXT → review_label_ext
        try:
            name = "review_label_inf" if idx == 0 else "review_label_ext"
            if name != self.active_label_col:
                self.active_label_col = name
                self.refresh_label_controls()
                self.populate_filter_controls()
                self._populate_value_filter()
                self._update_stats_quick()
                self.update_summary()
        except Exception:
            pass

    def refresh_label_controls(self) -> None:
        # Reset combobox
        self.cmb_label_col.blockSignals(True)
        self.cmb_label_col.clear()
        for name in self.label_map.keys():
            self.cmb_label_col.addItem(name)
        idx = list(self.label_map.keys()).index(self.active_label_col) if self.active_label_col in self.label_map else 0
        self.cmb_label_col.setCurrentIndex(idx)
        self.cmb_label_col.blockSignals(False)
        # Rebuild choice buttons (1..n shortcuts)
        while self.choice_buttons_layout.count():
            item = self.choice_buttons_layout.takeAt(0)
            w = item.widget()
            if w:
                w.setParent(None)
        opts = self.label_map.get(self.active_label_col, [])
        for i, opt in enumerate(opts):
            btn = QtWidgets.QPushButton(f"{i+1}. {opt}")
            btn.clicked.connect(lambda _, j=i: self.on_assign_index(j))
            self.choice_buttons_layout.addWidget(btn, i // 3, i % 3)
        # Update dropdown options
        self.cmb_choice.blockSignals(True)
        self.cmb_choice.clear()
        self.cmb_choice.addItem("Select…")
        for opt in opts:
            self.cmb_choice.addItem(opt)
        self.cmb_choice.setCurrentIndex(0)
        self.cmb_choice.blockSignals(False)
        # value filter options depend on active label column
        self._populate_value_filter()

    def populate_filter_controls(self) -> None:
        # origin_class dropdown
        self.cmb_origin.blockSignals(True)
        self.cmb_origin.clear()
        self.cmb_origin.addItem("(all)")
        if self.df is not None and "origin_class" in self.df.columns:
            try:
                # Use unique values from the full dataframe, not filtered
                vals = pd.Series(self.df["origin_class"].astype(str)).dropna().unique().tolist()
                for v in sorted([str(x) for x in vals]):
                    self.cmb_origin.addItem(v)
            except Exception:
                pass
        self.cmb_origin.setCurrentIndex(0)
        self.cmb_origin.blockSignals(False)
        # Sort columns
        self.cmb_sort_col.blockSignals(True)
        self.cmb_sort_col.clear()
        self.cmb_sort_col.addItem("(no sort)")
        if self.df is not None:
            for col in list(self.df.columns):
                self.cmb_sort_col.addItem(col)
        self.cmb_sort_col.blockSignals(False)
        # pred_seg_results unique values → checkboxes
        while self.pred_checks_layout.count():
            it = self.pred_checks_layout.takeAt(0)
            w = it.widget()
            if w:
                w.setParent(None)
        if self.df is not None and "pred_seg_results" in self.df.columns:
            try:
                uniques: List[str] = []
                for v in self.df["pred_seg_results"].fillna(""):
                    for item in parse_pred_list(v):
                        if item and item not in uniques:
                            uniques.append(item)
                uniques = sorted(uniques)
                self.pred_checkboxes: Dict[str, QtWidgets.QCheckBox] = {}
                for i, val in enumerate(uniques):
                    cb = QtWidgets.QCheckBox(val)
                    self.pred_checkboxes[val] = cb
                    self.pred_checks_layout.addWidget(cb, i // 3, i % 3)
            except Exception:
                pass

    def _populate_value_filter(self) -> None:
        try:
            self.cmb_label_value.blockSignals(True)
            self.cmb_label_value.clear()
            self.cmb_label_value.addItem("(all)")
            if self.df is not None and self.active_label_col in self.df.columns:
                vals = (
                    pd.Series(self.df[self.active_label_col].astype(str))
                    .replace("nan", "")
                    .replace("None", "")
                    .dropna()
                    .unique()
                    .tolist()
                )
                vals = sorted({v for v in vals if v})
                for v in vals:
                    self.cmb_label_value.addItem(str(v))
            self.cmb_label_value.blockSignals(False)
        except Exception:
            try:
                self.cmb_label_value.blockSignals(False)
            except Exception:
                pass

    def apply_filters(self) -> None:
        if self.df is None:
            return
        
        # Proactive memory cleanup before filtering
        self._proactive_memory_cleanup()
        
        # Check memory before filtering
        if self._manage_memory():
            QtWidgets.QMessageBox.warning(self, "Memory Warning", 
                "Memory usage is high. Some operations may be slower.")
        
        # Use copy only if necessary (for large datasets, avoid unnecessary copying)
        if len(self.df) > 5000:  # Reduced from 10000 to 5000
            # For large datasets, work with view instead of copy
            df = self.df
            use_view = True
        else:
            df = self.df.copy()
            use_view = False
        
        # origin_class filter
        origin_sel = self.cmb_origin.currentText()
        if origin_sel and origin_sel != "(all)" and "origin_class" in df.columns:
            if use_view:
                mask = df["origin_class"].astype(str) == origin_sel
                df = df[mask]
            else:
                df = df[df["origin_class"].astype(str) == origin_sel]
        
        # text contains across img_path and pred
        t = self.edt_text.text().strip()
        if t:
            t_low = t.lower()
            mask = pd.Series(False, index=df.index)
            for col in ("img_path", "filename", "pred_seg_results"):
                if col in df.columns:
                    mask = mask | df[col].astype(str).str.lower().str.contains(t_low, na=False)
            df = df[mask]
        
        # value filter for active label column
        val_sel = self.cmb_label_value.currentText() if hasattr(self, 'cmb_label_value') else "(all)"
        if val_sel and val_sel != "(all)" and self.active_label_col in df.columns:
            df = df[df[self.active_label_col].astype(str) == val_sel]
        
        # bookmark-only filter (JSON-backed)
        if hasattr(self, 'chk_bookmarks') and self.chk_bookmarks.isChecked():
            try:
                json_path = self.json_path or default_json_path(self.output_excel_path or self.excel_path)
                store = load_label_store(json_path)
                labels = store.get("labels", {})
                bookmarked_ids = set()
                for k, entry in labels.items():
                    try:
                        ridx = int(k)
                    except Exception:
                        continue
                    if bool(entry.get("bookmark", False)) and ridx in df.index:
                        bookmarked_ids.add(ridx)
                df = df[df.index.isin(bookmarked_ids)]
            except Exception:
                pass
        
        # label state filter
        state = self.cmb_label_state.currentText()
        if state == "Unlabeled" and self.active_label_col in df.columns:
            df = df[(df[self.active_label_col].isna()) | (df[self.active_label_col] == "")]
        elif state == "Labeled" and self.active_label_col in df.columns:
            df = df[(~df[self.active_label_col].isna()) & (df[self.active_label_col] != "")]
        
        # legacy checkbox support
        if self.chk_unlabeled.isChecked() and self.active_label_col in df.columns:
            df = df[(df[self.active_label_col].isna()) | (df[self.active_label_col] == "")]
        
        # pred_seg_results filter logic
        try:
            selected: List[str] = []
            if hasattr(self, 'pred_checkboxes'):
                for k, cb in self.pred_checkboxes.items():
                    if cb.isChecked():
                        selected.append(k)
            if selected:
                exclusive = self.chk_pred_exclusive.isChecked() if hasattr(self, 'chk_pred_exclusive') else False
                exclude = self.chk_pred_exclude.isChecked() if hasattr(self, 'chk_pred_exclude') else False
                keep_mask = []
                for ridx, v in df['pred_seg_results'].fillna("").items() if 'pred_seg_results' in df.columns else []:
                    items = set(parse_pred_list(v))
                    if exclude:
                        # drop rows that contain any selected items
                        keep = len(items.intersection(selected)) == 0
                    elif exclusive:
                        # keep only rows whose set equals selected
                        keep = items and items.issubset(set(selected)) and set(selected).issubset(items)
                    else:
                        # keep rows that contain at least one selected item
                        keep = len(items.intersection(selected)) > 0
                    keep_mask.append(keep)
                if 'pred_seg_results' in df.columns:
                    df = df[pd.Series(keep_mask, index=df.index)]
        except Exception:
            pass
        
        # sort
        sort_col = self.cmb_sort_col.currentText()
        if sort_col and sort_col != "(no sort)" and sort_col in df.columns:
            df = df.sort_values(by=sort_col, ascending=not self.chk_sort_desc.isChecked(), kind="mergesort")
        
        # update indices/preview list
        self.filtered_indices = list(df.index)
        self.current_idx = 0 if self.filtered_indices else 0
        
        # Preserve current sort
        header = self.table_preview.horizontalHeader()
        sort_col = header.sortIndicatorSection() if hasattr(header, 'sortIndicatorSection') else -1
        sort_order = header.sortIndicatorOrder() if hasattr(header, 'sortIndicatorOrder') else QtCore.Qt.AscendingOrder
        
        # Populate preview table with row limit for performance
        self.table_preview.blockSignals(True)
        self.table_preview.setSortingEnabled(False)
        self.table_preview.clearContents()
        
        # Limit table rows for performance
        display_indices = self.filtered_indices[:self.max_table_rows]
        if len(self.filtered_indices) > self.max_table_rows:
            self.log(f"Showing first {self.max_table_rows} of {len(self.filtered_indices)} filtered rows")
        
        self.table_preview.setRowCount(len(display_indices))
        for r, idx in enumerate(display_indices):
            row = self.df.loc[idx]
            disp = str(row.get("img_path", row.get("filename", idx)))
            # INF/EXT values for list columns
            inf_val = str(row.get("review_label_inf", "")) if "review_label_inf" in self.df.columns else ""
            ext_val = str(row.get("review_label_ext", "")) if "review_label_ext" in self.df.columns else ""
            # Active column value for quick flag
            active_val = str(row.get(self.active_label_col, "")) if self.active_label_col in self.df.columns else ""
            label_flag = "1" if active_val else "0"  # for sorting
            self.table_preview.setItem(r, 0, QtWidgets.QTableWidgetItem(str(idx)))
            self.table_preview.setItem(r, 1, QtWidgets.QTableWidgetItem(label_flag))
            self.table_preview.setItem(r, 2, QtWidgets.QTableWidgetItem(disp))
            self.table_preview.setItem(r, 3, QtWidgets.QTableWidgetItem(inf_val))
            self.table_preview.setItem(r, 4, QtWidgets.QTableWidgetItem(ext_val))
        
        self.table_preview.setSortingEnabled(True)
        # Re-apply preserved sort if any
        if sort_col is not None and sort_col >= 0 and self.table_preview.rowCount() > 0:
            self.table_preview.sortByColumn(sort_col, sort_order)
        self.table_preview.blockSignals(False)
        
        # default-select the top-most row
        if self.filtered_indices and self.table_preview.rowCount() > 0:
            self.current_idx = 0
            self.table_preview.selectRow(0)
        
        # live stats (filtered + overall)
        total = len(self.df) if self.df is not None else 0
        overall_unlabeled = 0
        overall_labeled = 0
        if self.df is not None and self.active_label_col in self.df.columns:
            overall_unlabeled = int(((self.df[self.active_label_col].isna()) | (self.df[self.active_label_col] == "")).sum())
            overall_labeled = total - overall_unlabeled
        
        # filtered counts
        f_total = len(self.filtered_indices)
        f_labeled = 0
        f_unlabeled = 0
        if f_total > 0 and self.active_label_col in self.df.columns:
            sub = self.df.loc[self.filtered_indices, self.active_label_col]
            f_unlabeled = int(((sub.isna()) | (sub == "")).sum())
            f_labeled = f_total - f_unlabeled
        
        self.lbl_stats.setText(
            f"Filtered: {f_total} | Labeled: {f_labeled} | Unlabeled: {f_unlabeled}  ||  Overall: {total} (L:{overall_labeled} U:{overall_unlabeled})"
        )
        self.refresh_view()
        
        # Ensure current row is selected in the list for visibility
        # ensure selected row in table remains in sync
        try:
            if self.filtered_indices and 0 <= self.current_idx < len(self.filtered_indices):
                self.table_preview.blockSignals(True)
                self.table_preview.clearSelection()
                sel_idx = self.filtered_indices[self.current_idx]
                row_in_table = self._find_list_row_by_index(sel_idx)
                if row_in_table >= 0:
                    self.table_preview.selectRow(row_in_table)
                self.table_preview.blockSignals(False)
        except Exception:
            pass
        
        # Update summary after any filter change
        self.update_summary()
        
        # Final memory check after filtering
        self._proactive_memory_cleanup()

    def on_clear_sort(self) -> None:
        try:
            i = self.cmb_sort_col.findText("(no sort)")
            if i >= 0:
                self.cmb_sort_col.setCurrentIndex(i)
            self.chk_sort_desc.setChecked(False)
        except Exception:
            pass
        self.apply_filters()

    def update_summary(self) -> None:
        if self.df is None or self.df.empty:
            self.txt_summary.setPlainText("No data loaded.")
            return
        total = len(self.df)
        labeled = 0
        unlabeled = 0
        if self.active_label_col in self.df.columns:
            unlabeled = int(((self.df[self.active_label_col].isna()) | (self.df[self.active_label_col] == "")).sum())
            labeled = total - unlabeled
        prog_pct = (labeled / total * 100.0) if total else 0.0
        # Label distribution (top 10)
        label_dist = []
        if self.active_label_col in self.df.columns:
            try:
                vc = self.df[self.active_label_col].fillna("").replace("", "(empty)").value_counts()
                for k, v in vc.head(10).items():
                    label_dist.append(f"  - {k}: {v}")
            except Exception:
                pass
        # origin_class distribution (top 10)
        origin_dist = []
        if "origin_class" in self.df.columns:
            try:
                vc2 = self.df["origin_class"].astype(str).value_counts()
                for k, v in vc2.head(10).items():
                    origin_dist.append(f"  - {k}: {v}")
            except Exception:
                pass
        lines = [
            f"Total: {total}",
            f"Labeled: {labeled} | Unlabeled: {unlabeled} | Progress: {prog_pct:.1f}%",
            "",
            f"Active label: {self.active_label_col}",
            "Label distribution (top 10):",
            *label_dist,
            "",
            "origin_class distribution (top 10):",
            *origin_dist,
        ]
        self.txt_summary.setPlainText("\n".join(lines))

    def log(self, message: str) -> None:
        try:
            ts = datetime.now().strftime("%H:%M:%S")
            if hasattr(self, 'txt_log') and self.txt_log is not None:
                self.txt_log.appendPlainText(f"[{ts}] {message}")
        except Exception:
            pass

    # -------- Batched JSON save helpers --------
    def _queue_update(self, row_idx: int, updater: Dict[str, object]) -> None:
        json_path = self.json_path or default_json_path(self.output_excel_path or self.excel_path)
        self._pending_json_path = json_path
        self._pending_ops.append(("meta", row_idx, updater, {}))
        self._save_timer.start()

    def _queue_set_values(self, row_idx: int, values: Dict[str, str], keys_for_row: Dict[str, str]) -> None:
        json_path = self.json_path or default_json_path(self.output_excel_path or self.excel_path)
        self._pending_json_path = json_path
        self._pending_ops.append(("values", row_idx, values, keys_for_row))
        self._save_timer.start()

    def _flush_pending_ops(self) -> None:
        if not self._pending_ops:
            return
        json_path = self._pending_json_path or (self.json_path or default_json_path(self.output_excel_path or self.excel_path))
        store = load_label_store(json_path)
        for kind, row_idx, payload, keys_for_row in self._pending_ops:
            key = str(row_idx)
            entry = store["labels"].get(key) or {}
            # Ensure identity keys present
            for k, v in keys_for_row.items():
                entry[k] = v
            if kind == "values":
                vals = entry.get("values") or {}
                for k, v in payload.items():
                    vals[k] = v
                entry["values"] = vals
            else:
                for k, v in payload.items():
                    entry[k] = v
            store["labels"][key] = entry
        ok = save_label_store(json_path, store)
        self._pending_ops.clear()
        if ok:
            self.status.showMessage("Saved JSON")
        else:
            self.status.showMessage("Save JSON failed")

    def reset_filters(self) -> None:
        if self.df is None:
            return
        self.edt_text.clear()
        self.chk_unlabeled.setChecked(False)
        self.cmb_origin.setCurrentIndex(0)
        if hasattr(self, 'cmb_label_value'):
            self.cmb_label_value.setCurrentIndex(0)
        self.cmb_sort_col.setCurrentIndex(0 if self.cmb_sort_col.count() > 0 else -1)
        self.chk_sort_desc.setChecked(False)
        # default to Unlabeled for active label column
        try:
            i = self.cmb_label_state.findText("Unlabeled")
            if i >= 0:
                self.cmb_label_state.setCurrentIndex(i)
        except Exception:
            pass
        # re-apply to rebuild list and select top
        self.apply_filters()

    def on_table_select(self) -> None:
        rows = self.table_preview.selectionModel().selectedRows()
        if not rows:
            return
        row = rows[0].row()
        try:
            idx_item = self.table_preview.item(row, 0)
            if not idx_item:
                return
            df_idx = int(idx_item.text())
            if df_idx in self.filtered_indices:
                self.current_idx = self.filtered_indices.index(df_idx)
                self.refresh_view()
        except Exception:
            pass

    def on_toggle_bookmark(self) -> None:
        if self.df is None or not self.filtered_indices:
            return
        row_idx = self.filtered_indices[self.current_idx]
        entry = get_json_entry(self.json_path or default_json_path(self.output_excel_path or self.excel_path), row_idx)
        curr = bool(entry.get("bookmark", False))
        # Queue bookmark toggle
        self._queue_update(row_idx, {"bookmark": not curr})
        self.status.showMessage("Bookmark " + ("ON" if not curr else "OFF"))
        self.log(f"Bookmark {'ON' if not curr else 'OFF'} for row {row_idx}")

    def on_save_memo(self) -> None:
        if self.df is None or not self.filtered_indices:
            return
        row_idx = self.filtered_indices[self.current_idx]
        memo = self.edt_memo.toPlainText()
        # Queue memo save
        self._queue_update(row_idx, {"memo": memo})
        self.status.showMessage("Memo queued")
        self.log(f"Memo saved for row {row_idx} ({len(memo)} chars)")

    def _resolve_img_for_row(self, row_idx: int) -> Tuple[Optional[str], Optional[str], Optional[str], str]:
        if self.df is None or self.images_base == "":
            return None, None, None, ""
        r = self.df.loc[row_idx]
        p = str(r.get("img_path", "")) or str(r.get("filename", ""))
        resolved_infer = resolve_image_path(self.images_base, p)
        # Resolve original using same relative path or basename match
        resolved_orig = None
        if self.images_base_orig:
            rel = normalize_relative_path(p)
            cand = os.path.join(self.images_base_orig, rel)
            if os.path.exists(cand):
                resolved_orig = cand
            else:
                base = os.path.basename(rel)
                base_no_ext, _ = os.path.splitext(base)
                for pattern in [
                    os.path.join(self.images_base_orig, "**", base),
                    os.path.join(self.images_base_orig, "**", f"{base_no_ext}.*"),
                ]:
                    m = glob.glob(pattern, recursive=True)
                    if m:
                        resolved_orig = m[0]
                        break
        # Resolve extra similarly
        resolved_extra = None
        if self.images_base_extra:
            rel = normalize_relative_path(p)
            cand = os.path.join(self.images_base_extra, rel)
            if os.path.exists(cand):
                resolved_extra = cand
            else:
                base = os.path.basename(rel)
                base_no_ext, _ = os.path.splitext(base)
                for pattern in [
                    os.path.join(self.images_base_extra, "**", base),
                    os.path.join(self.images_base_extra, "**", f"{base_no_ext}.*"),
                    os.path.join(self.images_base_extra, "**", f"*{base_no_ext}*.*"),
                ]:
                    m = glob.glob(pattern, recursive=True)
                    if m:
                        resolved_extra = m[0]
                        break
        return resolved_infer, resolved_orig, resolved_extra, p

    def refresh_view(self) -> None:
        if self.df is None or not self.filtered_indices:
            self.image_label_infer.setPixmap(QtGui.QPixmap())
            self.image_label_orig.setPixmap(QtGui.QPixmap())
            self.lbl_info.setText("Open Excel/CSV and set Images Bases.")
            return
        
        # Check memory before refreshing view
        self._manage_memory()
        
        row_idx = self.filtered_indices[self.current_idx]
        resolved_infer, resolved_orig, resolved_extra, disp = self._resolve_img_for_row(row_idx)
        self._set_image_on_label(self.image_label_infer, self.scroll_infer, resolved_infer)
        self._set_image_on_label(self.image_label_orig, self.scroll_orig, resolved_orig)
        self._set_image_on_label(self.image_label_extra, self.scroll_extra, resolved_extra)
        # Show paths under each image
        try:
            self.path_label_infer.setText(resolved_infer or "-")
            self.path_label_orig.setText(resolved_orig or "-")
            self.path_label_extra.setText(resolved_extra or "-")
        except Exception:
            pass
        inf_txt = "OK" if resolved_infer else "not found"
        org_txt = "OK" if resolved_orig else "not found"
        ext_txt = "OK" if resolved_extra else ("not found" if self.images_base_extra else "not set")
        self.lbl_info.setText(
            f"Row {self.current_idx+1}/{len(self.filtered_indices)}  |  INF: {inf_txt}  |  ORG: {org_txt}  |  EXT: {ext_txt}\n{disp}"
        )
        # Also show on status bar permanently
        if hasattr(self, 'lbl_status_io'):
            self.lbl_status_io.setText(f"INF: {inf_txt}   ORG: {org_txt}   EXT: {ext_txt}")
        # Load memo for current row
        json_path = self.json_path or default_json_path(self.output_excel_path or self.excel_path)
        entry = get_json_entry(json_path, row_idx)
        self.edt_memo.blockSignals(True)
        self.edt_memo.setPlainText(str(entry.get("memo", "")))
        self.edt_memo.blockSignals(False)
        # Update banner style (label state + bookmark)
        label_val = str(self.df.loc[row_idx].get(self.active_label_col, "")) if (self.df is not None and self.active_label_col in self.df.columns) else ""
        bookmarked = bool(entry.get("bookmark", False))
        if label_val:
            lv = str(label_val).strip().upper()
            if lv.startswith("NG"):
                color = "#c62828"  # red for NG
            else:
                color = "#2e7d32"  # green for OK or others
            text = f"Labeled: {label_val}"
        else:
            color = "#c62828"  # red
            text = "Unlabeled"
        if bookmarked:
            text = "★ " + text
        self.lbl_banner.setText(text)
        self.lbl_banner.setStyleSheet(
            f"background:{color}22; color:{color}; border:2px solid {color}; font-size:24px; font-weight:800; padding:4px; border-radius:4px;"
        )
        # Refresh AS-IS / TO-BE panel
        try:
            self._refresh_as_is_tobe_panel()
        except Exception:
            pass

    def _set_image_on_label(self, label: QtWidgets.QLabel, scroll: QtWidgets.QScrollArea, path: Optional[str]) -> None:
        if not path or not os.path.exists(path):
            label.setPixmap(QtGui.QPixmap())
            return
        
        # Check cache first
        if path in self._image_cache:
            pixmap = self._image_cache[path]
        else:
            # Load image with memory management
            try:
                # Check memory usage before loading
                if check_memory_limit(self.max_memory_mb):
                    # Clear old cache entries if memory is high
                    self._clear_image_cache()
                
                pixmap = QtGui.QPixmap(path)
                if not pixmap.isNull():
                    # Add to cache (limit cache size)
                    if len(self._image_cache) >= self.image_cache_size:
                        # Remove oldest entry
                        oldest_key = next(iter(self._image_cache))
                        del self._image_cache[oldest_key]
                    self._image_cache[path] = pixmap
                else:
                    label.setPixmap(QtGui.QPixmap())
                    return
            except Exception as e:
                self.log(f"Error loading image {path}: {str(e)}")
                label.setPixmap(QtGui.QPixmap())
                return
        
        if not getattr(self, 'fit_to_window', True):
            label.setPixmap(pixmap)
            return
        
        vp_size = scroll.viewport().size()
        if vp_size.width() <= 0 or vp_size.height() <= 0:
            label.setPixmap(pixmap)
            return
        
        # Scale image to fit viewport
        scaled = pixmap.scaled(vp_size, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
        label.setPixmap(scaled)
    
    def _clear_image_cache(self):
        """Clear image cache to free memory"""
        self._image_cache.clear()
        force_garbage_collection()
        self.log("Image cache cleared to free memory")
    
    def _manage_memory(self):
        """Proactive memory management"""
        memory_usage = get_memory_usage()
        if memory_usage > self.max_memory_mb * 0.8:  # Reduced from 0.9 to 0.8 for earlier intervention
            self._clear_image_cache()
            force_garbage_collection()
            self.log(f"Memory management triggered: {memory_usage:.1f}MB")
            return True
        return False

    def _proactive_memory_cleanup(self):
        """More aggressive memory cleanup"""
        memory_usage = get_memory_usage()
        if memory_usage > self.max_memory_mb * 0.7:  # Even earlier intervention
            self.log(f"Proactive memory cleanup: {memory_usage:.1f}MB")
            # Clear image cache
            self._image_cache.clear()
            # Force garbage collection multiple times
            for _ in range(3):
                force_garbage_collection()
            # Clear any temporary variables
            if hasattr(self, '_temp_data'):
                del self._temp_data
            return True
        return False

    def _show_memory_info(self):
        """Show current memory usage information"""
        memory_usage = get_memory_usage()
        cache_size = len(self._image_cache)
        df_size = len(self.df) if self.df is not None else 0
        
        info = f"""Memory Usage: {memory_usage:.1f}MB
Image Cache: {cache_size} images
DataFrame Rows: {df_size}
Filtered Rows: {len(self.filtered_indices)}
Memory Limit: {self.max_memory_mb}MB"""
        
        QtWidgets.QMessageBox.information(self, "Memory Information", info)
    
    def _load_more_data(self):
        """Load additional data in chunks"""
        if not self.excel_path:
            QtWidgets.QMessageBox.information(self, "Load More", "No file loaded.")
            return
        
        try:
            current_rows = len(self.df) if self.df is not None else 0
            
            if self.excel_path.lower().endswith(".csv"):
                # Load next chunk of CSV
                chunk_df = pd.read_csv(self.excel_path, encoding="utf-8-sig", 
                                     skiprows=range(1, current_rows + 1), 
                                     nrows=self.chunk_size)
                if chunk_df.empty:
                    QtWidgets.QMessageBox.information(self, "Load More", "No more data to load.")
                    return
                
                # Append to existing DataFrame
                self.df = pd.concat([self.df, chunk_df], ignore_index=True)
                
            else:
                # For Excel files, load next chunk
                chunk_df = pd.read_excel(self.excel_path, sheet_name=self.sheet_name,
                                       skiprows=range(1, current_rows + 1),
                                       nrows=self.chunk_size)
                if chunk_df.empty:
                    QtWidgets.QMessageBox.information(self, "Load More", "No more data to load.")
                    return
                
                # Append to existing DataFrame
                self.df = pd.concat([self.df, chunk_df], ignore_index=True)
            
            # Update UI
            self.filtered_indices = list(self.df.index)
            self.populate_filter_controls()
            self.apply_filters()
            
            memory_usage = get_memory_usage()
            self.log(f"Loaded {len(chunk_df)} more rows. Total: {len(self.df)}. Memory: {memory_usage:.1f}MB")
            
            if memory_usage > self.max_memory_mb * 0.8:
                QtWidgets.QMessageBox.warning(self, "Memory Warning", 
                    f"Memory usage is high ({memory_usage:.1f}MB). Consider clearing cache.")
                    
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Load More Failed", str(e))
            self.log(f"Error loading more data: {str(e)}")

    def _force_memory_cleanup(self):
        """Force immediate memory cleanup"""
        self.log("Forcing memory cleanup...")
        # Clear image cache
        self._image_cache.clear()
        # Force garbage collection multiple times
        for i in range(5):
            force_garbage_collection()
        # Clear any temporary data
        if hasattr(self, '_temp_data'):
            del self._temp_data
        memory_usage = get_memory_usage()
        self.log(f"Memory cleanup completed. Current usage: {memory_usage:.1f}MB")
        QtWidgets.QMessageBox.information(self, "Memory Cleanup", 
            f"Memory cleanup completed.\nCurrent usage: {memory_usage:.1f}MB")

    def _optimize_memory_settings(self):
        """Optimize memory settings based on current system state"""
        system_memory = get_system_memory()
        current_usage = get_memory_usage()
        
        # Calculate optimal settings
        available_memory = system_memory - current_usage
        optimal_limit = min(1024, available_memory * 0.3)  # 30% of available memory
        
        # Update settings
        old_limit = self.max_memory_mb
        self.max_memory_mb = optimal_limit
        
        # Adjust other settings based on available memory
        if available_memory < 2048:  # Less than 2GB available
            self.chunk_size = 250
            self.max_table_rows = 1000
            self.image_cache_size = 3
        elif available_memory < 4096:  # Less than 4GB available
            self.chunk_size = 500
            self.max_table_rows = 2000
            self.image_cache_size = 5
        else:  # 4GB+ available
            self.chunk_size = 1000
            self.max_table_rows = 3000
            self.image_cache_size = 8
        
        info = f"""Memory settings optimized:
System Memory: {system_memory:.0f}MB
Current Usage: {current_usage:.1f}MB
Available Memory: {available_memory:.0f}MB

New Settings:
- Memory Limit: {self.max_memory_mb:.0f}MB (was {old_limit:.0f}MB)
- Chunk Size: {self.chunk_size}
- Max Table Rows: {self.max_table_rows}
- Image Cache Size: {self.image_cache_size}"""
        
        self.log("Memory settings optimized")
        QtWidgets.QMessageBox.information(self, "Memory Settings Optimized", info)

    def _select_all_pred_filters(self):
        for cb in self.pred_checkboxes.values():
            cb.setChecked(True)

    def _select_none_pred_filters(self):
        for cb in self.pred_checkboxes.values():
            cb.setChecked(False)


def main() -> None:
    parser = argparse.ArgumentParser(description="PySide6 Local Labeler")
    parser.add_argument("--file", dest="file", type=str, default="", help="Path to CSV or Excel file to open")
    parser.add_argument("--images", dest="images", type=str, default="", help="Path to inference/viz images base directory")
    parser.add_argument("--orig-images", dest="orig_images", type=str, default="", help="Path to original images base directory")
    parser.add_argument("--extra-images", dest="extra_images", type=str, default="", help="Path to extra images base directory")
    args, qt_args = parser.parse_known_args()

    app = QtWidgets.QApplication([sys.argv[0], *qt_args])
    w = LabelerWindow()

    # Apply CLI args
    try:
        if args.file and os.path.exists(args.file):
            w.load_excel_from_path(args.file)
        if args.images and os.path.isdir(args.images):
            w.images_base = args.images
        if args.orig_images and os.path.isdir(args.orig_images):
            w.images_base_orig = args.orig_images
        if args.extra_images and os.path.isdir(args.extra_images):
            w.images_base_extra = args.extra_images
        # Refresh view if any base set
        if args.images or args.orig_images or args.extra_images:
            w.refresh_view()
    except Exception:
        pass

    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()


