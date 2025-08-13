#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import glob
import csv
from openpyxl import Workbook
from openpyxl.drawing import image as openpyxl_image
from PIL import Image as PILImage
import uuid

def load_inference_results(base_path):
    """Load all inference_results.csv files and create a lookup dictionary"""
    print("=== 추론 결과 파일 검색 시작 ===")
    csv_files = glob.glob(os.path.join(base_path, "**", "inference_results.csv"), recursive=True)
    print(f"발견된 CSV 파일 수: {len(csv_files)}")
    
    results_dict = {}
    
    for i, csv_file in enumerate(csv_files):
        try:
            print(f"[{i+1}/{len(csv_files)}] CSV 파일 로딩 중: {csv_file}")
            row_count = 0
            with open(csv_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    filename = row.get('filename', '').replace('.bmp', '')
                    results_dict[filename] = {
                        'gt_status': row.get('gt_status', ''),
                        'pred_status': row.get('pred_status', ''),
                        'dominant_class': row.get('dominant_class', ''),
                        'csv_source': os.path.basename(os.path.dirname(csv_file))
                    }
                    row_count += 1
            print(f"   └── {row_count}개 결과 로딩 완료")
        except Exception as e:
            print(f"   └── 오류 발생: {e}")
    
    print(f"=== 총 {len(results_dict)}개 파일의 추론 결과 로딩 완료 ===\n")
    return results_dict

def find_image_pairs(base_path):
    """Find image pairs (viz.png and .bmp files) in the given path"""
    print("=== 이미지 파일 검색 시작 ===")
    print("PNG 파일 (viz) 검색 중...")
    png_files = glob.glob(os.path.join(base_path, "**", "images", "**", "*_viz.png"), recursive=True)
    print(f"발견된 PNG 파일 수: {len(png_files)}")
    
    print("BMP 파일 검색 중...")
    bmp_files = glob.glob(os.path.join(base_path, "**", "images", "**", "*.bmp"), recursive=True)
    print(f"발견된 BMP 파일 수: {len(bmp_files)}")
    
    print("이미지 쌍 매칭 중...")
    pairs = []
    matched_count = 0
    
    for i, png_file in enumerate(png_files):
        if i % 100 == 0:
            print(f"   진행률: {i}/{len(png_files)} PNG 파일 처리 중...")
        
        # Extract base name by removing "_viz.png"
        base_name = os.path.basename(png_file).replace("_viz.png", "")
        
        # Find corresponding .bmp file
        bmp_file = None
        for bmp in bmp_files:
            if os.path.basename(bmp).replace(".bmp", "") == base_name:
                bmp_file = bmp
                matched_count += 1
                break
        
        if bmp_file:
            pairs.append({
                'filename': base_name,
                'viz_img_path': png_file,
                'img_path': bmp_file
            })
    
    print(f"=== 총 {len(pairs)}개 이미지 쌍 매칭 완료 ===\n")
    return pairs

def create_excel_with_images_and_results(pairs, results_dict, output_file):
    """Create Excel file with images and inference results"""
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Image Pairs with Results"
        
        # Set headers
        ws['A1'] = 'filename'
        ws['B1'] = 'img'
        ws['C1'] = 'viz_img'
        ws['D1'] = 'gt_status'
        ws['E1'] = 'pred_status'
        ws['F1'] = 'dominant_class'
        ws['G1'] = 'csv_source'
        
        # Set column widths to fit images better (2.5 inches)
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 25  # Width for BMP image (2.5 inches)
        ws.column_dimensions['C'].width = 50  # Width for PNG image (2:1 ratio, 2.5 inches height)
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 30
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        current_row = 2
        temp_files = []  # Track temporary files for cleanup
        
        print("=== 엑셀 파일에 이미지 및 결과 삽입 시작 ===")
        total_pairs = len(pairs)
        
        for i, pair in enumerate(pairs):
                
            try:
                print(f"[{i+1}/{total_pairs}] 처리 중: {pair['filename']}")
                
                # Add filename
                ws[f'A{current_row}'] = pair['filename']
                print(f"   └── 파일명 추가: {pair['filename']}")
                
                # Add inference results if available
                if pair['filename'] in results_dict:
                    result = results_dict[pair['filename']]
                    ws[f'D{current_row}'] = result['gt_status']
                    ws[f'E{current_row}'] = result['pred_status']
                    ws[f'F{current_row}'] = result['dominant_class']
                    ws[f'G{current_row}'] = result['csv_source']
                    print(f"   └── 추론 결과 추가: GT={result['gt_status']}, Pred={result['pred_status']}, Class={result['dominant_class']}")
                else:
                    ws[f'D{current_row}'] = 'Not found'
                    ws[f'E{current_row}'] = 'Not found'
                    ws[f'F{current_row}'] = 'Not found'
                    ws[f'G{current_row}'] = 'Not found'
                    print(f"   └── 추론 결과 없음")
                
                # Set row height for images (2.5 inches = 180 points)
                ws.row_dimensions[current_row].height = 180
                
                # Add BMP image (img column)
                print(f"   └── BMP 이미지 처리 중: {pair['img_path']}")
                if os.path.exists(pair['img_path']):
                    try:
                        # Create unique temporary filename in current directory
                        temp_bmp_path = f"temp_bmp_{uuid.uuid4().hex[:8]}.png"
                        temp_files.append(temp_bmp_path)
                        
                        # Resize BMP image
                        with PILImage.open(pair['img_path']) as pil_img:
                            original_size = pil_img.size
                            # Convert to RGB if necessary
                            if pil_img.mode != 'RGB':
                                pil_img = pil_img.convert('RGB')
                            pil_img.thumbnail((300, 300), PILImage.Resampling.LANCZOS)
                            new_size = pil_img.size
                            pil_img.save(temp_bmp_path, "PNG")
                        
                        # Add to Excel with proper sizing (2.5 inches = 180 pixels)
                        img = openpyxl_image.Image(temp_bmp_path)
                        img.width = 180  # Set width to 180 pixels (2.5 inches)
                        img.height = 180  # Set height to 180 pixels (2.5 inches)
                        img.anchor = f'B{current_row}'
                        ws.add_image(img)
                        print(f"       └── BMP 이미지 삽입 완료 ({original_size[0]}x{original_size[1]} → {new_size[0]}x{new_size[1]})")
                        
                    except Exception as e:
                        ws[f'B{current_row}'] = f"Error: {str(e)}"
                        print(f"       └── BMP 이미지 처리 오류: {e}")
                else:
                    ws[f'B{current_row}'] = "BMP not found"
                    print(f"       └── BMP 파일 없음")
                
                # Add PNG image (viz_img column)
                print(f"   └── PNG 이미지 처리 중: {pair['viz_img_path']}")
                if os.path.exists(pair['viz_img_path']):
                    try:
                        # Create unique temporary filename in current directory
                        temp_png_path = f"temp_png_{uuid.uuid4().hex[:8]}.png"
                        temp_files.append(temp_png_path)
                        
                        # Resize PNG image (keep 2:1 aspect ratio)
                        with PILImage.open(pair['viz_img_path']) as pil_img:
                            original_size = pil_img.size
                            # Convert to RGB if necessary
                            if pil_img.mode != 'RGB':
                                pil_img = pil_img.convert('RGB')
                            # For viz images, maintain 2:1 aspect ratio (600x300)
                            pil_img.thumbnail((600, 300), PILImage.Resampling.LANCZOS)
                            new_size = pil_img.size
                            pil_img.save(temp_png_path, "PNG")
                        
                        # Add to Excel with proper sizing (2:1 aspect ratio, 2.5 inches height)
                        img = openpyxl_image.Image(temp_png_path)
                        img.width = 360  # Set width to 360 pixels (2:1 ratio, 2.5 inches height)
                        img.height = 180  # Set height to 180 pixels (2.5 inches)
                        img.anchor = f'C{current_row}'
                        ws.add_image(img)
                        print(f"       └── PNG 이미지 삽입 완료 ({original_size[0]}x{original_size[1]} → {new_size[0]}x{new_size[1]})")
                        
                    except Exception as e:
                        ws[f'C{current_row}'] = f"Error: {str(e)}"
                        print(f"       └── PNG 이미지 처리 오류: {e}")
                else:
                    ws[f'C{current_row}'] = "PNG not found"
                    print(f"       └── PNG 파일 없음")
                
                current_row += 1
                print(f"   └── 행 {current_row-1} 처리 완료\n")
                
            except Exception as e:
                print(f"   └── 쌍 처리 오류: {e}\n")
                continue
        
        # Save workbook with proper encoding
        print("엑셀 파일 저장 중...")
        wb.save(output_file)
        
        # Clean up temporary files
        print("임시 파일 정리 중...")
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except Exception as e:
                print(f"Error cleaning up {temp_file}: {e}")
        
        print(f"엑셀 파일 생성 완료: {output_file}")
        print(f"처리된 총 쌍의 수: {len(pairs)}")
        
        return True
        
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        # Clean up temporary files even on error
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except:
                pass
        return False

def main():
    base_path = "/Users/rtm/Downloads/v0.3"
    output_file = "/Users/rtm/Downloads/v0.3/image_pairs_with_results.xlsx"
    
    print("추론 결과 로딩 중...")
    results_dict = load_inference_results(base_path)
    
    print("이미지 쌍 검색 중...")
    pairs = find_image_pairs(base_path)
    
    print(f"발견된 이미지 쌍: {len(pairs)}개")
    
    if pairs:
        print("이미지와 결과가 포함된 엑셀 파일 생성 중...")
        success = create_excel_with_images_and_results(pairs, results_dict, output_file)
        
        if success:
            print(f"엑셀 파일 생성 완료: {output_file}")
        else:
            print("엑셀 파일 생성 실패")
    else:
        print("이미지 쌍을 찾을 수 없습니다")

if __name__ == "__main__":
    main()